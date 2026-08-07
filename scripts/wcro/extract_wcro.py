"""
extract_wcro.py
---------------
Everde — WCRO (West Coast Retail Opportunity) portal extractor.

Reads published report workbooks only (no engine recompute). Sprint 1:
validate Four Numbers against the known 5.29 / 2026-08-06 targets, then
publish wcro_data.json for the portal.

Usage:
  python extract_wcro.py
  python extract_wcro.py --reports "\\\\192.168.190.10\\Claude Sandbox\\DataDrops\\_HANDOFF_WCRO_2026-08-06\\reports"
  python extract_wcro.py --out data/wcro_data.json --skip-validate   # escape hatch only

Dependencies: pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# ── Scope constants (engine rules — enforce / document, do not recompute) ──

WCRO_ORGS = {"WIN", "BRA", "STE", "FAL", "MIR", "PAU", "HUN", "PIR", "ESC", "FOR"}
TX_FL_ORGS = {"GFL", "MCR", "BNL", "OAS", "HOM"}
AB_GRADES = {"A", "B"}
QC_GRADES = {"SS", "SN", "S2N", "GS", "GN", "S"}

DEFAULT_HANDOFF = Path(
    r"\\192.168.190.10\Claude Sandbox\DataDrops\_HANDOFF_WCRO_2026-08-06"
)
DEFAULT_WEEKLYDROP = DEFAULT_HANDOFF / "WeeklyDrop"
DEFAULT_REPORTS = DEFAULT_HANDOFF / "reports"

SET_FOLDER_NAMES = {
    "set1": "Sales Variance and Allocation",
    "set2": "Store Driven Sales Recommendation",
    "set3": "On Hand and Register Sales Analysis",
    "set4": "Transfers",
    "set5": "Rep Orders",
}

# §2 / handoff validation targets (5.29 / 2026-08-06 only)
TARGETS = {
    "ship_this_week": 1_873_239,
    "to_transfer": 585_910,
    "nn_plan": 998_309,
    # Demand-sensed *pool-level* units (memory: NN_Customer units pool).
    # Combined Summary labels this "NN Cust Pool (units)" (~145k); gross
    # "NN Cust Store (units)" is ~380k and must NOT be used for this tile.
    "nn_cust_store": 150_293,
}
BASELINE_REFRESH = ("5.29", "2026-08-06")

REFRESH_RE = re.compile(
    r"Refresh\s+(\d+\.\d+)\s*[-–—]\s*(\d{4}-\d{2}-\d{2})", re.I
)


def safe_read(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        return None
    return value


def norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def as_float(value: Any) -> float | None:
    v = safe_read(value)
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def as_int_round(value: Any) -> int | None:
    f = as_float(value)
    return None if f is None else int(round(f))


def parse_refresh_from_name(name: str) -> tuple[str | None, str | None]:
    m = REFRESH_RE.search(name)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def find_header_row(rows: list[tuple], required_substr: str, max_scan: int = 40) -> int | None:
    needle = required_substr.lower()
    for i, row in enumerate(rows[:max_scan]):
        for cell in row:
            if isinstance(cell, str) and needle in norm_header(cell).lower():
                return i
    return None


def col_map(header_row: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = norm_header(cell)
        if key and key not in out:
            out[key] = i
    return out


def find_col(cmap: dict[str, int], *candidates: str) -> int | None:
    lower = {k.lower(): i for k, i in cmap.items()}
    for cand in candidates:
        c = cand.lower()
        if c in lower:
            return lower[c]
        for k, i in lower.items():
            if c in k:
                return i
    return None


def load_sheet_rows(path: Path, sheet: str | None = None) -> tuple[str, list[tuple]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        name = sheet if sheet and sheet in wb.sheetnames else wb.sheetnames[0]
        ws = wb[name]
        rows = [tuple(safe_read(c) for c in row) for row in ws.iter_rows(values_only=True)]
        return name, rows
    finally:
        wb.close()


def list_xlsx(folder: Path) -> list[Path]:
    if not folder.is_dir():
        return []
    return sorted(
        p
        for p in folder.glob("*.xlsx")
        if not p.name.startswith("~$") and "Archive" not in p.name
    )


def list_xlsx_recursive(folder: Path) -> list[Path]:
    if not folder.is_dir():
        return []
    return sorted(
        p
        for p in folder.rglob("*.xlsx")
        if not p.name.startswith("~$") and "Archive" not in p.name
    )


def classify_flat_file(name: str) -> str | None:
    """Map a flat WeeklyDrop filename to set1..set5."""
    n = name.lower()
    if " - orders (" in n or n.endswith(" - orders.xlsx") or " orders (refresh" in n:
        return "set5"
    if "on hand and register" in n:
        return "set3"
    if "transfer" in n and "store driven" not in n:
        return "set4"
    if "store driven" in n or "combined summary" in n:
        return "set2"
    if (
        "sales variance" in n
        or "sales manager summary" in n
        or "actuals vs suggested" in n
    ):
        return "set1"
    return None


def materialize_flat_sets(root: Path, staging: Path) -> dict[str, Path]:
    """Copy/classify flat xlsx into set subfolders under staging."""
    import shutil

    staging.mkdir(parents=True, exist_ok=True)
    sets: dict[str, Path] = {}
    for key, folder_name in SET_FOLDER_NAMES.items():
        d = staging / folder_name
        d.mkdir(parents=True, exist_ok=True)
        # clear prior staged copies so deleted drops don't linger
        for old in list_xlsx(d):
            try:
                old.unlink()
            except OSError:
                pass
        sets[key] = d

    for p in list_xlsx(root):
        key = classify_flat_file(p.name)
        if not key:
            print(f"WARN: unclassified WeeklyDrop file skipped: {p.name}", file=sys.stderr)
            continue
        dest = sets[key] / p.name
        shutil.copy2(p, dest)
    return sets


def resolve_set_dirs(reports_root: Path) -> tuple[Path, dict[str, Path]]:
    """
    Resolve the five set folders from either:
      reports_root/<Set Name>/*.xlsx   (canonical)
      reports_root/*.xlsx              (flat WeeklyDrop — staged under .wcro_extract_sets)
    """
    canonical = {
        k: reports_root / name for k, name in SET_FOLDER_NAMES.items()
    }
    if all(d.is_dir() and list_xlsx(d) for d in [canonical["set2"], canonical["set5"]]):
        # Prefer canonical when Set 2 + Rep Orders exist (minimum for Four Numbers + index)
        if all(d.is_dir() for d in canonical.values()):
            return reports_root, canonical

    flat = list_xlsx(reports_root)
    if flat:
        staging = reports_root / ".wcro_extract_sets"
        sets = materialize_flat_sets(reports_root, staging)
        if not list_xlsx(sets["set2"]):
            raise FileNotFoundError(
                f"WeeklyDrop has xlsx but no Store Driven / Combined Summary files in {reports_root}"
            )
        return staging, sets

    # Partial canonical (some folders only)
    if any(d.is_dir() for d in canonical.values()):
        missing = [k for k, d in canonical.items() if not d.is_dir()]
        if missing:
            raise FileNotFoundError(
                f"Incomplete WCRO layout under {reports_root}; missing: "
                + ", ".join(SET_FOLDER_NAMES[m] for m in missing)
            )
        return reports_root, canonical

    raise FileNotFoundError(
        f"No WCRO workbooks found under {reports_root}. "
        "Drop the five set folders (or flat published xlsx) into WeeklyDrop."
    )


def pick_reports_root(weeklydrop: Path | None, reports: Path | None) -> Path:
    """Prefer WeeklyDrop when it has content; else handoff reports\\; else explicit --reports."""
    wd = weeklydrop or DEFAULT_WEEKLYDROP
    rp = reports or DEFAULT_REPORTS

    def has_content(p: Path) -> bool:
        if not p.is_dir():
            return False
        if list_xlsx(p):
            return True
        return any(
            (p / name).is_dir() and list_xlsx(p / name)
            for name in SET_FOLDER_NAMES.values()
        )

    if has_content(wd):
        return wd
    if reports is not None and has_content(rp):
        return rp
    if has_content(rp):
        print(
            f"WeeklyDrop empty — using handoff reports folder: {rp}",
            file=sys.stderr,
        )
        return rp
    if weeklydrop is not None:
        return wd
    return rp


# ── Combined Summary → Four Numbers ───────────────────────────────────────


def extract_combined_summary(path: Path) -> dict[str, Any]:
    _, rows = load_sheet_rows(path)
    hi = find_header_row(rows, "Segment")
    if hi is None:
        raise RuntimeError(f"Combined Summary: no Segment header in {path.name}")
    cmap = col_map(rows[hi])

    i_seg = find_col(cmap, "Segment")
    i_plan_var = find_col(cmap, "Plan Var (net) whlsl $", "Plan Var (net)")
    i_nn_plan_u = find_col(cmap, "NN Plan (units)")
    i_nn_plan_d = find_col(cmap, "NN Plan whlsl $")
    i_nn_cust_u = find_col(cmap, "NN Cust Store (units)")
    i_nn_cust_d = find_col(cmap, "NN Cust Store whlsl $")
    i_nn_pool_u = find_col(cmap, "NN Cust Pool (units)")
    i_nn_pool_d = find_col(cmap, "NN Cust Pool whlsl $")
    i_ship_u = find_col(cmap, "Ship This Week (units)")
    i_ship_d = find_col(cmap, "Ship This Week whlsl $")
    i_order_d = find_col(cmap, "Order $")
    i_for_d = find_col(cmap, "FOR direct $")
    i_xfer_u = find_col(cmap, "To Transfer (units)")
    i_xfer_d = find_col(cmap, "To Transfer whlsl $")
    i_ab_d = find_col(cmap, "A+B on hand whlsl $")

    required = {
        "Segment": i_seg,
        "NN Plan (units)": i_nn_plan_u,
        "NN Cust Pool (units)": i_nn_pool_u,
        "Ship This Week whlsl $": i_ship_d,
        "To Transfer whlsl $": i_xfer_d,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise RuntimeError(f"Combined Summary missing columns: {missing}")

    segments: list[dict[str, Any]] = []
    combined: dict[str, Any] | None = None

    for row in rows[hi + 1 :]:
        if not row or i_seg is None:
            break
        seg = row[i_seg]
        if seg is None:
            # blank row ends the segment table
            if segments:
                break
            continue
        seg_s = str(seg).strip()
        if not seg_s or seg_s.upper().startswith("PLAN VARIANCE"):
            break

        def g(idx: int | None) -> float | None:
            return as_float(row[idx]) if idx is not None and idx < len(row) else None

        rec = {
            "segment": seg_s,
            "plan_var_net_$": g(i_plan_var),
            "nn_plan_u": g(i_nn_plan_u),
            "nn_plan_$": g(i_nn_plan_d),
            "nn_cust_store_u": g(i_nn_cust_u),
            "nn_cust_store_$": g(i_nn_cust_d),
            "nn_cust_pool_u": g(i_nn_pool_u),
            "nn_cust_pool_$": g(i_nn_pool_d),
            "ship_this_week_u": g(i_ship_u),
            "ship_this_week_$": g(i_ship_d),
            "order_$": g(i_order_d),
            "for_direct_$": g(i_for_d),
            "to_transfer_u": g(i_xfer_u),
            "to_transfer_$": g(i_xfer_d),
            "ab_on_hand_$": g(i_ab_d),
        }
        segments.append(rec)
        if seg_s.lower().startswith("combined"):
            combined = rec

    if combined is None:
        raise RuntimeError("Combined Summary: Combined Total row not found")

    # Plan variance summary section (optional)
    plan_var_summary: list[dict[str, Any]] = []
    pvi = find_header_row(rows, "Sales Plan Variance")
    if pvi is not None:
        pcm = col_map(rows[pvi])
        ps = find_col(pcm, "Segment")
        pn = find_col(pcm, "Sales Plan Variance (net) $", "Sales Plan Variance")
        pa = find_col(pcm, "Ahead of Plan $", "Ahead of Plan")
        pb = find_col(pcm, "Behind Plan (gross) $", "Behind Plan")
        for row in rows[pvi + 1 :]:
            if not row or ps is None or row[ps] is None:
                if plan_var_summary:
                    break
                continue
            seg_s = str(row[ps]).strip()
            if not seg_s or "RETAIL" in seg_s.upper():
                break
            plan_var_summary.append(
                {
                    "segment": seg_s,
                    "net_$": as_float(row[pn]) if pn is not None else None,
                    "ahead_$": as_float(row[pa]) if pa is not None else None,
                    "behind_gross_$": as_float(row[pb]) if pb is not None else None,
                }
            )

    refresh, date = parse_refresh_from_name(path.name)
    # Prefer banner line if present
    for row in rows[:5]:
        for cell in row:
            if isinstance(cell, str) and "Refresh" in cell:
                m = re.search(
                    r"Refresh\s+(\d+\.\d+).*?(\d{4}-\d{2}-\d{2})", cell, re.I
                )
                if m:
                    refresh, date = m.group(1), m.group(2)

    four = {
        "ship_this_week": as_int_round(combined["ship_this_week_$"]),
        "to_transfer": as_int_round(combined["to_transfer_$"]),
        "nn_plan": as_int_round(combined["nn_plan_u"]),
        # Pool-level demand-sensed units → Four Numbers "NN Cust Store" tile
        "nn_cust_store": as_int_round(combined["nn_cust_pool_u"]),
        "nn_cust_store_gross_u": as_int_round(combined["nn_cust_store_u"]),
        "nn_cust_pool_u": as_int_round(combined["nn_cust_pool_u"]),
        "note": (
            "NN Plan is plan-driven. Four Numbers NN Cust Store uses pool-level "
            "demand-sensed units (Combined Summary 'NN Cust Pool'). Gross "
            "NN Cust Store units are larger and measure maldistribution vs pool."
        ),
    }

    return {
        "source_file": path.name,
        "unc_path": str(path),
        "refresh": refresh,
        "date": date,
        "four_numbers": four,
        "segments": segments,
        "plan_var_summary": plan_var_summary,
        "plan_var_net_combined_summary_$": as_int_round(combined["plan_var_net_$"]),
        "notes": [
            "A+B on hand $ is shared physical stock — never sum across HD/LOW segments.",
            "Sales Manager Summary plan variance NET may differ cosmetically "
            "from Combined Summary (known label/cutoff display gap).",
        ],
    }


# ── Build Health ──────────────────────────────────────────────────────────


def extract_build_health(path: Path, channel: str) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Build Health" not in wb.sheetnames:
            return {
                "channel": channel,
                "file": path.name,
                "status": "MISSING",
                "checks": [],
            }
        ws = wb["Build Health"]
        status = "UNKNOWN"
        checks: list[dict[str, str]] = []
        for row in ws.iter_rows(values_only=True):
            vals = [safe_read(c) for c in row]
            if not vals or vals[0] is None:
                continue
            first = str(vals[0])
            if first.upper().startswith("BUILD STATUS"):
                m = re.search(r"(PASS|FAIL)", first, re.I)
                status = m.group(1).upper() if m else first.split(":", 1)[-1].strip().upper()
                continue
            if first == "Check":
                continue
            detail = vals[1] if len(vals) > 1 else None
            result = vals[2] if len(vals) > 2 else None
            if result is None and len(vals) > 1:
                for v in reversed(vals):
                    if isinstance(v, str) and v.upper() in {"PASS", "FAIL"}:
                        result = v
                        break
            if isinstance(result, str) and result.upper() in {"PASS", "FAIL"}:
                checks.append(
                    {
                        "check": first,
                        "detail": "" if detail is None else str(detail),
                        "result": result.upper(),
                    }
                )
        return {
            "channel": channel,
            "file": path.name,
            "status": status,
            "checks": checks,
        }
    finally:
        wb.close()


# ── Transfers ─────────────────────────────────────────────────────────────


def extract_transfers(path: Path, channel: str) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        tabs: dict[str, Any] = {}
        total_u = 0.0
        total_d = 0.0
        for name in wb.sheetnames:
            if name.lower() in {"read me", "readme", "changes"}:
                continue
            if not name.upper().startswith("TO "):
                continue
            ws = wb[name]
            rows = [tuple(safe_read(c) for c in r) for r in ws.iter_rows(values_only=True)]
            hi = find_header_row(rows, "Transfer Qty")
            if hi is None:
                hi = find_header_row(rows, "From")
            if hi is None:
                continue
            cmap = col_map(rows[hi])
            i_from = find_col(cmap, "From")
            i_items = find_col(cmap, "Items in Group")
            i_desc = find_col(cmap, "Description")
            i_genus = find_col(cmap, "Genus")
            i_form = find_col(cmap, "Form")
            i_size = find_col(cmap, "Size")
            i_qty = find_col(cmap, "Transfer Qty (u)", "Transfer Qty")
            i_dol = find_col(cmap, "Wholesale $")
            i_dest = find_col(cmap, "Dest. Stores", "Dest Stores")
            lines: list[dict[str, Any]] = []
            su = sd = 0.0
            for row in rows[hi + 1 :]:
                if not row or (i_from is not None and row[i_from] is None and row[0] is None):
                    continue
                if i_qty is not None and i_qty < len(row):
                    q = as_float(row[i_qty]) or 0.0
                else:
                    q = 0.0
                if i_dol is not None and i_dol < len(row):
                    d = as_float(row[i_dol]) or 0.0
                else:
                    d = 0.0
                if q == 0 and d == 0:
                    # skip empty
                    fr = row[i_from] if i_from is not None else row[0]
                    if fr is None:
                        continue
                su += q
                sd += d
                lines.append(
                    {
                        "from": row[i_from] if i_from is not None else None,
                        "items_in_group": row[i_items] if i_items is not None else None,
                        "description": row[i_desc] if i_desc is not None else None,
                        "genus": row[i_genus] if i_genus is not None else None,
                        "form": row[i_form] if i_form is not None else None,
                        "size": row[i_size] if i_size is not None else None,
                        "transfer_qty_u": q,
                        "wholesale_$": d,
                        "dest_stores": row[i_dest] if i_dest is not None else None,
                    }
                )
            tabs[name] = {
                "line_count": len(lines),
                "transfer_qty_u": round(su, 2),
                "wholesale_$": round(sd, 2),
                "lines": lines,
            }
            total_u += su
            total_d += sd
        return {
            "channel": channel,
            "file": path.name,
            "unc_path": str(path),
            "total_transfer_u": round(total_u, 2),
            "total_transfer_$": round(total_d, 2),
            "tabs": tabs,
        }
    finally:
        wb.close()


# ── Store Driven compact extract ──────────────────────────────────────────


def extract_store_driven(path: Path, channel: str) -> dict[str, Any]:
    """Totals from By-Pool tabs + Build Health. Skip Oracle Order/FOR tabs."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        markets: dict[str, Any] = {}
        for sheet in wb.sheetnames:
            if not sheet.startswith("By-Pool "):
                continue
            if "Oracle" in sheet or sheet.endswith("Order") or sheet.endswith("FOR"):
                continue
            region = sheet.replace("By-Pool ", "").strip()
            ws = wb[sheet]
            rows = [tuple(safe_read(c) for c in r) for r in ws.iter_rows(values_only=True)]
            hi = find_header_row(rows, "Ship $ (this week)")
            if hi is None:
                hi = find_header_row(rows, "NN Cust Store (gross)")
            if hi is None:
                continue
            cmap = col_map(rows[hi])
            i_ship = find_col(cmap, "Ship $ (this week)")
            i_xfer = find_col(cmap, "To Transfer $")
            i_ops = find_col(cmap, "Ops-Miss $")
            i_nn_g = find_col(cmap, "NN Cust Store (gross) $")
            i_nn_p = find_col(cmap, "NN Pool $ (ref)")
            i_pool_u = find_col(cmap, "NN Pool (u)")
            i_gross_u = find_col(cmap, "Gross Need (u)")
            i_ab = find_col(cmap, "A+B on hand $")
            i_genus = find_col(cmap, "Genus")
            i_form = find_col(cmap, "Form")
            i_size = find_col(cmap, "Size")

            sums = {
                "ship_$": 0.0,
                "to_transfer_$": 0.0,
                "ops_miss_$": 0.0,
                "nn_cust_store_gross_$": 0.0,
                "nn_pool_$": 0.0,
                "nn_pool_u": 0.0,
                "gross_need_u": 0.0,
                "ab_on_hand_$": 0.0,
            }
            pool_count = 0
            top: list[tuple[float, dict[str, Any]]] = []

            for row in rows[hi + 1 :]:
                if not row or row[0] is None:
                    continue
                ship = as_float(row[i_ship]) if i_ship is not None else None
                if ship is None and all(
                    as_float(row[i]) is None
                    for i in (i_xfer, i_nn_g, i_pool_u)
                    if i is not None
                ):
                    continue
                pool_count += 1

                def add(key: str, idx: int | None) -> float:
                    v = as_float(row[idx]) if idx is not None else None
                    if v is None:
                        return 0.0
                    sums[key] += v
                    return v

                add("ship_$", i_ship)
                add("to_transfer_$", i_xfer)
                add("ops_miss_$", i_ops)
                nn_g = add("nn_cust_store_gross_$", i_nn_g)
                add("nn_pool_$", i_nn_p)
                add("nn_pool_u", i_pool_u)
                add("gross_need_u", i_gross_u)
                add("ab_on_hand_$", i_ab)

                top.append(
                    (
                        nn_g,
                        {
                            "genus": row[i_genus] if i_genus is not None else None,
                            "form": row[i_form] if i_form is not None else None,
                            "size": row[i_size] if i_size is not None else None,
                            "nn_cust_store_gross_$": nn_g,
                            "ship_$": as_float(row[i_ship]) if i_ship is not None else None,
                        },
                    )
                )

            top.sort(key=lambda t: t[0], reverse=True)
            markets[region] = {
                "pool_count": pool_count,
                "totals": {k: round(v, 2) for k, v in sums.items()},
                "top_pools_by_nn_cust_store": [t[1] for t in top[:30]],
            }

        return {
            "channel": channel,
            "file": path.name,
            "unc_path": str(path),
            "markets": markets,
            "build_health": extract_build_health(path, channel),
        }
    finally:
        wb.close()


# ── On Hand & Register (Set 3) ────────────────────────────────────────────


def extract_ohr(path: Path) -> dict[str, Any]:
    m = re.match(
        r"(HD|LOW)\s+(N\.CA|S\.CA)\s+On Hand and Register Analysis\s+\((Weekly|YTD)\)",
        path.name,
        re.I,
    )
    channel = m.group(1).upper() if m else None
    region = m.group(2) if m else None
    edition = m.group(3) if m else None

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        kpis: dict[str, Any] = {}
        nn_block: dict[str, Any] = {}
        if "Exec Summary" in wb.sheetnames:
            ws = wb["Exec Summary"]
            rows = [tuple(safe_read(c) for c in r) for r in ws.iter_rows(values_only=True)]
            # Key metrics table (rows with Metric / This year / Last year / Var $)
            hi = find_header_row(rows, "This year")
            if hi is not None:
                for row in rows[hi + 1 :]:
                    if not row or row[0] is None:
                        if kpis:
                            break
                        continue
                    label = str(row[0]).strip()
                    if label.startswith("Why REGISTER") or label.startswith("NET NEED"):
                        break
                    kpis[label] = {
                        "ty": as_float(row[1]) if len(row) > 1 else None,
                        "ly": as_float(row[2]) if len(row) > 2 else None,
                        "var_$": as_float(row[3]) if len(row) > 3 else None,
                    }
            for row in rows:
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                if "NN Cust Store" in label:
                    nn_block["nn_cust_store_$"] = as_float(row[1])
                elif "Shippable now from A+B" in label:
                    nn_block["shippable_ab_$"] = as_float(row[1])
                    nn_block["shippable_ab_note"] = row[2] if len(row) > 2 else None
                elif "additional from QC" in label:
                    nn_block["additional_qc_$"] = as_float(row[1])
                elif "Total shippable now" in label:
                    nn_block["total_shippable_$"] = as_float(row[1])
                elif label.startswith("Uncovered"):
                    nn_block["uncovered_$"] = as_float(row[1])

        genus_master_rows = 0
        if "Genus Master" in wb.sheetnames:
            ws = wb["Genus Master"]
            # count data rows cheaply
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if row and row[0] is not None:
                    genus_master_rows += 1

        return {
            "channel": channel,
            "region": region,
            "edition": edition,
            "file": path.name,
            "unc_path": str(path),
            "sheet_count": len(wb.sheetnames),
            "kpis": kpis,
            "net_need_block": nn_block,
            "genus_master_row_count": genus_master_rows,
        }
    finally:
        wb.close()


# ── Sales Manager / Set 1 index ───────────────────────────────────────────


def extract_sales_manager_summary(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        out: dict[str, Any] = {
            "file": path.name,
            "unc_path": str(path),
            "headline": {},
            "nn_customer": {},
            "plan_var_net_$": None,
        }
        if "Executive Summary" not in wb.sheetnames:
            return out
        ws = wb["Executive Summary"]
        rows = [tuple(safe_read(c) for c in r) for r in ws.iter_rows(values_only=True)]
        for row in rows:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            if label == "Combined Total" and len(row) > 6:
                # HEADLINE table: Plan_thru … Net Need (units) …
                out["headline"] = {
                    "plan_thru_u": as_float(row[1]),
                    "actual_thru_u": as_float(row[2]),
                    "plan_var_u": as_float(row[3]),
                    "plan_var_$_retail": as_float(row[4]),
                    "plan_var_$_wholesale": as_float(row[5]),
                    "nn_plan_u": as_float(row[6]),
                    "nn_plan_$_retail": as_float(row[7]) if len(row) > 7 else None,
                    "nn_plan_$_wholesale": as_float(row[8]) if len(row) > 8 else None,
                }
                out["plan_var_net_$"] = as_int_round(row[5])
            if label.startswith("NET NEED (Customer)"):
                out["nn_customer"] = {
                    "units": as_float(row[1]),
                    "$_retail": as_float(row[2]) if len(row) > 2 else None,
                    "$_wholesale": as_float(row[3]) if len(row) > 3 else None,
                }
            if label.startswith("NET NEED (Plan)"):
                out["nn_plan_block"] = {
                    "units": as_float(row[1]),
                    "$_retail": as_float(row[2]) if len(row) > 2 else None,
                    "$_wholesale": as_float(row[3]) if len(row) > 3 else None,
                }
        return out
    finally:
        wb.close()


def index_set1_file(path: Path) -> dict[str, Any]:
    """Index only for huge Store Detail files; light summary for others."""
    refresh, date = parse_refresh_from_name(path.name)
    info: dict[str, Any] = {
        "file": path.name,
        "unc_path": str(path),
        "size_bytes": path.stat().st_size,
        "refresh": refresh,
        "date": date,
        "role": "summary",
    }
    if "Store Detail" in path.name:
        info["role"] = "store_detail_index_only"
        info["note"] = (
            "Large companion file — not fully loaded. Portal should offer UNC download."
        )
        # sheet names only
        wb = load_workbook(path, read_only=True)
        try:
            info["sheets"] = list(wb.sheetnames)
        finally:
            wb.close()
        return info

    if path.name.startswith("Sales Manager Summary"):
        info["role"] = "sales_manager_summary"
        info["extract"] = extract_sales_manager_summary(path)
        return info

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        info["sheets"] = list(wb.sheetnames)
        if "Build Health" in wb.sheetnames:
            # status line only
            ws = wb["Build Health"]
            for row in ws.iter_rows(values_only=True):
                if row and row[0] and "BUILD STATUS" in str(row[0]).upper():
                    info["build_status"] = str(row[0])
                    break
    finally:
        wb.close()
    return info


# ── Rep Orders index ──────────────────────────────────────────────────────


def index_rep_order(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        rep_name = None
        tabs: list[dict[str, Any]] = []
        channels: set[str] = set()
        regions: set[str] = set()
        total_ship = 0.0
        total_for = 0.0
        all_stores: set[str] = set()

        for sheet_name in wb.sheetnames:
            parts = sheet_name.split()
            if len(parts) < 3:
                continue
            cust, region, source = parts[0], parts[1], parts[2]
            if cust not in {"HD", "LOW"} or source not in {"Order", "FOR"}:
                continue
            channels.add("LOW" if cust == "LOW" else "HD")
            regions.add(region)

            ws = wb[sheet_name]
            materials = 0
            stores: list[str] = []
            dollars = 0.0
            header_seen = False

            for row in ws.iter_rows(values_only=True):
                vals = [safe_read(c) for c in row]
                if not vals:
                    continue
                if vals[0] == "# of materials":
                    materials = int(as_float(vals[1]) or 0)
                    continue
                if vals[0] == "Rep":
                    header_seen = True
                    continue
                if not header_seen:
                    continue
                if vals[0] is None:
                    continue
                # data row
                if rep_name is None and isinstance(vals[0], str):
                    rep_name = vals[0].strip()
                store = vals[1]
                if store is not None:
                    stores.append(str(store).strip())
                    all_stores.add(str(store).strip())
                d = as_float(vals[5]) if len(vals) > 5 else None
                if d:
                    dollars += d

            if source == "FOR":
                total_for += dollars
            else:
                total_ship += dollars

            tabs.append(
                {
                    "tab": sheet_name,
                    "channel": "LOW" if cust == "LOW" else "HD",
                    "region": region,
                    "source": source,
                    "materials": materials,
                    "store_count": len(set(stores)),
                    "total_ship_$": round(dollars, 2),
                }
            )

        # Channel/region for index: primary = first tab; also expose all
        primary_ch = sorted(channels)[0] if channels else None
        primary_rg = sorted(regions)[0] if regions else None
        note = None
        stem = path.stem.upper()
        if "BARCENAS" in stem or "BARCENES" in stem:
            note = (
                "BARCENAS/BARCENES Customer Master typo pair — keep both files; "
                "do not deduplicate."
            )

        return {
            "rep_name": rep_name or path.stem.split(" - ")[0],
            "channel": primary_ch,
            "region": primary_rg,
            "channels": sorted(channels),
            "regions": sorted(regions),
            "filename": path.name,
            "unc_path": str(path),
            "store_count": len(all_stores),
            "total_ship": round(total_ship, 2),
            "total_transfer": 0.0,  # transfers live in Transfers set, not rep books
            "total_for": round(total_for, 2),
            "tabs": tabs,
            "note": note,
        }
    finally:
        wb.close()


# ── Supply / org scan (when org columns exist) ─────────────────────────────


def scan_org_leak(path: Path) -> list[str]:
    """Return any TX/FL org codes found in a workbook (empty = clean)."""
    found: set[str] = set()
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            hi = None
            for i, row in enumerate(rows[:30]):
                if not row:
                    continue
                headers = [norm_header(c).lower() for c in row]
                if any(h in {"org", "org code", "org_code", "farm", "source farm"} for h in headers):
                    hi = i
                    cmap = col_map(row)
                    break
            if hi is None:
                continue
            i_org = find_col(cmap, "Org Code", "Org", "org_code", "Farm", "Source Farm")
            if i_org is None:
                continue
            for row in rows[hi + 1 :]:
                if not row or i_org >= len(row) or row[i_org] is None:
                    continue
                code = str(row[i_org]).strip().upper()
                if code in TX_FL_ORGS:
                    found.add(code)
    finally:
        wb.close()
    return sorted(found)


# ── Validation ────────────────────────────────────────────────────────────


def validate(output: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    fn = output["four_numbers"]
    snap = output.get("snapshot") or {}
    refresh = str(snap.get("refresh") or "")
    date = str(snap.get("date") or "")
    is_baseline = (refresh, date) == BASELINE_REFRESH

    if is_baseline:
        checks = [
            ("ship_this_week", TARGETS["ship_this_week"], 1000),
            ("to_transfer", TARGETS["to_transfer"], 1000),
            ("nn_plan", TARGETS["nn_plan"], 5000),
            ("nn_cust_store", TARGETS["nn_cust_store"], 5000),
        ]
        for key, target, tol in checks:
            val = fn.get(key)
            if val is None:
                errors.append(f"{key}: missing")
                continue
            if abs(val - target) >= tol:
                errors.append(f"{key}: got {val}, expected ~{target} (±{tol})")
    else:
        for key in ("ship_this_week", "to_transfer", "nn_plan", "nn_cust_store"):
            val = fn.get(key)
            if val is None:
                errors.append(f"{key}: missing")
            elif not isinstance(val, (int, float)) or val < 0:
                errors.append(f"{key}: invalid value {val}")
        if (fn.get("ship_this_week") or 0) <= 0:
            errors.append("ship_this_week: expected positive wholesale $")

    rep_n = len(output.get("rep_orders", []))
    if rep_n == 0:
        errors.append("rep_orders: empty — expected published rep workbooks")
    elif is_baseline and rep_n != 40:
        errors.append(f"rep_orders: got {rep_n}, expected 40")

    if fn.get("nn_plan") == fn.get("nn_cust_store"):
        errors.append("NN Plan and NN Cust Store must not be equal")
    for bh in output.get("build_health", {}).get("store_driven", []):
        if bh.get("status") == "FAIL":
            errors.append(f"Build Health FAIL in {bh.get('file')}")
    leaks = output.get("build_health", {}).get("tx_fl_org_leaks", [])
    if leaks:
        errors.append(f"TX/FL farm inventory leaked: {leaks}")
    return errors


# ── Main ──────────────────────────────────────────────────────────────────


def build_output(reports: Path) -> dict[str, Any]:
    root, sets = resolve_set_dirs(reports)
    set1, set2, set3, set4, set5 = (
        sets["set1"],
        sets["set2"],
        sets["set3"],
        sets["set4"],
        sets["set5"],
    )

    combined_candidates = list(set2.glob("*Combined Summary*.xlsx"))
    if not combined_candidates:
        raise FileNotFoundError(f"No Combined Summary workbook in {set2}")
    combined_path = combined_candidates[0]
    combined = extract_combined_summary(combined_path)

    hd_sd = next(set2.glob("HD Store Driven*.xlsx"))
    low_sd = next(set2.glob("LOW Store Driven*.xlsx"))
    store_rec = [
        extract_store_driven(hd_sd, "HD"),
        extract_store_driven(low_sd, "LOW"),
    ]

    transfers = []
    for p in list_xlsx(set4):
        ch = "HD" if p.name.upper().startswith("HD") else "LOW"
        transfers.append(extract_transfers(p, ch))

    ohr: dict[str, Any] = {"weekly": {}, "ytd": {}}
    for p in list_xlsx(set3):
        rec = extract_ohr(p)
        key = f"{(rec['channel'] or '').lower()}_{(rec['region'] or '').lower().replace('.', '')}"
        bucket = "weekly" if rec["edition"] == "Weekly" else "ytd"
        ohr[bucket][key] = rec

    rep_orders = [index_rep_order(p) for p in list_xlsx(set5)]

    set1_index = [index_set1_file(p) for p in list_xlsx(set1)]
    sms = next((x for x in set1_index if x.get("role") == "sales_manager_summary"), None)

    bh_hd = extract_build_health(hd_sd, "HD")
    bh_low = extract_build_health(low_sd, "LOW")

    leaks: list[str] = []
    for p in list_xlsx(set1):
        if "Store Detail" in p.name:
            continue
        leaks.extend(scan_org_leak(p))
    leaks = sorted(set(leaks))

    refresh = combined.get("refresh") or "unknown"
    date = combined.get("date") or "unknown"

    file_counts = {
        "sales_variance_allocation": len(list_xlsx(set1)),
        "store_driven": len(list_xlsx(set2)),
        "on_hand_register": len(list_xlsx(set3)),
        "transfers": len(list_xlsx(set4)),
        "rep_orders": len(list_xlsx(set5)),
    }

    output: dict[str, Any] = {
        "snapshot": {
            "refresh": refresh,
            "date": date,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "reports_root": str(reports),
            "resolved_sets_root": str(root),
        },
        "four_numbers": combined["four_numbers"],
        "exec_summary": {
            "combined_summary": {
                "segments": combined["segments"],
                "plan_var_summary": combined["plan_var_summary"],
                "plan_var_net_combined_summary_$": combined[
                    "plan_var_net_combined_summary_$"
                ],
                "notes": combined["notes"],
            },
            "sales_manager_summary": sms.get("extract") if sms else None,
            "set1_files": set1_index,
            "stub_note": (
                "Set 1 is being phased toward retirement. Prefer Set 2 Combined "
                "Summary for Ship / Transfer / NN tiles."
            ),
        },
        "store_recommendation": store_rec,
        "on_hand_register": ohr,
        "transfers": transfers,
        "rep_orders": rep_orders,
        "build_health": {
            "refresh": refresh,
            "date": date,
            "file_counts": file_counts,
            "expected_file_counts": {
                "sales_variance_allocation": 7,
                "store_driven": 3,
                "on_hand_register": 8,
                "transfers": 2,
                "rep_orders": 40,
            },
            "store_driven": [bh_hd, bh_low],
            "supply_filter": {
                "wcro_orgs": sorted(WCRO_ORGS),
                "excluded_tx_fl_orgs": sorted(TX_FL_ORGS),
                "tx_fl_org_leaks": leaks,
                "result": "PASSED" if not leaks else "FAILED",
                "note": (
                    "Published report extract: org-code scan on Set 1 (non Store Detail). "
                    "Inventory Available QTY / grade rules are engine-side; extractor "
                    "does not recompute Net Need."
                ),
            },
            "grade_buckets": {
                "ab_grades": sorted(AB_GRADES),
                "qc_grades": sorted(QC_GRADES),
                "note": "Documented only — recomputation is out of scope for portal extract.",
            },
            "fiscal_calendar_note": "4-5-4; weeks end Saturday",
            "known_flags": [
                "CITHYF3305 has −96u A+B in Inventory Transform — clamped to 0, flagged",
                "265 items have HD plan but no SKU (~$308K plan) — logged, not failed",
                "Negative on-hand rows in vendor store feeds — log, do not fail",
                "BARCENAS / BARCENES Jose — two files for same rep (Customer Master typo)",
            ],
            "tx_fl_org_leaks": leaks,
        },
        "change_log": [
            {
                "refresh": refresh,
                "date": date,
                "source": "extractor seed",
                "notes": [
                    "Mid-week data refresh 5.29 / 2026-08-06; version unchanged.",
                    "Ship/Transfer from Combined Summary (5.28 definition).",
                ],
            }
        ],
        "rules_acknowledged": [
            "Net Need = max(0, Forward Demand − (Curr Inv + On Order))",
            "Plan variance = Plan − Actual; positive = behind",
            "Inventory column = Available QTY only (engine)",
            "Pool key = (item, form)",
            "NN_Customer must never alias to NN_Plan",
            "Ship This Week excludes cross-region transfers",
            "TX/FL orgs out of scope",
        ],
    }
    return output


def parse_args() -> argparse.Namespace:
    repo = Path(__file__).resolve().parents[2]
    p = argparse.ArgumentParser(description="WCRO portal extractor")
    p.add_argument(
        "--weeklydrop",
        type=Path,
        default=None,
        help=f"WeeklyDrop folder (default prefer: {DEFAULT_WEEKLYDROP})",
    )
    p.add_argument(
        "--reports",
        type=Path,
        default=None,
        help=f"Explicit reports root (fallback: {DEFAULT_REPORTS})",
    )
    p.add_argument(
        "--out",
        type=Path,
        default=repo / "data" / "wcro_data.json",
        help="Output JSON path",
    )
    p.add_argument(
        "--change-history",
        type=Path,
        default=repo / "data" / "change_history_wcro.json",
    )
    p.add_argument(
        "--skip-validate",
        action="store_true",
        help="Write even if gates fail (debug only)",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    reports = pick_reports_root(args.weeklydrop, args.reports)
    print(f"Reports root: {reports}")
    if not reports.is_dir():
        print(f"ERROR: reports folder not found: {reports}", file=sys.stderr)
        return 2

    print("Extracting…")
    output = build_output(reports)
    fn = output["four_numbers"]
    print(
        "Four Numbers:",
        f"Ship ${fn['ship_this_week']:,} |",
        f"Transfer ${fn['to_transfer']:,} |",
        f"NN Plan {fn['nn_plan']:,} u |",
        f"NN Cust Store (pool) {fn['nn_cust_store']:,} u",
        f"(gross {fn.get('nn_cust_store_gross_u'):,} u)",
    )
    print(f"Rep orders indexed: {len(output['rep_orders'])}")

    errors = validate(output)
    output["build_health"]["validation"] = {
        "targets": TARGETS,
        "passed": not errors,
        "errors": errors,
    }

    if errors:
        print("VALIDATION FAILED:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        if not args.skip_validate:
            return 1
        print("Continuing because --skip-validate was set", file=sys.stderr)
    else:
        print("VALIDATION PASSED (Four Numbers + rep count + NN distinct)")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(output, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {args.out} ({args.out.stat().st_size:,} bytes)")

    history = {
        "seeded_from": output["snapshot"],
        "entries": output["change_log"],
        "note": "Seed only — enrich from workbook Change Log tabs in later sprints.",
    }
    args.change_history.parent.mkdir(parents=True, exist_ok=True)
    args.change_history.write_text(json.dumps(history, indent=2), encoding="utf-8")
    print(f"Wrote {args.change_history}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
