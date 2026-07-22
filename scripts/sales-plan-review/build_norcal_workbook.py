"""
build_norcal_workbook.py
========================

Builds NOR CAL Forward Looking INV vs Sales Plan 050426.xlsx — combined
Operations + Sales workbook for Head of Operations and Sales leadership.
"""

from __future__ import annotations
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, "/sessions/optimistic-beautiful-ramanujan/mnt/Sales Plan Review")
import nor_cal_forward as M
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants & styling
# ---------------------------------------------------------------------------

PROJECT_DIR = Path("/sessions/optimistic-beautiful-ramanujan/mnt/Sales Plan Review")
ARCHIVE_DIR = PROJECT_DIR / "Archive"
OUT_PATH = PROJECT_DIR / "NOR CAL Forward Looking INV vs Sales Plan 050426.xlsx"


# ---------------------------------------------------------------------------
# Archive previous reports before generating new
# ---------------------------------------------------------------------------

def archive_previous_reports():
    """Move any prior NOR CAL forward-looking report files into Sales Plan Review/Archive/
    before generating a new one. Keeps only the current run's output in the top-level
    folder. All older versions go to Archive/ with original filename preserved."""
    import shutil
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    patterns = [
        "NOR CAL Forward Looking INV vs Sales Plan*.xlsx",
        "NOR CAL Forward Reproduction*.xlsx",
        "NOR CAL Forward Validation*.xlsx",
    ]
    archived = []
    for pat in patterns:
        for f in PROJECT_DIR.glob(pat):
            if f.name.startswith("~$"):
                continue
            # If the existing file's name matches the new output's, we still
            # archive it (as a renamed copy) so the prior build is preserved
            # rather than silently overwritten.
            dest = ARCHIVE_DIR / f.name
            if dest.exists():
                stem = dest.stem; suffix = dest.suffix; i = 1
                while True:
                    candidate = ARCHIVE_DIR / f"{stem}.bak{i}{suffix}"
                    if not candidate.exists():
                        dest = candidate; break
                    i += 1
            shutil.move(str(f), str(dest))
            archived.append(f.name)
    if archived:
        log(f"Archived {len(archived)} prior report(s) to Archive/: {', '.join(archived)}")
    else:
        log("No prior reports to archive (clean slate).")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", start_color="D9E2F3")
SUBHEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="000000")
SECTION_FILL = PatternFill("solid", start_color="EDEDED")
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color="000000")
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="000000")
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")
THIN = Side(border_style="thin", color="BFBFBF")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

NEG_FILL = PatternFill("solid", start_color="FCE4D6")
POS_FILL = PatternFill("solid", start_color="E2EFDA")
TOTAL_FILL = PatternFill("solid", start_color="F2F2F2")

FMT_INT = "#,##0;(#,##0);-"
FMT_DLR = "$#,##0;($#,##0);-"
FMT_DLR2 = "$#,##0.00;($#,##0.00);-"
FMT_PCT = "0.0%;(0.0%);-"

CUSTOMERS = ["HD", "Lowes", "Walmart", "West Coast", "Midwest"]
NBB_CUSTOMERS = {"West Coast", "Midwest"}
BB_CUSTOMERS  = {"HD", "Lowes", "Walmart"}
FWD_MONTHS = [5, 6, 7, 8, 9, 10, 11, 12]
YTD_MONTHS = [1, 2, 3, 4]
MONTH_LBL = {5:"May", 6:"Jun", 7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec"}
HIST_YEARS = [2023, 2024, 2025]

OFFICIAL_PLAN_DLR = 34_268_765   # NOR CAL official 2026 plan $ per CEO


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def apply_header(ws, row, n_cols, fill=HEADER_FILL, font=HEADER_FONT, height=None):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    if height:
        ws.row_dimensions[row].height = height


def set_col_widths(ws, widths):
    """widths is dict of {col_letter: width} or list of widths starting at col A."""
    if isinstance(widths, dict):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    else:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w


# ---------------------------------------------------------------------------
# Compute metrics — all the tables we need
# ---------------------------------------------------------------------------

def compute_metrics(d):
    """Compute all data tables needed across tabs."""
    log("Computing metrics for all tabs...")

    plan = d["plan"]
    plan_lifted = d["plan_lifted"]
    ytd = d["ytd"]
    hist = d["hist"]
    cust_alloc = d["cust_alloc"]
    prices = d["prices"]
    items = d["items"]
    smoothed = d["smoothed"]
    lift_groups = d["lift_groups"]
    inv = d["inv"]

    # Item descriptions
    item_meta = items[["Item Num", "Item Desc", "KI", "Genus", "Size", "Brand"]].drop_duplicates(subset=["Item Num"])

    # ---- Per (Customer, Item) full panel ----
    orig_yr = plan.groupby(["Customer", "Item Num"])["PlanQty"].sum().reset_index().rename(columns={"PlanQty": "OrigYE_QTY"})
    orig_md = plan[plan["MonthNum"].isin(FWD_MONTHS)].groupby(["Customer", "Item Num"])["PlanQty"].sum().reset_index().rename(columns={"PlanQty": "OrigMD_QTY"})
    orig_jad = plan[plan["MonthNum"].isin(YTD_MONTHS)].groupby(["Customer", "Item Num"])["PlanQty"].sum().reset_index().rename(columns={"PlanQty": "OrigJAD_QTY"})

    lifted_md = plan_lifted[plan_lifted["MonthNum"].isin(FWD_MONTHS)].groupby(["Customer", "Item Num"])["LiftedQty"].sum().reset_index().rename(columns={"LiftedQty": "LiftedMD_QTY"})

    ytd_g = ytd.groupby(["Customer", "Item Num"]).agg(YTD_QTY=("Qty", "sum"), YTD_Rev=("Revenue", "sum")).reset_index()
    # Two-stage forward fulfillment: PlanFilled (Stage 1) + LiftFilled (Stage 2)
    pf_g = cust_alloc.groupby(["Customer", "Item Num"])["PlanFilled"].sum().reset_index().rename(columns={"PlanFilled": "PlanFill_QTY"})
    lf_g = cust_alloc.groupby(["Customer", "Item Num"])["LiftFilled"].sum().reset_index().rename(columns={"LiftFilled": "LiftFill_QTY"})
    ff_g = pf_g.merge(lf_g, on=["Customer", "Item Num"], how="outer").fillna(0)
    ff_g["FwdFill_QTY"] = ff_g["PlanFill_QTY"] + ff_g["LiftFill_QTY"]

    # Build the full panel
    panel = orig_yr.merge(orig_md, on=["Customer","Item Num"], how="outer") \
                   .merge(orig_jad, on=["Customer","Item Num"], how="outer") \
                   .merge(lifted_md, on=["Customer","Item Num"], how="outer") \
                   .merge(ytd_g, on=["Customer","Item Num"], how="outer") \
                   .merge(ff_g[["Customer","Item Num","PlanFill_QTY","LiftFill_QTY","FwdFill_QTY"]],
                          on=["Customer","Item Num"], how="outer").fillna(0)
    panel = panel.merge(prices[["Customer","Item Num","Price","Tier"]], on=["Customer","Item Num"], how="left")
    panel["Price"] = panel["Price"].fillna(0)
    panel["Tier"] = panel["Tier"].fillna(0).astype(int)
    panel = panel.merge(item_meta, on="Item Num", how="left")
    # keep only customers we care about
    panel = panel[panel["Customer"].isin(CUSTOMERS)]

    panel["Achieved_QTY"] = panel["YTD_QTY"] + panel["FwdFill_QTY"]
    panel["Achieved_QTY_Capped"] = np.minimum(panel["Achieved_QTY"], panel["OrigYE_QTY"])
    panel["Miss_QTY"] = (panel["OrigYE_QTY"] - panel["Achieved_QTY_Capped"]).clip(lower=0)
    panel["Excess_QTY_vsPlan"] = (panel["Achieved_QTY"] - panel["OrigYE_QTY"]).clip(lower=0)

    panel["OrigYE_$"] = panel["OrigYE_QTY"] * panel["Price"]
    panel["LiftedMD_$"] = panel["LiftedMD_QTY"] * panel["Price"]
    panel["OrigJAD_$"] = panel["OrigJAD_QTY"] * panel["Price"]
    panel["YTD_$"] = panel["YTD_QTY"] * panel["Price"]
    panel["FwdFill_$"] = panel["FwdFill_QTY"] * panel["Price"]
    panel["PlanFill_$"] = panel["PlanFill_QTY"] * panel["Price"]
    panel["LiftFill_$"] = panel["LiftFill_QTY"] * panel["Price"]
    panel["Achieved_$"] = panel["Achieved_QTY_Capped"] * panel["Price"]
    panel["Miss_$"] = panel["Miss_QTY"] * panel["Price"]
    panel["Excess_$_vsPlan"] = panel["Excess_QTY_vsPlan"] * panel["Price"]

    # ---- Smoothed history per (Customer, KI) ----
    sm_ki = smoothed[smoothed["LiftGroup"].astype(str).str.startswith("KI:")].copy()
    sm_ki["KI_extracted"] = sm_ki["LiftGroup"].astype(str).str.replace("KI:", "", regex=False)

    # Aggregate smoothed history at (Customer, KI, Year separate columns)
    h_with_ki = hist.merge(items[["Item Num","KI"]].drop_duplicates(subset=["Item Num"]), on="Item Num", how="left")
    h_yr_ki = h_with_ki.dropna(subset=["KI"]).groupby(["Customer","KI","Year"])["Qty"].sum().reset_index()
    h_pivot = h_yr_ki.pivot_table(index=["Customer","KI"], columns="Year", values="Qty", fill_value=0).reset_index()
    h_pivot.columns = [str(c) if not isinstance(c, str) else c for c in h_pivot.columns]
    for y in HIST_YEARS:
        col = str(y)
        if col not in h_pivot.columns:
            h_pivot[col] = 0
    h_pivot["Hist_Avg"] = h_pivot[[str(y) for y in HIST_YEARS]].mean(axis=1)

    # outlier flag — drop a year if its value > 10× the mean of the OTHER years.
    # Earlier rule used raw_mean (inclusive of the year being checked) which was
    # mathematically impossible to trigger. Fixed per CEO direction 2026-05-07.
    def smooth_row(row):
        vals = []
        flags = []
        for y in HIST_YEARS:
            v = row[str(y)]
            others = [row[str(yy)] for yy in HIST_YEARS if yy != y]
            others_mean = (sum(others) / len(others)) if any(others) else 0
            if others_mean > 0 and v > 10 * others_mean:
                flags.append(str(y))
                continue
            vals.append(v)
        sm = float(np.mean(vals)) if vals else 0.0
        return pd.Series([sm, ",".join(flags) if flags else ""])
    h_pivot[["Smoothed_Hist", "Outlier_Years"]] = h_pivot.apply(smooth_row, axis=1)

    # ---- KI-level metrics ----
    ki_panel = panel.groupby(["Customer", "KI"]).agg(
        OrigYE_QTY=("OrigYE_QTY","sum"), OrigMD_QTY=("OrigMD_QTY","sum"),
        OrigJAD_QTY=("OrigJAD_QTY","sum"), LiftedMD_QTY=("LiftedMD_QTY","sum"),
        YTD_QTY=("YTD_QTY","sum"), FwdFill_QTY=("FwdFill_QTY","sum"),
        OrigYE_dlr=("OrigYE_$","sum"), Miss_QTY_item=("Miss_QTY","sum"),
        Miss_dlr_item=("Miss_$","sum"),
        YTD_dlr=("YTD_$","sum"),
    ).reset_index()
    # KI-level miss (capped at KI level, not per-item)
    ki_panel["Achieved_QTY_KI"] = np.minimum(ki_panel["YTD_QTY"] + ki_panel["FwdFill_QTY"], ki_panel["OrigYE_QTY"])
    ki_panel["Miss_QTY_KI"] = (ki_panel["OrigYE_QTY"] - ki_panel["Achieved_QTY_KI"]).clip(lower=0)
    # Apply weighted avg price per (Customer, KI) for KI-level miss $
    avg_price_q = panel.groupby(["Customer","KI"]).apply(
        lambda g: (g["OrigYE_QTY"] * g["Price"]).sum() / g["OrigYE_QTY"].sum() if g["OrigYE_QTY"].sum() > 0 else 0,
        include_groups=False,
    ).reset_index()
    avg_price_q.columns = ["Customer","KI","AvgPrice"]
    ki_panel = ki_panel.merge(avg_price_q, on=["Customer","KI"], how="left")
    ki_panel["AvgPrice"] = ki_panel["AvgPrice"].fillna(0)
    ki_panel["Miss_dlr_KI"] = ki_panel["Miss_QTY_KI"] * ki_panel["AvgPrice"]
    ki_panel["Excess_QTY_vsPlan"] = (ki_panel["YTD_QTY"] + ki_panel["FwdFill_QTY"] - ki_panel["OrigYE_QTY"]).clip(lower=0)
    ki_panel["Excess_dlr_vsPlan"] = ki_panel["Excess_QTY_vsPlan"] * ki_panel["AvgPrice"]
    ki_panel["Lift_QTY"] = (ki_panel["LiftedMD_QTY"] - ki_panel["OrigMD_QTY"]).clip(lower=0)
    ki_panel["Lift_dlr"] = ki_panel["Lift_QTY"] * ki_panel["AvgPrice"]

    # bring in smoothed history per (Customer, KI)
    ki_panel = ki_panel.merge(h_pivot[["Customer","KI","2023","2024","2025","Smoothed_Hist","Outlier_Years"]], on=["Customer","KI"], how="left")
    for c in ["2023","2024","2025","Smoothed_Hist"]:
        ki_panel[c] = ki_panel[c].fillna(0)
    ki_panel["Outlier_Years"] = ki_panel["Outlier_Years"].fillna("")

    # YTD pace metrics
    ki_panel["YTD_Pace_vs_Plan"] = np.where(ki_panel["OrigJAD_QTY"] > 0,
                                              ki_panel["YTD_QTY"] / ki_panel["OrigJAD_QTY"], np.nan)
    # YoY: same period (Jan-Apr) actuals 2025 vs 2026
    h_jan_apr_2025 = hist[(hist["Year"]==2025) & (hist["Month"].isin(YTD_MONTHS))].merge(
        items[["Item Num","KI"]].drop_duplicates(subset=["Item Num"]), on="Item Num", how="left"
    ).dropna(subset=["KI"]).groupby(["Customer","KI"])["Qty"].sum().reset_index().rename(columns={"Qty":"YTD_2025_QTY"})
    ki_panel = ki_panel.merge(h_jan_apr_2025, on=["Customer","KI"], how="left")
    ki_panel["YTD_2025_QTY"] = ki_panel["YTD_2025_QTY"].fillna(0)
    ki_panel["YTD_YoY"] = np.where(ki_panel["YTD_2025_QTY"] > 0,
                                     ki_panel["YTD_QTY"] / ki_panel["YTD_2025_QTY"], np.nan)

    # 2026 YTD avg sell price per (Customer, Item) → roll to (Customer, KI) weighted
    ytd_price_per_item = ytd.groupby(["Customer","Item Num"]).agg(qty=("Qty","sum"), rev=("Revenue","sum")).reset_index()
    ytd_price_per_item["AvgPrice_2026YTD"] = np.where(ytd_price_per_item["qty"]>0, ytd_price_per_item["rev"]/ytd_price_per_item["qty"], np.nan)
    h_25 = hist[hist["Year"]==2025].groupby(["Customer","Item Num"]).agg(qty=("Qty","sum"), rev=("Revenue","sum")).reset_index()
    h_25["AvgPrice_2025"] = np.where(h_25["qty"]>0, h_25["rev"]/h_25["qty"], np.nan)
    panel = panel.merge(ytd_price_per_item[["Customer","Item Num","AvgPrice_2026YTD"]], on=["Customer","Item Num"], how="left")
    panel = panel.merge(h_25[["Customer","Item Num","AvgPrice_2025"]], on=["Customer","Item Num"], how="left")
    panel["YoY_Price_Pct"] = (panel["AvgPrice_2026YTD"] - panel["AvgPrice_2025"]) / panel["AvgPrice_2025"]

    # Plan-implied price = OrigYE_$ / OrigYE_QTY (= Price column, but also explicit)
    panel["PlanImpliedPrice"] = np.where(panel["OrigYE_QTY"]>0, panel["OrigYE_$"]/panel["OrigYE_QTY"], np.nan)

    # ---- KI-level top-tier (across customers, total) ----
    ki_total = ki_panel.groupby("KI").agg(
        OrigYE_QTY=("OrigYE_QTY","sum"), OrigMD_QTY=("OrigMD_QTY","sum"),
        OrigJAD_QTY=("OrigJAD_QTY","sum"),
        LiftedMD_QTY=("LiftedMD_QTY","sum"), YTD_QTY=("YTD_QTY","sum"),
        FwdFill_QTY=("FwdFill_QTY","sum"),
        OrigYE_dlr=("OrigYE_dlr","sum"), Miss_QTY_KI=("Miss_QTY_KI","sum"),
        Miss_dlr_KI=("Miss_dlr_KI","sum"), Excess_QTY_vsPlan=("Excess_QTY_vsPlan","sum"),
        Excess_dlr_vsPlan=("Excess_dlr_vsPlan","sum"),
        Lift_QTY=("Lift_QTY","sum"), Lift_dlr=("Lift_dlr","sum"),
        Smoothed_Hist=("Smoothed_Hist","sum"),
    ).reset_index()
    ki_total["Achieved_QTY"] = ki_total["YTD_QTY"] + ki_total["FwdFill_QTY"]
    # Pct miss
    ki_total["Miss_pct"] = np.where(ki_total["OrigYE_dlr"]>0, ki_total["Miss_dlr_KI"]/ki_total["OrigYE_dlr"], 0)
    # Excess thresholds
    # vs Plan: already have
    # vs Smoothed History: max(0, Inv - max(Plan, History))
    # We need total inventory available across customers per KI for this:
    inv_avail = ki_total["YTD_QTY"] + ki_total["FwdFill_QTY"] + ki_total["Excess_QTY_vsPlan"]  # approx
    # Cleaner: Total available from pool walk per KI
    pool_walk = d["pool_walk"]
    # Total inventory available for KI = sum across months of (BB Filled + NBB Filled + BB Short + NBB Short + ending pool)
    # But actually simplest: total fulfillable + leftover, but we don't track leftover well
    # Use: total possible fulfillment = lifted demand if pool is sufficient, else pool capacity
    # Approximation: (ytd + fulfillable) is what we delivered. Excess = Inv - delivered = leftover.
    # Skip "vs history" excess threshold for now; use vs Plan and vs Lifted Plan.
    # Over-Plan vs Lifted YE Plan: KI-level excess after accounting for full-year lifted plan
    # (Lifted YE = Jan-Apr orig + May-Dec lifted).
    ki_total["LiftedYE_QTY_total"] = ki_total["OrigJAD_QTY"] + ki_total["LiftedMD_QTY"]
    ki_total["LiftedYE_QTY"] = ki_total["LiftedYE_QTY_total"]  # alias for tab writers using LiftedYE_QTY
    ki_total["Excess_QTY_vsLifted"] = (ki_total["YTD_QTY"] + ki_total["FwdFill_QTY"] - ki_total["LiftedYE_QTY_total"]).clip(lower=0)
    ki_total["Excess_QTY_vsHist"] = (ki_total["Excess_QTY_vsPlan"] -
                                        np.maximum(0, ki_total["Smoothed_Hist"] - ki_total["OrigMD_QTY"])).clip(lower=0)



    # ---- Excess at Farm: pool unused at end of Dec, per KI (combined all + restricted) ----
    pool_walk_df = d["pool_walk"]
    if not pool_walk_df.empty:
        dec = pool_walk_df[pool_walk_df["Month"] == 12].copy()
        if "End_Combined" in dec.columns:
            farm_df = dec[["KI", "End_Combined"]].rename(columns={"End_Combined": "ExcessAtFarm_QTY"}).copy()
        elif "End_All" in dec.columns and "End_Restricted" in dec.columns:
            dec["ExcessAtFarm_QTY"] = dec["End_All"] + dec["End_Restricted"]
            farm_df = dec[["KI", "ExcessAtFarm_QTY"]].copy()
        else:
            farm_df = pd.DataFrame(columns=["KI", "ExcessAtFarm_QTY"])
    else:
        farm_df = pd.DataFrame(columns=["KI", "ExcessAtFarm_QTY"])

    # Average price per KI weighted by Original Plan QTY across customers
    ki_avg_price = panel.groupby("KI").apply(
        lambda g: (g["OrigYE_QTY"] * g["Price"]).sum() / g["OrigYE_QTY"].sum() if g["OrigYE_QTY"].sum() > 0 else 0.0,
        include_groups=False,
    ).reset_index()
    ki_avg_price.columns = ["KI", "AvgPrice"]
    farm_df = farm_df.merge(ki_avg_price, on="KI", how="left").fillna({"AvgPrice": 0})
    # Also expose AvgPrice on ki_total for KI-only-level tabs (Miss Summary by KI, Lift Summary by KI)
    ki_total = ki_total.merge(ki_avg_price, on="KI", how="left").fillna({"AvgPrice": 0})
    farm_df["ExcessAtFarm_$"] = farm_df["ExcessAtFarm_QTY"] * farm_df["AvgPrice"]
    # Sort descending by $
    farm_df = farm_df.sort_values("ExcessAtFarm_$", ascending=False).reset_index(drop=True)

    # ---- Customer Over-Plan per (Customer, KI): when YTD + FwdFill > OrigYE ----
    # This is what the OLD "Excess by KI" tab actually showed. Now renamed.
    ki_panel["Customer_OverPlan_QTY"] = (ki_panel["YTD_QTY"] + ki_panel["FwdFill_QTY"] - ki_panel["OrigYE_QTY"]).clip(lower=0)
    ki_panel["Customer_OverPlan_$"]   = ki_panel["Customer_OverPlan_QTY"] * ki_panel["AvgPrice"]

    # ---- Period-specific walks (per Customer x KI capped) ----
    # YE Walk metrics already exist (Miss_QTY_KI / Customer_OverPlan_QTY).

    # May-Dec Walk: compare OrigMD vs FwdFill (no YTD)
    ki_panel["Miss_MD_QTY"]      = (ki_panel["OrigMD_QTY"] - ki_panel["FwdFill_QTY"]).clip(lower=0)
    ki_panel["Miss_MD_$"]        = ki_panel["Miss_MD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["OverPlan_MD_QTY"]  = (ki_panel["FwdFill_QTY"] - ki_panel["OrigMD_QTY"]).clip(lower=0)
    ki_panel["OverPlan_MD_$"]    = ki_panel["OverPlan_MD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["FwdFill_$"]        = ki_panel["FwdFill_QTY"] * ki_panel["AvgPrice"]
    ki_panel["OrigMD_$"]         = ki_panel["OrigMD_QTY"] * ki_panel["AvgPrice"]

    # YTD Walk: compare OrigJAD vs YTD_QTY (Jan-Apr only)
    ki_panel["Miss_YTD_QTY"]     = (ki_panel["OrigJAD_QTY"] - ki_panel["YTD_QTY"]).clip(lower=0)
    ki_panel["Miss_YTD_$"]       = ki_panel["Miss_YTD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["OverPlan_YTD_QTY"] = (ki_panel["YTD_QTY"] - ki_panel["OrigJAD_QTY"]).clip(lower=0)
    ki_panel["OverPlan_YTD_$"]   = ki_panel["OverPlan_YTD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["YTD_$"]            = ki_panel["YTD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["OrigJAD_$"]        = ki_panel["OrigJAD_QTY"] * ki_panel["AvgPrice"]

    # ---- Lifted YE Plan QTY: Jan-Apr orig + May-Dec lifted (matches Original Plan QTY YE for direct comparison) ----
    ki_panel["LiftedYE_QTY"]     = ki_panel["OrigJAD_QTY"] + ki_panel["LiftedMD_QTY"]
    ki_panel["LiftedYE_$"]       = ki_panel["LiftedYE_QTY"] * ki_panel["AvgPrice"]
    panel["LiftedYE_QTY"]        = panel["OrigJAD_QTY"] + panel["LiftedMD_QTY"]
    panel["LiftedYE_$"]          = panel["LiftedYE_QTY"] * panel["Price"]

    # Aggregate the walk metrics at KI level (sum across customers) — must run after
    # the new ki_panel columns are added.
    walk_agg = ki_panel.groupby("KI").agg(
        Miss_dlr_KI_total=("Miss_dlr_KI", "sum"),
        OverPlan_dlr_KI_total=("Customer_OverPlan_$", "sum"),
        Miss_MD_dlr_total=("Miss_MD_$", "sum"),
        OverPlan_MD_dlr_total=("OverPlan_MD_$", "sum"),
        Miss_YTD_dlr_total=("Miss_YTD_$", "sum"),
        OverPlan_YTD_dlr_total=("OverPlan_YTD_$", "sum"),
        OrigMD_dlr_total=("OrigMD_$", "sum"),
        OrigJAD_dlr_total=("OrigJAD_$", "sum"),
        FwdFill_dlr_total=("FwdFill_$", "sum"),
        YTD_dlr_total=("YTD_$", "sum"),
        LiftedYE_dlr_total=("LiftedYE_$", "sum"),
    ).reset_index()
    ki_total = ki_total.merge(walk_agg, on="KI", how="left")

    # Per-item Customer Over-Plan
    panel["Customer_OverPlan_QTY"] = (panel["YTD_QTY"] + panel["FwdFill_QTY"] - panel["OrigYE_QTY"]).clip(lower=0)
    panel["Customer_OverPlan_$"]   = panel["Customer_OverPlan_QTY"] * panel["Price"]

    # ---- KI x Month aggregations for the new "by KI x Month" tabs ----
    # Miss by KI x Month: aggregate cust_alloc (Customer x KI x Item x Month) up to (KI, Month)
    cust_alloc = d["cust_alloc"]
    if not cust_alloc.empty:
        # OrigPlan_QTY must come from ORIGINAL plan only (synthesis would inflate it).
        # LiftedPlan_QTY can use plan_lifted (which includes synthesized rows by design).
        plan_lifted = d["plan_lifted"]
        if "Plan_Source" in plan_lifted.columns:
            plan_orig_only = plan_lifted[plan_lifted["Plan_Source"] == "Original"]
        else:
            plan_orig_only = plan_lifted
        plan_orig_by_km = plan_orig_only.groupby(["KI","MonthNum"])["PlanQty"].sum().reset_index().rename(columns={"PlanQty":"OrigPlan_QTY"})
        plan_lifted_by_km = plan_lifted.groupby(["KI","MonthNum"])["LiftedQty"].sum().reset_index().rename(columns={"LiftedQty":"LiftedPlan_QTY"})
        plan_by_ki_month = plan_orig_by_km.merge(plan_lifted_by_km, on=["KI","MonthNum"], how="outer").fillna(0)
        # FwdFill QTY (Stage1 + Stage2) by KI x Month from cust_alloc
        ff_by_ki_month = cust_alloc.groupby(["KI","MonthNum"]).agg(
            FwdFill_QTY=("PlanFilled","sum"),
            LiftFilled_QTY=("LiftFilled","sum"),
        ).reset_index()
        ff_by_ki_month["FwdFill_QTY"] = ff_by_ki_month["FwdFill_QTY"] + ff_by_ki_month["LiftFilled_QTY"]
        # Avg price per KI for $ conversion
        avg_price_per_ki_dict = ki_avg_price.set_index("KI")["AvgPrice"].to_dict() if "AvgPrice" in ki_avg_price.columns else {}
        ki_month = plan_by_ki_month.merge(ff_by_ki_month, on=["KI","MonthNum"], how="left").fillna(0)
        ki_month["Lift_QTY"] = (ki_month["LiftedPlan_QTY"] - ki_month["OrigPlan_QTY"]).clip(lower=0)
        # Miss for FWD months only (May-Dec): max(0, OrigPlan - FwdFill)
        # For Jan-Apr, miss is computed against YTD actuals — handled in miss-by-KIx-month tab writer
        ki_month["AvgPrice"] = ki_month["KI"].map(avg_price_per_ki_dict).fillna(0)
        ki_month["OrigPlan_$"] = ki_month["OrigPlan_QTY"] * ki_month["AvgPrice"]
        ki_month["LiftedPlan_$"] = ki_month["LiftedPlan_QTY"] * ki_month["AvgPrice"]
        ki_month["FwdFill_$"] = ki_month["FwdFill_QTY"] * ki_month["AvgPrice"]
        ki_month["Lift_$"] = ki_month["Lift_QTY"] * ki_month["AvgPrice"]
    else:
        ki_month = pd.DataFrame()

    # ---- Item-level summary for Customer Miss Detail tab (already has panel; just keep it accessible) ----
    # No new aggregations needed; "panel" is already at (Customer x Item) level

    # ---- Lift by Customer x Item: aggregate plan_lifted by (Customer, Item) summed across months ----
    plan_lifted = d["plan_lifted"]
    fwd_pl = plan_lifted[plan_lifted["MonthNum"].isin(FWD_MONTHS)].copy()
    # OrigMD_QTY must come from ORIGINAL plan only (synthesis would inflate it)
    if "Plan_Source" in fwd_pl.columns:
        fwd_pl_orig = fwd_pl[fwd_pl["Plan_Source"] == "Original"]
    else:
        fwd_pl_orig = fwd_pl
    orig_md_cust_item = fwd_pl_orig.groupby(["Customer","Item Num","KI"])["PlanQty"].sum().reset_index().rename(columns={"PlanQty":"OrigMD_QTY"})
    lifted_md_cust_item = fwd_pl.groupby(["Customer","Item Num","KI"])["LiftedQty"].sum().reset_index().rename(columns={"LiftedQty":"LiftedMD_QTY"})
    lift_by_cust_item = orig_md_cust_item.merge(lifted_md_cust_item, on=["Customer","Item Num","KI"], how="outer").fillna(0)
    lift_by_cust_item["Lift_QTY"] = (lift_by_cust_item["LiftedMD_QTY"] - lift_by_cust_item["OrigMD_QTY"]).clip(lower=0)
    # Add Item Description and Avg Price per (Customer, Item Num) — use panel
    lift_by_cust_item = lift_by_cust_item.merge(
        panel[["Customer","Item Num","Item Desc","Price"]].drop_duplicates(subset=["Customer","Item Num"]),
        on=["Customer","Item Num"], how="left"
    )
    lift_by_cust_item["Lift_$"] = lift_by_cust_item["Lift_QTY"] * lift_by_cust_item["Price"].fillna(0)
    lift_by_cust_item["Lift_pct"] = np.where(lift_by_cust_item["OrigMD_QTY"]>0,
                                             lift_by_cust_item["Lift_QTY"]/lift_by_cust_item["OrigMD_QTY"], 0)

    # ============================================================================
    # GROW DAYS POST-PROCESS — classify each (Customer, KI, Month) miss as
    # RECOVERABLE (grow days fits) or LOCKED IN. KI-level recovery: if any
    # item in the customer's KI has fitting grow days, the entire KI miss for
    # that month is recoverable via within-KI substitution.
    # ============================================================================
    grow_times = d.get("grow_times")
    cust_alloc = d["cust_alloc"]
    plan = d["plan"]
    SNAP_DATE = M.SNAP_DATE
    import datetime
    def _days_to_month_end(month_num: int, year: int = 2026):
        last_day_map = {5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
        end = datetime.date(year, month_num, last_day_map.get(month_num, 31))
        snap = SNAP_DATE.date()
        return (end - snap).days

    # Build per-item Org Code lookup from plan (plan rows have Org Code; each NOR CAL
    # plan row is tagged WIN or BRA per CEO direction)
    plan_orig = d["plan"]  # post-load, has Org Code column
    item_org = plan_orig.groupby("Item Num")["Org Code"].first().to_dict()

    # Cascade-resolve grow days per item
    if grow_times is not None and not grow_times.empty:
        # Tier 1: exact (Item Num, Org) match
        gt_lookup = {(r["Org Code"], r["Item Num"]): float(r["GrowDays"]) for _, r in grow_times.iterrows()}
        # Tier 2: KI MIN — within each Org Code, MIN grow days per KI
        gt_with_ki = grow_times.merge(items[["Item Num","KI"]].drop_duplicates(subset=["Item Num"]),
                                       on="Item Num", how="left")
        ki_min_per_org = gt_with_ki.dropna(subset=["KI"]).groupby(["Org Code","KI"])["GrowDays"].min().to_dict()
        # Tier 3: Genus + Size MIN
        gt_with_gs = grow_times.merge(items[["Item Num","Genus","Size"]].drop_duplicates(subset=["Item Num"]),
                                       on="Item Num", how="left")
        gs_min_per_org = gt_with_gs.dropna(subset=["Genus","Size"]).groupby(["Org Code","Genus","Size"])["GrowDays"].min().to_dict()
        # Tier 4: Size MIN
        size_min_per_org = gt_with_gs.dropna(subset=["Size"]).groupby(["Org Code","Size"])["GrowDays"].min().to_dict()
    else:
        gt_lookup = {}
        ki_min_per_org = {}
        gs_min_per_org = {}
        size_min_per_org = {}

    item_meta_lookup = items[["Item Num","KI","Genus","Size","Item Desc"]].drop_duplicates(subset=["Item Num"]).set_index("Item Num").to_dict("index")

    def resolve_grow_days(item_num):
        """Return (grow_days, source_label) using cascade fallback. None if no match."""
        org = item_org.get(item_num)
        if not org:
            return None, "no Org Code"
        # Tier 1
        if (org, item_num) in gt_lookup:
            return gt_lookup[(org, item_num)], "exact match"
        # Get item metadata for Tier 2-4
        meta = item_meta_lookup.get(item_num, {})
        ki = meta.get("KI")
        genus = meta.get("Genus")
        size = meta.get("Size")
        # Tier 2: KI MIN
        if ki and (org, ki) in ki_min_per_org:
            return ki_min_per_org[(org, ki)], f"KI fallback: {ki}"
        # Tier 3: Genus + Size
        if genus and size and (org, genus, size) in gs_min_per_org:
            return gs_min_per_org[(org, genus, size)], f"Genus+Size fallback: {genus}/{size}"
        # Tier 4: Size only
        if size and (org, size) in size_min_per_org:
            return size_min_per_org[(org, size)], f"Size fallback: {size}"
        return None, "no match"

    # Pre-resolve grow days for all items in the universe (for performance)
    item_grow_days = {}
    for item_num in items["Item Num"].dropna().unique():
        gd, src = resolve_grow_days(item_num)
        item_grow_days[item_num] = (gd, src)

    # Days to month-end lookup
    days_to_me = {m: _days_to_month_end(m) for m in FWD_MONTHS}

    # Process each (Customer, KI, Month): if KI has any growable item with grow days
    # fitting that month, the full KI miss is recoverable.
    recovery_rows = []
    if not cust_alloc.empty and grow_times is not None and not grow_times.empty:
        ca = cust_alloc.copy()
        ca["GrowDays"] = ca["Item Num"].map(lambda x: item_grow_days.get(x, (None, ""))[0])
        ca["GrowSource"] = ca["Item Num"].map(lambda x: item_grow_days.get(x, (None, "no match"))[1])
        ca["DaysToMonthEnd"] = ca["MonthNum"].map(days_to_me)
        ca["FitsThisMonth"] = (ca["GrowDays"].notna() & (ca["GrowDays"] <= ca["DaysToMonthEnd"]))

        # Per (Customer, KI, Month): aggregate
        grouped = ca.groupby(["Customer","KI","MonthNum"], as_index=False).agg(
            KI_miss_qty=("PlanShort","sum"),
            any_fit=("FitsThisMonth","any"),
        )
        # Only KI/Month combos with miss > 0 AND at least one growable item
        recoverable_groups = grouped[(grouped["KI_miss_qty"] > 0) & (grouped["any_fit"])].copy()

        for _, g in recoverable_groups.iterrows():
            cust, ki, month = g["Customer"], g["KI"], int(g["MonthNum"])
            ki_miss_qty = float(g["KI_miss_qty"])
            # Find the production target (item with shortest grow days that fits this month)
            sub = ca[(ca["Customer"]==cust) & (ca["KI"]==ki) & (ca["MonthNum"]==month) & (ca["FitsThisMonth"])]
            if sub.empty: continue
            # Pick item with shortest grow days
            target = sub.sort_values("GrowDays").iloc[0]
            target_item = target["Item Num"]
            target_gd = float(target["GrowDays"])
            target_src = target["GrowSource"]
            # Items being substituted for (other items in KI for this customer/month with miss > 0)
            others = ca[(ca["Customer"]==cust) & (ca["KI"]==ki) & (ca["MonthNum"]==month) &
                          (ca["PlanShort"] > 0) & (ca["Item Num"] != target_item)]
            substitutes = list(others["Item Num"].unique())
            # Compute $ value using avg price for the KI (weighted)
            ki_panel_row = ki_panel[(ki_panel["Customer"]==cust) & (ki_panel["KI"]==ki)]
            avg_price = float(ki_panel_row["AvgPrice"].iloc[0]) if not ki_panel_row.empty else 0.0
            recovery_rows.append({
                "Customer": cust,
                "KI": ki,
                "MonthNum": month,
                "ProductionTargetItem": target_item,
                "ProductionTargetItemDesc": item_meta_lookup.get(target_item, {}).get("Item Desc", ""),
                "RecoverableQTY": ki_miss_qty,
                "RecoverableDlr": ki_miss_qty * avg_price,
                "MinGrowDays": target_gd,
                "DaysToMonthEnd": days_to_me.get(month, 0),
                "SlackDays": days_to_me.get(month, 0) - target_gd,
                "OrgCode": item_org.get(target_item, ""),
                "GrowDaysSource": target_src,
                "SubstitutesFor": ", ".join(substitutes) if substitutes else "(none — single item KI)",
            })
            # Zero out PlanShort for ALL items in this (Customer, KI, Month) — substitution covers
            cust_alloc.loc[(cust_alloc["Customer"]==cust) & (cust_alloc["KI"]==ki) &
                            (cust_alloc["MonthNum"]==month), "PlanShort"] = 0

        # Update d["cust_alloc"] in place so downstream re-reads see the modified values
        d["cust_alloc"] = cust_alloc

    production_recovery = pd.DataFrame(recovery_rows).sort_values("RecoverableDlr", ascending=False).reset_index(drop=True) if recovery_rows else pd.DataFrame(columns=["Customer","KI","MonthNum","ProductionTargetItem","ProductionTargetItemDesc","RecoverableQTY","RecoverableDlr","MinGrowDays","DaysToMonthEnd","SlackDays","OrgCode","GrowDaysSource","SubstitutesFor"])
    log(f"  Production Recovery: {len(production_recovery):,} (Customer, KI, Month) actions; total recoverable ${production_recovery['RecoverableDlr'].sum() if not production_recovery.empty else 0:,.0f}")

    # ============================================================================
    # Recompute downstream aggregations after miss reduction.
    # The recovery zeros out PlanShort (Miss). For the walks (YE / YTD / Forward)
    # to mathematically reconcile (YE Net = YTD Net + Forward Net per Cust x KI),
    # the recovered QTY must ALSO be ADDED to FwdFill_QTY so that
    # FwdFill_with_recovery + YTD = OrigYE - LockedInMiss.
    # ============================================================================
    # 1. Build recovery aggregations
    if not production_recovery.empty:
        rec_by_cust_ki = production_recovery.groupby(["Customer","KI"])["RecoverableQTY"].sum().reset_index()
        rec_by_cust_ki = rec_by_cust_ki.rename(columns={"RecoverableQTY":"RecoveredQTY_CK"})
        rec_by_cust_ki_month = production_recovery.groupby(["Customer","KI","MonthNum"])["RecoverableQTY"].sum().reset_index()
        rec_by_cust_ki_month = rec_by_cust_ki_month.rename(columns={"RecoverableQTY":"RecoveredQTY_CKM"})
    else:
        rec_by_cust_ki = pd.DataFrame(columns=["Customer","KI","RecoveredQTY_CK"])
        rec_by_cust_ki_month = pd.DataFrame(columns=["Customer","KI","MonthNum","RecoveredQTY_CKM"])

    # 2. Allocate KI-level recovery to items proportionally by plan share within KI
    #    (so each item's panel.FwdFill_QTY reflects its proportional coverage from the
    #    KI-level production action — recovery covers all items in the KI via substitution)
    if not rec_by_cust_ki.empty:
        plan_share = plan_lifted.groupby(["Customer","KI","Item Num"])["PlanQty"].sum().reset_index()
        plan_share_total = plan_share.groupby(["Customer","KI"])["PlanQty"].sum().reset_index().rename(columns={"PlanQty":"KI_PlanQty"})
        plan_share = plan_share.merge(plan_share_total, on=["Customer","KI"], how="left")
        plan_share["share"] = np.where(plan_share["KI_PlanQty"] > 0,
                                          plan_share["PlanQty"] / plan_share["KI_PlanQty"], 0)
        plan_share = plan_share.merge(rec_by_cust_ki, on=["Customer","KI"], how="inner")
        plan_share["RecoveredQTY_item"] = plan_share["share"] * plan_share["RecoveredQTY_CK"]
        item_recovery = plan_share[["Customer","Item Num","RecoveredQTY_item"]].copy()
    else:
        item_recovery = pd.DataFrame(columns=["Customer","Item Num","RecoveredQTY_item"])

    # 3. Re-aggregate per (Customer, Item) miss from cust_alloc post-recovery PlanShort
    pf_g_new = cust_alloc.groupby(["Customer","Item Num"]).agg(
        PlanShort_new=("PlanShort","sum"),
    ).reset_index()

    # Update panel: Miss + FwdFill + downstream
    panel = panel.drop(columns=["Miss_QTY","Miss_$","Customer_OverPlan_QTY","Customer_OverPlan_$"], errors="ignore")
    panel = panel.merge(pf_g_new, on=["Customer","Item Num"], how="left").fillna({"PlanShort_new": 0})
    panel = panel.merge(item_recovery, on=["Customer","Item Num"], how="left").fillna({"RecoveredQTY_item": 0})
    panel["FwdFill_QTY"] = panel["FwdFill_QTY"] + panel["RecoveredQTY_item"]
    panel["FwdFill_$"]   = panel["FwdFill_QTY"] * panel["Price"]
    panel["Miss_QTY"] = panel["PlanShort_new"]
    panel["Miss_$"]   = panel["Miss_QTY"] * panel["Price"]
    panel["Customer_OverPlan_QTY"] = (panel["YTD_QTY"] + panel["FwdFill_QTY"] - panel["OrigYE_QTY"]).clip(lower=0)
    panel["Customer_OverPlan_$"]   = panel["Customer_OverPlan_QTY"] * panel["Price"]
    panel["Achieved_QTY"] = panel["YTD_QTY"] + panel["FwdFill_QTY"]
    panel["Achieved_QTY_Capped"] = np.minimum(panel["Achieved_QTY"], panel["OrigYE_QTY"])
    panel["Achieved_$"] = panel["Achieved_QTY_Capped"] * panel["Price"]
    panel["Excess_QTY_vsPlan"] = (panel["Achieved_QTY"] - panel["OrigYE_QTY"]).clip(lower=0)
    panel["Excess_$_vsPlan"] = panel["Excess_QTY_vsPlan"] * panel["Price"]
    panel = panel.drop(columns=["PlanShort_new","RecoveredQTY_item"])

    # 4. Update ki_panel: Miss + FwdFill + downstream walk metrics
    ki_panel = ki_panel.drop(columns=["Miss_QTY_KI","Miss_dlr_KI","Miss_MD_QTY","Miss_MD_$","Miss_YTD_QTY","Miss_YTD_$",
                                        "Customer_OverPlan_QTY","Customer_OverPlan_$","FwdFill_$",
                                        "OverPlan_MD_QTY","OverPlan_MD_$","Excess_QTY_vsPlan","Excess_dlr_vsPlan"], errors="ignore")
    # Update ki_panel.FwdFill_QTY from updated panel
    fwdfill_agg = panel.groupby(["Customer","KI"])["FwdFill_QTY"].sum().reset_index().rename(columns={"FwdFill_QTY":"FwdFill_QTY_new"})
    ki_panel = ki_panel.drop(columns=["FwdFill_QTY"], errors="ignore")
    ki_panel = ki_panel.merge(fwdfill_agg, on=["Customer","KI"], how="left").fillna({"FwdFill_QTY_new":0})
    ki_panel = ki_panel.rename(columns={"FwdFill_QTY_new":"FwdFill_QTY"})
    # Compute (Cust, KI) CAPPED miss/over (not per-item sum) for proper walk reconciliation
    ki_panel["Miss_QTY_KI"] = (ki_panel["OrigYE_QTY"] - ki_panel["YTD_QTY"] - ki_panel["FwdFill_QTY"]).clip(lower=0)
    ki_panel["Miss_dlr_KI"] = ki_panel["Miss_QTY_KI"] * ki_panel["AvgPrice"]
    ki_panel["Excess_QTY_vsPlan"] = (ki_panel["YTD_QTY"] + ki_panel["FwdFill_QTY"] - ki_panel["OrigYE_QTY"]).clip(lower=0)
    ki_panel["Excess_dlr_vsPlan"] = ki_panel["Excess_QTY_vsPlan"] * ki_panel["AvgPrice"]
    ki_panel["FwdFill_$"] = ki_panel["FwdFill_QTY"] * ki_panel["AvgPrice"]
    ki_panel["Miss_MD_QTY"]      = (ki_panel["OrigMD_QTY"] - ki_panel["FwdFill_QTY"]).clip(lower=0)
    ki_panel["Miss_MD_$"]        = ki_panel["Miss_MD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["Miss_YTD_QTY"]     = (ki_panel["OrigJAD_QTY"] - ki_panel["YTD_QTY"]).clip(lower=0)
    ki_panel["Miss_YTD_$"]       = ki_panel["Miss_YTD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["OverPlan_MD_QTY"]  = (ki_panel["FwdFill_QTY"] - ki_panel["OrigMD_QTY"]).clip(lower=0)
    ki_panel["OverPlan_MD_$"]    = ki_panel["OverPlan_MD_QTY"] * ki_panel["AvgPrice"]
    ki_panel["Customer_OverPlan_QTY"] = (ki_panel["YTD_QTY"] + ki_panel["FwdFill_QTY"] - ki_panel["OrigYE_QTY"]).clip(lower=0)
    ki_panel["Customer_OverPlan_$"]   = ki_panel["Customer_OverPlan_QTY"] * ki_panel["AvgPrice"]

    # 5. Update ki_total: re-aggregate from updated ki_panel
    ki_total = ki_total.drop(columns=["Miss_QTY_KI","Miss_dlr_KI","Miss_pct","FwdFill_QTY","Excess_QTY_vsPlan","Excess_dlr_vsPlan",
                                       "Miss_dlr_KI_total","OverPlan_dlr_KI_total","Miss_MD_dlr_total","OverPlan_MD_dlr_total",
                                       "FwdFill_dlr_total"], errors="ignore")
    ki_total_new = ki_panel.groupby("KI").agg(
        Miss_QTY_KI=("Miss_QTY_KI","sum"),
        FwdFill_QTY=("FwdFill_QTY","sum"),
        Excess_QTY_vsPlan=("Excess_QTY_vsPlan","sum"),
        Miss_dlr_KI_total=("Miss_dlr_KI","sum"),
        OverPlan_dlr_KI_total=("Customer_OverPlan_$","sum"),
        Miss_MD_dlr_total=("Miss_MD_$","sum"),
        OverPlan_MD_dlr_total=("OverPlan_MD_$","sum"),
        FwdFill_dlr_total=("FwdFill_$","sum"),
    ).reset_index()
    ki_total = ki_total.merge(ki_total_new, on="KI", how="left").fillna(0)
    ki_total["Miss_dlr_KI"] = ki_total["Miss_QTY_KI"] * ki_total["AvgPrice"]
    ki_total["Excess_dlr_vsPlan"] = ki_total["Excess_QTY_vsPlan"] * ki_total["AvgPrice"]
    ki_total["Miss_pct"] = np.where(ki_total["OrigYE_dlr"]>0,
                                     ki_total["Miss_dlr_KI"]/ki_total["OrigYE_dlr"], 0)
    ki_total["Achieved_QTY"] = ki_total["YTD_QTY"] + ki_total["FwdFill_QTY"]

    # 6. Update ki_month for Miss by KI x Month tab — add KI x Month recovery to FwdFill
    if not rec_by_cust_ki_month.empty and "ki_month" in dir() and not ki_month.empty:
        rec_ki_month = rec_by_cust_ki_month.groupby(["KI","MonthNum"])["RecoveredQTY_CKM"].sum().reset_index()
        ki_month = ki_month.merge(rec_ki_month, on=["KI","MonthNum"], how="left").fillna({"RecoveredQTY_CKM":0})
        ki_month["FwdFill_QTY"] = ki_month["FwdFill_QTY"] + ki_month["RecoveredQTY_CKM"]
        ki_month["FwdFill_$"]   = ki_month["FwdFill_QTY"] * ki_month["AvgPrice"]
        ki_month = ki_month.drop(columns=["RecoveredQTY_CKM"])

    # ============================================================================
    # VERIFICATION: confirm walks reconcile (YE Net == YTD Net + Forward Net per Cust)
    # ============================================================================
    walk_check = ki_panel.groupby("Customer").agg(
        OrigYE=("OrigYE_QTY","sum"),
        OrigJAD=("OrigJAD_QTY","sum"),
        OrigMD=("OrigMD_QTY","sum"),
        YTD=("YTD_QTY","sum"),
        FwdFill=("FwdFill_QTY","sum"),
    ).reset_index()
    walk_check["YE_Net"]      = (walk_check["YTD"] + walk_check["FwdFill"]) - walk_check["OrigYE"]
    walk_check["YTD_Net"]     = walk_check["YTD"] - walk_check["OrigJAD"]
    walk_check["Forward_Net"] = walk_check["FwdFill"] - walk_check["OrigMD"]
    walk_check["Sum_Net"]     = walk_check["YTD_Net"] + walk_check["Forward_Net"]
    walk_check["Diff"]        = walk_check["YE_Net"] - walk_check["Sum_Net"]
    max_diff = walk_check["Diff"].abs().max()
    log(f"  Walk reconciliation: max |YE_Net - (YTD_Net + Forward_Net)| per customer = {max_diff:.2f} QTY")
    if max_diff > 1.0:
        log(f"  WARN: walks may not reconcile cleanly (max diff > 1 unit)")


    # ============================================================================
    # KI LIFECYCLE VIEW — full picture per KI (no customer dimension)
    # YTD Plan / YTD Actual / Starting INV (R + NR) / Forward Plan / Forward Actual /
    # Forward Inflows / Ending INV. Two tabs: Units (QTY) + Dollars ($).
    # Pool conservation: Starting (R + NR) + Forward Inflows = Forward Actual + Ending INV.
    # ============================================================================
    pools = d["pools"]
    pool_walk_df = d["pool_walk"]
    # Starting INV at KI level
    start_restricted = pools[pools["PoolCat"]=="restricted_then_all"].groupby("KI")["PoolQty"].sum().reset_index().rename(columns={"PoolQty":"StartINV_Restricted"})
    start_nonrestricted = pools[pools["PoolCat"]=="all_initial"].groupby("KI")["PoolQty"].sum().reset_index().rename(columns={"PoolQty":"StartINV_NonRestricted"})

    # Forward Inflows: "fresh" inflow during May-Dec
    # = (Total - Avail) for SS/GS grads (the non-Available portion that finishes growing)
    # + Total for SN/GN at-ready (whole lot enters fresh)
    grads_fresh = pools[pools["PoolCat"]=="restricted_then_all"].copy()
    grads_fresh["FreshInflow"] = grads_fresh["Total QTY"] - grads_fresh["Available QTY"]
    grads_fresh = grads_fresh[grads_fresh["FreshInflow"] > 0]
    grads_inflow = grads_fresh.groupby("KI")["FreshInflow"].sum().reset_index().rename(columns={"FreshInflow":"Inflow_Grads"})
    sngn_inflow = pools[pools["PoolCat"]=="sngn_inflow_at_ready"].groupby("KI")["PoolQty"].sum().reset_index().rename(columns={"PoolQty":"Inflow_SNGN"})

    # YTD Plan + YTD Actual at KI level (sum across customers)
    ytd_plan_ki = ki_panel.groupby("KI")["OrigJAD_QTY"].sum().reset_index().rename(columns={"OrigJAD_QTY":"YTDPlan_QTY"})
    ytd_actual_ki = ki_panel.groupby("KI")["YTD_QTY"].sum().reset_index().rename(columns={"YTD_QTY":"YTDActual_QTY"})

    # Forward Plan + Forward Actual at KI level
    # FwdActual must include off-plan customer Stage 2 fills (Southeast, MLC, Other) which
    # are NOT in ki_panel. Otherwise lifecycle conservation breaks for KIs with off-plan demand.
    fwd_plan_ki = ki_panel.groupby("KI")["OrigMD_QTY"].sum().reset_index().rename(columns={"OrigMD_QTY":"FwdPlan_QTY"})
    fwd_actual_planned_ki = ki_panel.groupby("KI")["FwdFill_QTY"].sum().reset_index().rename(columns={"FwdFill_QTY":"FwdActual_Planned_QTY"})
    offplan_alloc_df = d.get("offplan_alloc", pd.DataFrame())
    if not offplan_alloc_df.empty:
        offplan_filled_ki = offplan_alloc_df.groupby("KI")["Stage2Filled"].sum().reset_index().rename(columns={"Stage2Filled":"FwdActual_OffPlan_QTY"})
    else:
        offplan_filled_ki = pd.DataFrame(columns=["KI","FwdActual_OffPlan_QTY"])
    fwd_actual_ki = fwd_actual_planned_ki.merge(offplan_filled_ki, on="KI", how="left").fillna({"FwdActual_OffPlan_QTY": 0})
    fwd_actual_ki["FwdActual_QTY"] = fwd_actual_ki["FwdActual_Planned_QTY"] + fwd_actual_ki["FwdActual_OffPlan_QTY"]
    fwd_actual_ki = fwd_actual_ki[["KI","FwdActual_QTY"]]

    # Ending INV: pool at end of Dec from pool_walk
    if not pool_walk_df.empty and "End_Combined" in pool_walk_df.columns:
        ending_inv = pool_walk_df[pool_walk_df["Month"]==12][["KI","End_Combined"]].copy().rename(columns={"End_Combined":"EndingINV"})
    else:
        ending_inv = pd.DataFrame(columns=["KI","EndingINV"])

    # Recovery inflows per KI (Source 1 fix - Task 48; Method B per QA #67)
    # Recovery QTY for Lifecycle Units tab; Recovery $ uses Cust × KI Method B from production_recovery.RecoverableDlr
    if not production_recovery.empty:
        recovery_per_ki = production_recovery.groupby("KI").agg(
            Inflow_Recovery=("RecoverableQTY","sum"),
            Inflow_Recovery_dlr_B=("RecoverableDlr","sum"),
        ).reset_index()
    else:
        recovery_per_ki = pd.DataFrame(columns=["KI","Inflow_Recovery","Inflow_Recovery_dlr_B"])

    # Method B customer-aware $ aggregation per KI (QA #64).
    # ki_panel already has per-(Customer × KI) dollars using customer-aware AvgPrice.
    # Sum these to KI level for use on the Lifecycle ($) tab.
    custaware_dlrs_ki = ki_panel.groupby("KI").agg(
        YTDPlan_dlr_B=("OrigJAD_$","sum"),
        YTDActual_dlr_B_planned=("YTD_$","sum"),
        FwdPlan_dlr_B=("OrigMD_$","sum"),
        FwdActual_planned_dlr_B=("FwdFill_$","sum"),
    ).reset_index()
    # QA #65: add off-plan customer YTD shipments (Southeast/MLC/Other) per KI
    ytd_raw = d.get("ytd")
    plan_custs_set = set(ki_panel["Customer"].unique())
    items_for_ki_map = d.get("items")
    if ytd_raw is not None and not ytd_raw.empty:
        item_to_ki_map = items_for_ki_map.drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict() if items_for_ki_map is not None else {}
        ytd_op = ytd_raw[~ytd_raw["Customer"].isin(plan_custs_set) & ytd_raw["Customer"].notna()].copy()
        ytd_op["KI"] = ytd_op["Item Num"].map(item_to_ki_map)
        ytd_op = ytd_op.dropna(subset=["KI"])
        if "Revenue" in ytd_op.columns:
            ytd_op_dlr_ki = ytd_op.groupby("KI")["Revenue"].sum().reset_index().rename(columns={"Revenue":"YTDActual_dlr_B_offplan"})
        else:
            ki_avg = ki_total.set_index("KI")["AvgPrice"].to_dict()
            ytd_op["AvgPrice"] = ytd_op["KI"].map(ki_avg).fillna(0)
            ytd_op["YTD_$"] = ytd_op["Qty"] * ytd_op["AvgPrice"]
            ytd_op_dlr_ki = ytd_op.groupby("KI")["YTD_$"].sum().reset_index().rename(columns={"YTD_$":"YTDActual_dlr_B_offplan"})
        custaware_dlrs_ki = custaware_dlrs_ki.merge(ytd_op_dlr_ki, on="KI", how="left").fillna({"YTDActual_dlr_B_offplan":0})
    else:
        custaware_dlrs_ki["YTDActual_dlr_B_offplan"] = 0
    custaware_dlrs_ki["YTDActual_dlr_B"] = custaware_dlrs_ki["YTDActual_dlr_B_planned"] + custaware_dlrs_ki["YTDActual_dlr_B_offplan"]
    # Add off-plan FwdFill dollars (Southeast/MLC/Other Stage 2 fills)
    offplan_alloc_df = d.get("offplan_alloc", pd.DataFrame())
    if not offplan_alloc_df.empty:
        # Need to convert offplan QTY to $ — use KI's avg price as proxy since off-plan customers
        # don't have a Cust×KI AvgPrice in ki_panel. Off-plan share is small.
        op_with_price = offplan_alloc_df.merge(
            ki_total[["KI","AvgPrice"]], on="KI", how="left"
        ).fillna({"AvgPrice": 0})
        op_with_price["FwdFill_$_offplan"] = op_with_price["Stage2Filled"] * op_with_price["AvgPrice"]
        offplan_dlr_ki = op_with_price.groupby("KI")["FwdFill_$_offplan"].sum().reset_index()
        custaware_dlrs_ki = custaware_dlrs_ki.merge(offplan_dlr_ki, on="KI", how="left").fillna({"FwdFill_$_offplan":0})
        custaware_dlrs_ki["FwdActual_dlr_B"] = custaware_dlrs_ki["FwdActual_planned_dlr_B"] + custaware_dlrs_ki["FwdFill_$_offplan"]
    else:
        custaware_dlrs_ki["FwdActual_dlr_B"] = custaware_dlrs_ki["FwdActual_planned_dlr_B"]

    # Combine into one DataFrame at KI level
    ki_lifecycle = ki_total[["KI","AvgPrice"]].copy()
    for df_to_merge in [ytd_plan_ki, ytd_actual_ki, start_restricted, start_nonrestricted,
                         fwd_plan_ki, fwd_actual_ki, grads_inflow, sngn_inflow, recovery_per_ki, ending_inv,
                         custaware_dlrs_ki]:
        ki_lifecycle = ki_lifecycle.merge(df_to_merge, on="KI", how="left")
    ki_lifecycle = ki_lifecycle.fillna(0)
    # Pool inflows = grads + sngn; Recovery shown separately (Source 1 fix)
    ki_lifecycle["FwdInflows_Pool"] = ki_lifecycle["Inflow_Grads"] + ki_lifecycle["Inflow_SNGN"]
    ki_lifecycle["FwdInflows_Recovery"] = ki_lifecycle["Inflow_Recovery"]
    ki_lifecycle["FwdInflows"] = ki_lifecycle["FwdInflows_Pool"] + ki_lifecycle["FwdInflows_Recovery"]
    # Sort by Forward Plan $ desc
    ki_lifecycle["FwdPlan_$"] = ki_lifecycle["FwdPlan_QTY"] * ki_lifecycle["AvgPrice"]
    ki_lifecycle = ki_lifecycle.sort_values("FwdPlan_$", ascending=False).reset_index(drop=True)
    log(f"  KI Lifecycle: {len(ki_lifecycle):,} KIs computed (pool-walk-eligible)")

    # Source 2 fix (Task 50): augment ki_lifecycle with stranded KIs (in pools but not pool_walk)
    items_df = d["items"]
    ytd_raw = d["ytd"]
    item_to_ki = items_df[["Item Num","KI"]].drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict()
    ytd_with_ki = ytd_raw.copy()
    ytd_with_ki["KI"] = ytd_with_ki["Item Num"].map(item_to_ki)

    kis_in_walk = set(pool_walk_df["KI"].unique()) if not pool_walk_df.empty else set()
    kis_in_pool = set(pools["KI"].unique())
    stranded_kis_set = kis_in_pool - kis_in_walk

    # PART A: KIs that ARE already in ki_lifecycle (via ki_total) — just update Ending INV
    # to reflect that nothing was consumed. Their Start INV and Inflows are already populated
    # via the merges above; only EndingINV is wrong (defaulted to 0 because no pool_walk row).
    in_lifecycle_already = set(ki_lifecycle["KI"].unique())
    stranded_in_lifecycle = stranded_kis_set & in_lifecycle_already
    if stranded_in_lifecycle:
        mask = ki_lifecycle["KI"].isin(stranded_in_lifecycle)
        ki_lifecycle.loc[mask, "EndingINV"] = (
            ki_lifecycle.loc[mask, "StartINV_Restricted"]
            + ki_lifecycle.loc[mask, "StartINV_NonRestricted"]
            + ki_lifecycle.loc[mask, "FwdInflows_Pool"]
            + ki_lifecycle.loc[mask, "FwdInflows_Recovery"]
        )

    # PART B: KIs in pools but NOT in ki_lifecycle (truly new) — synthesize fresh rows
    stranded_new = stranded_kis_set - in_lifecycle_already
    stranded_rows = []
    ki_avg_price_lookup = ki_total.set_index("KI")["AvgPrice"].to_dict()
    for ki in stranded_new:
        p = pools[pools["KI"] == ki]
        start_all = float(p[p["PoolCat"] == "all_initial"]["PoolQty"].sum())
        start_res = float(p[p["PoolCat"] == "restricted_then_all"]["PoolQty"].sum())
        grad_total = float(p[p["PoolCat"] == "restricted_then_all"]["Total QTY"].sum())
        grad_avail = float(p[p["PoolCat"] == "restricted_then_all"]["Available QTY"].sum())
        sngn = float(p[p["PoolCat"] == "sngn_inflow_at_ready"]["PoolQty"].sum())
        pool_inflow = max(0.0, grad_total - grad_avail) + sngn
        ytd_actual_qty = float(ytd_with_ki[ytd_with_ki["KI"] == ki]["Qty"].sum())
        avg_price = ki_avg_price_lookup.get(ki, 0.0)
        if avg_price == 0.0:
            ki_items = items_df[items_df["KI"] == ki]["Item Num"].tolist()
            prices_d = d.get("prices", {})
            ki_prices = [prices_d.get(it, 0.0) for it in ki_items if prices_d.get(it, 0.0) > 0]
            avg_price = sum(ki_prices) / len(ki_prices) if ki_prices else 0.0
        ending_inv_qty = start_all + start_res + pool_inflow
        stranded_rows.append({
            "KI": ki, "AvgPrice": avg_price, "YTDPlan_QTY": 0.0,
            "YTDActual_QTY": ytd_actual_qty,
            "StartINV_Restricted": start_res, "StartINV_NonRestricted": start_all,
            "FwdPlan_QTY": 0.0, "FwdActual_QTY": 0.0,
            "Inflow_Grads": max(0.0, grad_total - grad_avail), "Inflow_SNGN": sngn,
            "Inflow_Recovery": 0.0, "EndingINV": ending_inv_qty,
            "FwdInflows_Pool": pool_inflow, "FwdInflows_Recovery": 0.0,
            "FwdInflows": pool_inflow, "FwdPlan_$": 0.0,
        })

    if stranded_rows:
        stranded_df = pd.DataFrame(stranded_rows)
        ki_lifecycle = pd.concat([ki_lifecycle, stranded_df], ignore_index=True)
    ki_lifecycle = ki_lifecycle.sort_values("FwdPlan_$", ascending=False).reset_index(drop=True)
    log(f"  KI Lifecycle: stranded handling — updated {len(stranded_in_lifecycle)} existing rows + added {len(stranded_new)} new rows")

    # Track for Build Health
    if stranded_kis_set:
        # Compute stranded totals across ALL stranded (both Part A and Part B)
        stranded_mask = ki_lifecycle["KI"].isin(stranded_kis_set)
        sub = ki_lifecycle[stranded_mask]
        _stranded_count = len(stranded_kis_set)
        _stranded_units = float(sub["StartINV_Restricted"].sum() + sub["StartINV_NonRestricted"].sum() + sub["FwdInflows_Pool"].sum())
        _stranded_dollars = float((sub["StartINV_Restricted"] + sub["StartINV_NonRestricted"] + sub["FwdInflows_Pool"]) @ sub["AvgPrice"])
    else:
        _stranded_count = 0
        _stranded_units = 0.0
        _stranded_dollars = 0.0

    # Source 2 fix (continued): augment farm_df (Excess by KI) with stranded KIs
    if stranded_kis_set:
        existing_farm_kis = set(farm_df["KI"].unique()) if not farm_df.empty else set()
        new_excess_rows = []
        ki_avg_price_lookup = ki_total.set_index("KI")["AvgPrice"].to_dict()
        for ki in stranded_kis_set:
            if ki in existing_farm_kis:
                continue
            p = pools[pools["KI"] == ki]
            start_all = float(p[p["PoolCat"] == "all_initial"]["PoolQty"].sum())
            start_res = float(p[p["PoolCat"] == "restricted_then_all"]["PoolQty"].sum())
            grad_total = float(p[p["PoolCat"] == "restricted_then_all"]["Total QTY"].sum())
            grad_avail = float(p[p["PoolCat"] == "restricted_then_all"]["Available QTY"].sum())
            sngn = float(p[p["PoolCat"] == "sngn_inflow_at_ready"]["PoolQty"].sum())
            full_supply = start_all + start_res + max(0.0, grad_total - grad_avail) + sngn
            avg_price = ki_avg_price_lookup.get(ki, 0.0)
            if avg_price == 0.0:
                ki_items = items_df[items_df["KI"] == ki]["Item Num"].tolist()
                prices_d = d.get("prices", {})
                ki_prices = [prices_d.get(it, 0.0) for it in ki_items if prices_d.get(it, 0.0) > 0]
                avg_price = sum(ki_prices) / len(ki_prices) if ki_prices else 0.0
            new_excess_rows.append({
                "KI": ki,
                "ExcessAtFarm_QTY": full_supply,
                "AvgPrice": avg_price,
                "ExcessAtFarm_$": full_supply * avg_price,
            })
        if new_excess_rows:
            new_excess_df = pd.DataFrame(new_excess_rows)
            farm_df = pd.concat([farm_df, new_excess_df], ignore_index=True)
            farm_df = farm_df.sort_values("ExcessAtFarm_$", ascending=False).reset_index(drop=True)
            log(f"  Excess by KI: augmented with {len(new_excess_rows)} stranded KIs")

    # Task 51: Reason for Excess classifier
    fwd_plan_for_class = d["plan_lifted"]
    fwd_plan_for_class = fwd_plan_for_class[fwd_plan_for_class["MonthNum"].between(5, 12)]
    offplan_demand_df = d.get("offplan_demand", pd.DataFrame())
    hist_df_for_class = d["hist"].copy()
    if "KI" not in hist_df_for_class.columns:
        hist_df_for_class["KI"] = hist_df_for_class["Item Num"].map(item_to_ki)
    kis_with_fwd_plan = set(fwd_plan_for_class["KI"].unique())
    kis_with_offplan = set(offplan_demand_df["KI"].unique()) if not offplan_demand_df.empty else set()
    kis_with_hist = set(hist_df_for_class.dropna(subset=["KI"])["KI"].unique())

    def classify_excess_reason(ki):
        has_plan = ki in kis_with_fwd_plan
        has_offplan = ki in kis_with_offplan
        has_hist = ki in kis_with_hist
        if not has_plan and not has_offplan and not has_hist:
            return "No forward plan, no history"
        if not has_plan and (has_offplan or has_hist):
            return "No forward plan, has history"
        walk_ki = pool_walk_df[pool_walk_df["KI"] == ki] if not pool_walk_df.empty else pd.DataFrame()
        if walk_ki.empty:
            return "Plan met, no history available to lift"
        s2_demand = float(walk_ki["Stage2_Demand_Total"].sum())
        s2_filled = float(walk_ki["Stage2_Filled"].sum())
        if s2_demand <= 0:
            return "Plan met, no history available to lift"
        s2_fill_ratio = s2_filled / s2_demand if s2_demand > 0 else 0
        if s2_fill_ratio < 0.95:
            return "Inventory timing mismatch - supply ready after peak demand"
        return "Plan met + lift fully filled - true overproduction"

    # Synthesis stats from plan_lifted
    pl = d.get("plan_lifted")
    if pl is not None and "Plan_Source" in pl.columns:
        synth_rows = pl[pl["Plan_Source"] == "Synthesized"]
        _synthesis_count = len(synth_rows)
        _synthesis_units = float(synth_rows["LiftedQty"].sum()) if not synth_rows.empty else 0.0
        if not synth_rows.empty:
            prices_lookup = d.get("prices", {})
            if isinstance(prices_lookup, pd.DataFrame):
                pr_map = prices_lookup.set_index(["Customer","Item Num"])["Price"].to_dict() if all(c in prices_lookup.columns for c in ["Customer","Item Num","Price"]) else {}
            else:
                pr_map = prices_lookup
            def _price(row):
                # Try (Customer, Item Num) first, then Item Num key
                p = pr_map.get((row["Customer"], row["Item Num"])) if isinstance(pr_map, dict) else None
                if p is None: p = pr_map.get(row["Item Num"]) if isinstance(pr_map, dict) else 0
                return p or 0
            _synthesis_dollars = float((synth_rows["LiftedQty"] * synth_rows.apply(_price, axis=1)).sum())
        else:
            _synthesis_dollars = 0.0
    else:
        _synthesis_count = 0
        _synthesis_units = 0.0
        _synthesis_dollars = 0.0

    if not farm_df.empty:
        farm_df["ExcessReason"] = farm_df["KI"].apply(classify_excess_reason)
        reason_counts = farm_df.groupby("ExcessReason").agg(
            count=("KI", "count"),
            qty=("ExcessAtFarm_QTY", "sum"),
            dollars=("ExcessAtFarm_$", "sum"),
        ).sort_values("qty", ascending=False)
        log(f"  Excess Reason distribution:")
        for reason, row in reason_counts.iterrows():
            log(f"    {reason}: {int(row['count'])} KIs, {row['qty']:,.0f} units, ${row['dollars']:,.0f}")

    # QA #69: Build off-plan customer Net $ per (Customer × KI) for Plan by KI tab
    # Off-plan customers don't have plan, so Net = YTD $ + Forward FwdFill $ (all positive)
    plan_custs_set_for_op = set(ki_panel["Customer"].unique())
    items_for_op = d.get("items")
    item_to_ki_op = items_for_op.drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict() if items_for_op is not None else {}

    op_records = []
    # YTD $ per (off-plan cust × KI)
    ytd_for_op = d.get("ytd")
    if ytd_for_op is not None and not ytd_for_op.empty:
        ytd_op_only = ytd_for_op[~ytd_for_op["Customer"].isin(plan_custs_set_for_op) & ytd_for_op["Customer"].notna()].copy()
        ytd_op_only["KI"] = ytd_op_only["Item Num"].map(item_to_ki_op)
        ytd_op_only = ytd_op_only.dropna(subset=["KI"])
        if "Revenue" in ytd_op_only.columns:
            ytd_op_grouped = ytd_op_only.groupby(["Customer", "KI"])["Revenue"].sum().reset_index().rename(columns={"Revenue":"YTD_$"})
        else:
            ki_avg_for_op = ki_total.set_index("KI")["AvgPrice"].to_dict()
            ytd_op_only["AvgPrice"] = ytd_op_only["KI"].map(ki_avg_for_op).fillna(0)
            ytd_op_only["YTD_$"] = ytd_op_only["Qty"] * ytd_op_only["AvgPrice"]
            ytd_op_grouped = ytd_op_only.groupby(["Customer", "KI"])["YTD_$"].sum().reset_index()
        for _, row in ytd_op_grouped.iterrows():
            op_records.append({"Customer": row["Customer"], "KI": row["KI"], "Net_$": row["YTD_$"]})

    # Forward FwdFill $ per (off-plan cust × KI)
    offplan_alloc_for_op = d.get("offplan_alloc", pd.DataFrame())
    if not offplan_alloc_for_op.empty:
        ki_avg_for_op = ki_total.set_index("KI")["AvgPrice"].to_dict()
        op_alloc = offplan_alloc_for_op.copy()
        op_alloc["AvgPrice"] = op_alloc["KI"].map(ki_avg_for_op).fillna(0)
        op_alloc["FwdFill_$"] = op_alloc["Stage2Filled"] * op_alloc["AvgPrice"]
        op_fwd_grouped = op_alloc.groupby(["Customer","KI"])["FwdFill_$"].sum().reset_index()
        for _, row in op_fwd_grouped.iterrows():
            op_records.append({"Customer": row["Customer"], "KI": row["KI"], "Net_$": row["FwdFill_$"]})

    if op_records:
        op_df = pd.DataFrame(op_records)
        # Sum across YTD + Forward per (Cust, KI)
        op_df = op_df.groupby(["Customer", "KI"])["Net_$"].sum().reset_index()
        offplan_net_pivot = op_df.pivot(index="KI", columns="Customer", values="Net_$").fillna(0)
    else:
        offplan_net_pivot = pd.DataFrame()

    # QA #70: Add off-plan over-plan to ki_total.OverPlan_dlr_KI_total so Plan by KI / Over-Plan by KI
    # show consistent customer coverage with Customer Over-Plan Detail / Exec Summary.
    # Off-plan customers have Plan=0, so all their ship $ is over-plan by definition.
    if not offplan_net_pivot.empty:
        op_overplan_per_ki = offplan_net_pivot.sum(axis=1).clip(lower=0)
        op_overplan_df = op_overplan_per_ki.reset_index().rename(columns={0: "OverPlan_offplan_dlr"})
        op_overplan_df.columns = ["KI", "OverPlan_offplan_dlr"]
        ki_total = ki_total.merge(op_overplan_df, on="KI", how="left").fillna({"OverPlan_offplan_dlr": 0})
        ki_total["OverPlan_dlr_KI_total"] = ki_total["OverPlan_dlr_KI_total"] + ki_total["OverPlan_offplan_dlr"]
        ki_total = ki_total.drop(columns=["OverPlan_offplan_dlr"])
        log(f"  ki_total: added off-plan over-plan to OverPlan_dlr_KI_total")

    return {
        "panel": panel,
        "ki_panel": ki_panel,
        "ki_month": ki_month,
        "lift_by_cust_item": lift_by_cust_item,
        "excess_at_farm": farm_df,
        "ki_lifecycle": ki_lifecycle,
        "production_recovery": production_recovery,
        "ki_total": ki_total,
        "h_pivot": h_pivot,
        "item_meta": item_meta,
        "stranded_ki_count": _stranded_count,
        "stranded_ki_units": _stranded_units,
        "stranded_ki_dollars": _stranded_dollars,
        "synthesis_row_count": _synthesis_count,
        "synthesis_dollars": _synthesis_dollars,
        "synthesis_units": _synthesis_units,
        "_offplan_net_pivot": offplan_net_pivot,
    }


# ---------------------------------------------------------------------------
# Tab writers
# ---------------------------------------------------------------------------

def write_tab1_readme(wb):
    """Tab 1: Read Me & Methodology — narrative + tab guide. Updated for two-stage."""
    log("Writing Tab 1: Read Me & Methodology...")
    ws = wb.create_sheet("Read Me & Methodology")
    set_col_widths(ws, {"A": 100})
    r = 1

    def block(text, style="normal", indent=0, height=None):
        nonlocal r
        cell = ws.cell(row=r, column=1, value=("    "*indent) + text)
        if style == "title":
            cell.font = TITLE_FONT
            cell.fill = SECTION_FILL
            ws.row_dimensions[r].height = 28
        elif style == "section":
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            ws.row_dimensions[r].height = 22
        elif style == "bold":
            cell.font = BOLD_FONT
        elif style == "note":
            cell.font = NOTE_FONT
        else:
            cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if height: ws.row_dimensions[r].height = height
        r += 1

    block("NOR CAL — Forward Looking Inventory vs Sales Plan", style="title")
    block("Snapshot: 2026-04-27  |  Generated: " + datetime.now().strftime('%Y-%m-%d %H:%M') + "  |  Owner: Jonathan Saperstein, CEO Everde Growers", style="note")
    r += 1

    # Section 1
    block("1. Purpose & Scope", style="section")
    block("This workbook combines two views built from the same underlying model:")
    block("•   Operations view — where we are short of plan and why, so production/expediting decisions can be prioritized", indent=1)
    block("•   Sales view — where we have inventory above plan or above history, so commercial action / pricing decisions can be made", indent=1)
    block("Scope: NOR CAL region only. Forward window: May–Dec 2026. YTD context: Jan–Apr 2026 actuals vs plan.", indent=0)
    block("Engine note: as of this build, the model engine is region-agnostic — the same code base will run for FL, OR, SO CAL, and TX with no methodology changes (just a region constant). Customer channels are loaded dynamically from the plan file.", style="note", indent=0)
    r += 1

    # Section 2 — Two-Stage Methodology (NEW)
    block("2. Two-Stage Allocation Methodology", style="section")
    block("The forward-fulfillment walk runs in two passes for each Key Item Grouping × month from May → Dec:", indent=1)
    block("STAGE 1 — Defend the Original Plan", style="bold", indent=1)
    block("•   Each month, after inflows (graduations + new SN/GN ready), customers compete for inventory pool to fulfill their ORIGINAL committed plan.", indent=2)
    block("•   NBB customers (West Coast, Midwest, Southeast, MLC) get first dibs on the SS/GS pre-ready restricted pool — but only their plan demand applies.", indent=2)
    block("•   Then everyone — BB and NBB — competes pro-rata in the all-customer pool. When pool is short, every customer's plan defense suffers proportionally to their plan demand share. No BB-vs-NBB priority within the all-pool.", indent=2)
    block("•   Output: Plan Filled (S1) and Plan Miss (S1) per customer per item per month.", indent=2)
    block("STAGE 2 — Apply Lift From Surplus", style="bold", indent=1)
    block("•   After Stage 1 completes for all months, Stage 2 layers on top using whatever pool is left over.", indent=2)
    block("•   Lift demand for each planned customer = max(0, Lifted QTY − Original Plan QTY) — i.e., the additional demand from smoothed history beyond the contracted plan.", indent=2)
    block("•   Off-plan customers (Other / MLC / Southeast in NOR CAL — channels with history but no NOR CAL plan) bring their full smoothed history as Stage 2 demand.", indent=2)
    block("•   Stage 2 in any month M is constrained by min(end_pool[m] for m ∈ [M, Dec]) — this protects later months from losing pool that they need for plan defense.", indent=2)
    block("•   Same two-step pool allocation applies: NBB takes any remaining restricted pool first; everyone competes pro-rata in the all-pool.", indent=2)
    block("•   Output: Lift Filled (S2) per customer.", indent=2)
    block("Why two stages: Stage 1 defends plan first when pool is short, exposing the TRUE plan miss (the prior single-stage walk allocated pool by lifted-demand share, which over-allocated to high-lift items and hid plan miss elsewhere). Stage 2 layers lift on top so we still see realistic upside potential.", style="note", indent=1)
    r += 1

    # Section 3 — Channel universe
    block("3. Customer Channel Universe", style="section")
    block("Canonical company-wide buckets (the same 7 + Other apply across all 5 regions):", style="bold", indent=1)
    block("•   HD (Big Box) — pool eligibility: all-pool only", indent=2)
    block("•   Lowes (Big Box) — pool eligibility: all-pool only", indent=2)
    block("•   Walmart (Big Box) — pool eligibility: all-pool only", indent=2)
    block("•   West Coast (Non-Big-Box) — gets restricted SS/GS pre-ready pool", indent=2)
    block("•   Midwest (Non-Big-Box) — gets restricted pool. Includes COSTCO PNW, SITEONE MIDWEST, GREEN ACRES.", indent=2)
    block("•   Southeast (Non-Big-Box) — gets restricted pool. Includes 'HD FL - SOUTHEAST' (regional distribution variant).", indent=2)
    block("•   MLC (Non-Big-Box) — gets restricted pool. SO CAL-specific in the 2026 plan.", indent=2)
    block("•   Other (catch-all, BB-like) — anything that doesn't match. No pre-ready privilege.", indent=2)
    block("For NOR CAL specifically, the planned customer set (from the 2026 plan file) is HD, Lowes, Walmart, West Coast, Midwest. Southeast / MLC / Other appear in NOR CAL only via Stage 2 (history-driven demand, no plan).", indent=1)
    r += 1

    # Section 4 — Data Sources
    block("4. Data Sources", style="section")
    block("Source files used (all paths relative to user's connected folders):", style="bold")
    sources = [
        ("Sales Plan",            "Shared/Sales Plan/2026 Sales Plan by Item.xlsx",
         "2026 Sales Plan tab. Filtered to Region='NOR CAL'. Planned customer set derived dynamically from the file (5 channels for NOR CAL)."),
        ("V158 Key Item Report",  "Shared/Key Item report V158/Key Item Report V158.xlsx",
         "'Region Dataset' (KI labels, region item mapping) + 'Demand Data' (Customer × Item × Region with KI Tagged column). NOT used for pricing per CEO direction."),
        ("Inventory Transform",   "Key Item Review/Archive/Inventory Transform 042726.xlsx",
         "4/27/26 snapshot. NOR CAL inventory at lot level: Item Num, Grade, Demand Window, Ready Date, Available QTY, Total QTY."),
        ("Sales Actuals 2023-25", "Shared/Sales Data/{2023,2024,2025} Sales by Item.xlsx", "Historical sales. Cached as parquet. Used for smoothing baseline + Tier 2 reference price + YoY comparison."),
        ("2026 YTD Actuals",      "Key Item Review/Archive/2026 Sales by Item 042726.xlsx", "Jan–Apr 2026 actual sales. Tier-1 actual pricing + YTD performance baseline."),
        ("HD Cross-Reference",    "Shared/Inventory Cross References/Home Depot Corp-VN=PO xref rev.04222026.xlsb", "HD SKU ↔ Item mapping. Used to build HD-SKU-level lift groups."),
        ("Lowes Cross-Reference", "Shared/Inventory Cross References/LOWE'S xref rev.04292029.xlsb", "Lowe's SKU ↔ Item mapping. Used to build Lowes-SKU-level lift groups."),
    ]
    for name, path, desc in sources:
        block(f"{name}: {path}", style="bold", indent=1)
        block(desc, indent=2)
    r += 1

    # Section 5 — Inventory pool
    block("5. Inventory Pool Construction", style="section")
    block("Grade rules — what's in the pool:", style="bold", indent=1)
    block("•  A, B (post-ready): All-customer pool from May 1, using Total QTY", indent=2)
    block("•  SS, GS pre-ready (Ready Date > 4/30/26): Restricted pool (NBB only) using Available QTY", indent=2)
    block("•  SS, GS becoming ready in month M: graduate at start of M — Available QTY leaves restricted pool, Total QTY enters all-customer pool", indent=2)
    block("•  SN, GN, S2N pre-ready: not in any pool until ready", indent=2)
    block("•  SN, GN, S2N becoming ready in month M: enter all-customer pool with Total QTY", indent=2)
    block("•  C, D: excluded (damage / quarantine)", indent=2)
    block("•  Any P-prefix grade (PN, P2N, P, etc.): excluded (production-stage, not yet sellable)", indent=2)
    block("Demand Window filter: keep only rows where Demand Window ∈ {2026 H1, 2026 H2, 2025, 2024, blank}.", indent=1)
    block("Substitution: items within same Key Item Grouping pool together. HD/Lowes additionally have SKU-level grouping for lift.", indent=1)
    r += 1

    # Section 6 — Lift methodology
    block("6. Lift Methodology — Smoothed History Baseline", style="section")
    block("Aggregation unit for lift varies by customer:", style="bold", indent=1)
    block("•  HD: HD SKU (from HD xref). Multiple-SKU items use majority-history SKU as primary.", indent=2)
    block("•  Lowes: Lowes SKU (same logic).", indent=2)
    block("•  HD/Lowes items without SKU mapping but with HD/Lowes plan or history: KI Grouping fallback.", indent=2)
    block("•  HD/Lowes items with no SKU AND no plan AND no history: excluded from lift.", indent=2)
    block("•  Walmart, West Coast, Midwest, Southeast, MLC, Other: Key Item Grouping.", indent=2)
    block("Smoothing rule (>10× outlier drop):", style="bold", indent=1)
    block("•  Compute raw 3-yr mean across 2023, 2024, 2025 monthly QTYs at the (Customer, LiftGroup, Month) level.", indent=2)
    block("•  Drop any year where that year's value > 10× the mean of the OTHER years (treats one-time spikes as outliers).", indent=2)
    block("•  Re-average over remaining years.", indent=2)
    block("Lift application: when smoothed history exceeds (Customer, LiftGroup, Month) plan total, scale every item's plan QTY proportionally upward by (smoothed / grouping plan). Result = Lifted Plan QTY.", indent=1)
    r += 1

    # Section 7 — Pricing
    block("7. Pricing Methodology — How Items Were Priced", style="section")
    block("All pricing comes from sales actuals — NOT from V158 plan price tables (per CEO direction).", indent=1)
    block("6-tier cascade per (Customer, Item):", style="bold", indent=1)
    block("•  Tier 1: 2026 YTD avg Sell Price by (Customer, Item)", indent=2)
    block("•  Tier 2: 2025 avg Sell Price by (Customer, Item)", indent=2)
    block("•  Tier 3: Customer × LiftGroup avg (using 2025+2026 actuals)", indent=2)
    block("•  Tier 4: LiftGroup-only avg (across customers)", indent=2)
    block("•  Tier 5: Genus + Size avg", indent=2)
    block("•  Tier 6: Size-only avg (final fallback)", indent=2)
    block("Coverage: 100% of plan items have a price.", indent=1)
    block("Total plan dollars implied: $34.4M vs CEO official ~$34.27M (within ~0.4% — cascade approximation noise on items where actuals differ from negotiated contract).", indent=1)
    r += 1

    # Section 8 — Tab guide (UPDATED)
    block("8. How to Read Each Tab", style="section")
    block("This workbook has 22 tabs grouped by domain (color-coded for navigation):", indent=1)
    block("Reference / Utility (grey) — Read Me, Build Health, Changes", indent=1, style="bold")
    block("Top-level summary — Exec Summary (navy), Plan by KI (teal), KI Lifecycle pair (slate)", indent=1, style="bold")
    block("Domain detail — Miss (red), Production Recovery (orange), Lift (green), Excess by KI (gold), Customer Over-Plan (blue), Cross-cut views (purple)", indent=1, style="bold")
    r += 1
    tab_guide = [
        ("Tab 1 — Read Me & Methodology",
         "This tab. Documents purpose, methodology, data sources, and how each downstream tab is built. The first place to look when reading the workbook fresh, or when a number on another tab is unclear."),
        ("Tab 2 — Build Health (NEW Phase A)",
         "Auto-populated dashboard for verification + dataset health. Five sections: (1) Verification status — walks reconciliation + pool conservation pass/fail with top failing KIs if any; (2) Methodology Stats — counts and totals for Production Recovery, Synthesized plan rows (Option B), Stranded KIs; (3) Inventory Snapshot in $; (4) Data Quality Issues — flags items in plan with no pricing, items with no item-master, KIs with negative YTD Actual (credits/returns), etc.; (5) Comparison to Prior Build — deltas in Plan Miss $, Forward Forecast $, Excess at Farm $, synthesized rows since the last rebuild. Check this first whenever the numbers feel off."),
        ("Tab 3 — Changes (version history)",
         "Per-build changelog. Newest entries at top. Each version block tags changes by category ([Methodology], [Tab Added], [Cosmetic], etc.). Read this when comparing versions or to understand why current numbers differ from a prior file."),
        ("Tab 4 — Exec Summary",
         "Top-line dashboard. Three sequential walks (YE, May–Dec, YTD) showing Original Plan $ → Plan Miss [Down] → Customer Over-Plan [Up] → Net Position per customer. The 'achieved' column is labeled 'Projected $' on the YE and May–Dec walks (engine forecast) and 'Actual $' on the YTD walk (real Jan–Apr sales). Read this first for the headline."),
        ("Tab 5 — Plan by KI",
         "Per-KI roll-up: Original Plan $ YE, Plan Miss $ YE, Over-Plan $ YE, Net $ YE — with per-customer columns showing each customer's net Up/Down position. Sort by any of these to find the biggest contributors."),
        ("Tab 6 — KI Lifecycle (Units) (NEW Phase A)",
         "Full-year unit picture per Key Item Grouping (no customer dimension). 13 columns: YTD Plan / YTD Actual / Start INV (R + NR) / Forward Plan / Forward Sales Projected / Forward Inflows (Pool + Recovery) / Ending INV / YE Plan / YE Projected. Pool conservation: Start (R+NR) + Inflows = Forward Projected + Ending INV — should hold for every row. The two YE columns at the right give the at-a-glance plan-vs-projected comparison."),
        ("Tab 7 — KI Lifecycle ($) (NEW Phase A)",
         "Same structure as Tab 6 but in dollars (each unit column × KI's average price). Use to size dollar impact of patterns spotted on the Units tab."),
        ("Tab 8 — Miss Summary by KI",
         "All KIs with positive plan miss, ranked by Miss $. KI-level totals (no customer dimension). Quick scan for biggest miss contributors."),
        ("Tab 9 — Miss by Customer × KI",
         "Same data as Tab 8 but with customer dimension. Use to see which customer drives miss for each KI."),
        ("Tab 10 — Miss by KI × Month",
         "Same KI-level miss but split by forward month (May–Dec). Shows whether miss is concentrated early or late in the year."),
        ("Tab 11 — Customer Miss Detail",
         "Item-level drill-down (Customer × Item × Month) showing per-row miss with grow-days context — what the production target item would be if recovered."),
        ("Tab 12 — Production Recovery (Phase A)",
         "Action list of plan miss recoverable via production within grow-days window. Each row is a (Customer × KI × Month) action with target item to grow, recoverable QTY/$, min grow days, days to month-end, slack days, items being substituted for. Production Recovery zeroes those misses on the Miss tabs and adds them to FwdFill on the Forward Sales Projected column. Recovery dollars also surface as a separate inflow column on the KI Lifecycle tabs."),
        ("Tab 13 — Lift Summary by KI",
         "Per-KI roll-up of lift demand (smoothed history > plan). Includes synthesized rows from Phase C / Option B where original plan was zero — these appear as Plan_orig=0, Lifted=smoothed_history, Lift=full smoothed."),
        ("Tab 14 — Lift by Customer × KI",
         "Lift summary but per (Customer × KI). Use to find which customer is driving lift on each KI. Also includes synthesized rows."),
        ("Tab 15 — Lift by KI × Month",
         "Lift broken out by month for each KI. Shows when in the year the lift demand peaks."),
        ("Tab 16 — Lift by Customer × Item",
         "Item-level lift detail (Customer × Item) with Smoothed History QTY shown alongside Plan and Lifted Plan. The deepest drill-down on lift; use to audit specific item lift factors."),
        ("Tab 17 — Excess by KI",
         "Pool inventory unused at end of December, per KI. This is 'excess at the farm' — physical stock left over after plan defense + lift. Includes a 'Reason for Excess' column with one of five categories: 'No forward plan, no history' / 'No forward plan, has history' / 'Plan met, no history available to lift' / 'Plan met + lift fully filled — true overproduction' / 'Inventory timing mismatch — supply ready after peak demand'. Stranded KIs (no forward plan, has inventory) appear here with their full supply as excess."),
        ("Tab 18 — Over-Plan by KI",
         "KIs where Customer Over-Plan QTY/$ vs Original Plan YE > 0. Three over-plan thresholds shown: vs Original Plan (most aggressive), vs Smoothed History, vs Lifted YE Plan (most conservative). The 'Why no further lift' column explains lift construction; the 'Flags' column attributes over-plan to its driver (YTD over-shipped, forecast from history, etc.)."),
        ("Tab 19 — Customer Over-Plan Detail",
         "Per-customer view of over-plan with multi-year history (2023–2025), YTD pace vs plan, and pricing reference (2026 YTD vs 2025). 'Why no further lift' and 'Flags' columns added at the right give per-row attribution and lift construction explanation."),
        ("Tab 20 — Channel Summary",
         "Per-customer YE roll-up: their plan, their projected achievement, top contributing KIs ranked under each customer."),
        ("Tab 21 — YTD Performance",
         "Jan–Apr plan vs actual by Customer × KI. Sets context for 'are we already behind' coming into May."),
        ("Tab 22 — Pricing Comparison",
         "Cascade pricing vs the Pricing Look up reference file, per Customer × Item. Shows where prices differ and the dollar impact at plan QTY."),
    ]
    for tab, desc in tab_guide:
        block(tab, style="bold", indent=1)
        block(desc, indent=2)
    r += 1

    # Section 9 — Phase A/B/C Methodology Updates (NEW)
    block("9. Phase A / B / C Methodology Updates (recent — read this if you've used prior versions)", style="section")
    block("Three substantial methodology improvements were applied across recent builds. Each is described below.", indent=1)
    r += 1

    block("9a. Production Recovery (Phase A — added Tab 12)", style="bold", indent=1)
    block("For each (Customer × KI × Month) plan miss, the model checks if any item in that KI has grow days short enough to fit between snapshot date (4/27) and month-end. If yes, the entire KI miss for that customer/month is treated as RECOVERABLE — production grows the target item to substitute within the KI. Recovery zeroes the miss and adds the recovered QTY to Forward Sales Projected on downstream tabs. Recovery is shown as a separate column on KI Lifecycle (Forward Inflows - Recovery) so it's visible vs existing pool inflows.", indent=2)
    r += 1

    block("9b. Lot-identity tracking in restricted pool (Phase B — engine fix)", style="bold", indent=1)
    block("Prior to Phase B, the engine over-allocated forward shipments by ~55K units across ~200 KIs. Root cause: when NBB pre-consumed the restricted SS/GS pool before lots graduated, the engine then re-added the full Total QTY of those lots to all-pool at graduation, double-counting the consumed Avail portion. Phase B fix: each restricted lot now tracks remaining_avail and consumed_pre_grad. NBB consumption is allocated FIFO by ready_month. At graduation, all-pool gains (Total - consumed_pre_grad) — the actual physical residual — instead of the full Total. Pool conservation now holds for all 470 KIs (zero diff). The shift moved ~$320K of phantom forward shipments from Forward Forecast back into Plan Miss.", indent=2)
    r += 1

    block("9c. History-augmented lift (Phase C — Option B synthesis)", style="bold", indent=1)
    block("The lift mechanism prior to Phase C only amplified plan rows that already existed — if Plan = 0 in a given month, lift × 0 = 0 forever, even if history showed strong demand. This silently dropped customers/items where the planner entered plan only in some months. Example: HD #010 Inch MANDEVILLA had 2026 plan of 3,000 in April only; HD historical pattern showed ~5,000 units/year in May–Dec. Pre-Phase-C, the model forecast 0 for HD May–Dec.", indent=2)
    block("Phase C synthesis: for each (Customer × LiftGroup × Month) combination in May–Dec where smoothed history > 0 but no plan row exists, the engine now synthesizes a plan row with PlanQty = LiftedQty = smoothed history. These synthesized rows are tagged Plan_Source='Synthesized' (originals are tagged Plan_Source='Original'). HD/Lowes use SKU-level synthesis (matched to most-historically-shipped item under each HDSKU/LOWESSKU); other planning customers (Walmart, West Coast, Midwest) use KI-level synthesis attributed to the customer's most-historically-shipped item in that KI.", indent=2)
    block("Critical framing: Original Plan stays pristine. OrigYE used as Customer Over-Plan denominator is computed from ORIGINAL plan rows only (Plan_Source='Original'), NOT augmented with synthesized rows. The synthesis drives forecast/Forward Sales Projected, but the planner's commitment is sacrosanct. So a customer like HD MANDEVILLA shows OrigYE = 3,000, Forward Projected ≈ 2,485 from synthesis, and Customer Over-Plan = small/positive once total YTD+Projected exceeds the original 3,000.", indent=2)
    block("Synthesized rows appear on all four Lift tabs as lift-over-zero-plan (Plan_orig=0, Lifted=smoothed_history, Lift=full smoothed). Use the Lift tabs to audit what the synthesis is forecasting and whether the historical pattern still applies.", indent=2)
    r += 1

    # Section 10 — Pool Conservation Identity
    block("10. Pool Conservation Identity (KI Lifecycle Tabs)", style="section")
    block("The KI Lifecycle (Units) and KI Lifecycle ($) tabs are built around a verifiable identity:", indent=1)
    block("    Start INV (R + NR) + Forward Inflows (Pool + Recovery) = Forward Sales Projected + Ending INV Available", style="bold", indent=2)
    block("In words: every unit of inventory that exists at the start of May, plus every unit that comes ready (graduates from SS/GS pre-ready or arrives as new SN/GN) plus every unit produced via Production Recovery, must equal what we forecast to ship May–Dec plus what's left over at end of December. Build Health Section 1 verifies this identity holds for all 470 KIs after every rebuild — any failure flags as a red conservation gap and is debuggable from the lifecycle tab directly.", indent=1)
    block("This is distinct from the walks reconciliation (YE Net = YTD Net + Forward Net), which holds at the customer × KI level on Plan Miss / Customer Over-Plan tabs. Pool conservation is a SUPPLY-side identity; walks is a DEMAND-side identity.", indent=1)
    r += 1

    # Section 11 — Reason for Excess categories
    block("11. Reason for Excess — Categorization Logic", style="section")
    block("The Excess by KI tab includes a 'Reason for Excess' column with one of five labels. Each KI is classified at build time:", indent=1)
    block("•  No forward plan, no history — KI has inventory but no 2026 May–Dec plan rows AND no historical sales of any year. Truly orphan inventory. Should drive an assortment / cull conversation.", indent=2)
    block("•  No forward plan, has history — KI has no 2026 May–Dec plan but did have historical sales in prior years. Includes KIs where the planner entered Q1-only plan (no May–Dec rows) but history shows year-round demand. Should trigger a 'should this be in the plan?' review.", indent=2)
    block("•  Plan met, no history available to lift — Plan was filled by Stage 1 within available pool, but smoothed history doesn't exceed plan, so Stage 2 lift had nothing to add. The remaining pool became excess.", indent=2)
    block("•  Plan met + lift fully filled — true overproduction — Plan filled, lift demand fully consumed pool, and the model still has inventory leftover. This is real overproduction relative to demonstrated demand. Should trigger a production planning conversation.", indent=2)
    block("•  Inventory timing mismatch — supply ready after peak demand — Plan exists, lift demand exists, but Stage 2 couldn't fully fill because supply timing doesn't match demand timing (e.g., demand peaked May–Aug, but supply comes ready Sep–Dec). Should review production scheduling.", indent=2)
    r += 1

    # Section 12 — Stranded vs Zero-Forward-Demand
    block("12. Stranded KIs vs Zero-Forward-Demand KIs (Build Health Section 2)", style="section")
    block("Two related-but-different metrics track inventory with no forward demand:", indent=1)
    block("•  Stranded KIs — KIs that have inventory in the pools dataset but never enter the engine's pool walk because they have NO plan rows AND NO off-plan demand AND no synthesized rows. Truly orphaned. Currently 9 KIs.", indent=2)
    block("•  KIs with zero forward fulfillment — KIs that end up with FwdPlan=0 AND FwdProjected=0 in the lifecycle tab. This is a SUPERSET of stranded — it includes the 9 stranded plus an additional ~15 KIs that have plan rows somewhere but the engine couldn't allocate any forward fulfillment (e.g., plan exists but inventory was 100% consumed in YTD). Currently ~24 KIs.", indent=2)
    block("Practical interpretation: 'Stranded' is the truly orphan count — inventory with no demand pathway anywhere. 'Zero forward fulfillment' is the broader set — KIs where forward shipments will be zero for any reason. Both are visible on Build Health Section 2 for context.", indent=1)
    r += 1

    # Section 13 — YTD Actuals: negative values explanation
    block("13. Why YTD Sales Actual Can Be Negative", style="section")
    block("The 'YTD Sales Actual (Units)' column on the lifecycle tab can show small negative values for some KIs. These represent net position after credits and returns from the YTD source data — the customer paid us back for product that didn't ship correctly, so the net QTY for that KI ends up below zero. This is real data and intentionally not clipped at zero; clipping would hide the credit/return signal. Build Health Section 4 (Data Quality) lists KIs with negative YTD Actual values for visibility.", indent=1)
    r += 1

    # Section 14 — Forward Sales Projected vs YTD Sales Actual
    block("14. Why It's Called 'Forward Sales Projected' Not 'Forward Sales Actual'", style="section")
    block("Earlier versions of the workbook used 'Forward Sales Actual' as a column header. This was misleading — May–Dec values are the engine's FORECAST (Stage 1 plan defense + Stage 2 lift fulfillment + Production Recovery), not yet-realized actual sales. The header was renamed to 'Forward Sales Projected' on the KI Lifecycle tabs. The same logic applies on Exec Summary: the 'achieved' column is labeled 'Projected $' on YE and May–Dec walks (engine forecasts) and 'Actual $' only on the YTD walk (real Jan–Apr data).", indent=1)
    r += 1

    # Section 15 — Pricing Methodology & Margin of Error (NEW)
    block("15. Pricing Methodology & Margin of Error", style="section")
    block("Dollar values in this workbook are computed using one of three pricing methods, depending on the metric. Understanding when each is used helps explain why the same dollar concept may differ slightly across tabs.", indent=1)
    r += 1

    block("Three pricing methods at play:", style="bold", indent=1)
    block("•  Method A — Per-item × per-customer pricing (most accurate). Each item-customer combination uses its specific actual sell price (Tier 1 actuals, Tier 2 historical avg, Tier 3 V158 plan price, etc.). Used for raw item-level dollar values like panel.OrigYE_$, panel.YTD_$, panel.FwdFill_$.", indent=2)
    block("•  Method B — Customer-aware KI-level pricing (used for most workbook-level totals). For each (Customer × KI), AvgPrice = Σ(item OrigYE × item Price) / Σ(item OrigYE) = OrigYE-weighted average. KI-level dollar = Σ over customers of (Cust×KI quantity × Cust×KI AvgPrice). Used by Plan by KI Original Plan $, Miss Summary by KI Miss $, Customer Over-Plan $, Exec Summary YE walks, and KI Lifecycle ($) customer-attributable columns.", indent=2)
    block("•  Method C — KI-level pricing (used for supply-side metrics with no customer dimension). KI-level AvgPrice = Σ(Cust×KI quantity × Cust×KI AvgPrice) / Σ(Cust×KI quantity) — a 'KI-wide' single price. Used for Start INV ($), Forward Inflows ($), Ending INV ($), Excess at Farm $.", indent=2)
    r += 1

    block("Why three methods exist (it's not arbitrary):", style="bold", indent=1)
    block("Method A is used wherever item-level data exists (per-item shipping, per-item plan). It's the most accurate. But for aggregated metrics that need to roll up across items within a KI, we have to pick a representative price.", indent=2)
    block("Method B (customer-aware KI-avg) is used because Plan Miss QTY and Customer Over-Plan QTY are computed at the (Customer × KI) level with intra-KI substitution capping (per the methodology approved 2026-05-07: customers can substitute items within a KI, so net miss is capped at the KI level). Since Miss QTY itself isn't at the item level, we can only price it at the (Cust × KI) average level. Method B's price is implicitly equivalent to allocating the Cust × KI miss to items proportionally to OrigYE share and pricing each at item level — algebraically the same.", indent=2)
    block("Method C (KI-only avg) is used for supply-side metrics — pool inventory, inflows, ending INV — because these don't have a customer dimension. A unit of inventory in the pool is unassigned to a customer until it's allocated, so it can't use customer-aware pricing.", indent=2)
    r += 1

    block("Margin of error from pricing:", style="bold", indent=1)
    block("Where customer pricing within a KI is uniform, all three methods give the same answer. Where customer pricing varies (different customers buy different items at different prices within a KI, or the same item at different price tiers), the methods diverge.", indent=2)
    block("Quantified margin of error in the current workbook:", indent=2)
    block("•  At workbook total: ~1-2% between Method B (used by Plan by KI, Miss Summary by KI, Lifecycle $, Exec Summary) and Method A (most granular).", indent=3)
    block("•  Per-KI: up to 5-10% on KIs with diverse customer pricing (e.g., #005 BUXUS where West Coast price is ~$13/unit and Lowes price is ~$22.50/unit — the OrigYE-weighted average lies in between).", indent=3)
    block("•  Per-customer-row: same 1-10% range on individual customer rows in walks.", indent=3)
    r += 1

    block("Coverage scope difference (separate from pricing margin of error):", style="bold", indent=1)
    block("Some tabs have different customer coverage by design, which causes dollar totals to differ across tabs in a way that's NOT a pricing issue:", indent=2)
    block("•  Exec Summary YE/May-Dec/YTD walks: planned customers only (HD/Lowes/Walmart/West Coast/Midwest). Off-plan customers (Southeast/MLC/Other) don't have a plan, so the walks intentionally exclude them.", indent=2)
    block("•  KI Lifecycle ($) YE Projected and Forward Sales Projected columns: all customers including off-plan fills (since these are supply-side metrics — every unit shipped from the pool counts, regardless of whether the customer was planned).", indent=2)
    block(f"•  Quantified gap: in this build, Lifecycle YE Projected exceeds Exec Summary YE Projected by ~$135K (~0.4%) — exactly the off-plan Stage 2 fill total. This is correct behavior given each tab's purpose.", indent=2)
    block("•  If you need 'plan customers only' totals, use the Exec Summary walks. If you need 'all forward shipping' totals, use the Lifecycle tab.", indent=2)
    r += 1

    block("Where this margin of error matters (and where it doesn't):", style="bold", indent=1)
    block("•  Doesn't matter much for headline decisions: a $5,000 mix-effect gap on a $5M Plan Miss is well inside the noise of any planning action. The directional signal (which KIs miss, which customers are over-plan) is unaffected.", indent=2)
    block("•  Does matter when reconciling tabs: occasionally the same 'Miss $' value at different aggregation levels won't tie out exactly. Within ~1-2% is normal; larger gaps are a methodology bug worth investigating.", indent=2)
    block("•  Does matter for walks identity in $: Plan - Miss + Over-Plan = Achieved holds in QTY perfectly but in $ has small mix-effect approximation when actual ship mix ≠ plan mix.", indent=2)
    r += 1

    block("How tabs use which method (quick reference):", style="bold", indent=1)
    block("•  Plan by KI (Original Plan $ YE, Plan Miss $, Over-Plan $): Method B", indent=2)
    block("•  KI Lifecycle ($) customer-attributable columns (YTD Plan/Actual, Forward Plan/Projected, YE Plan/Projected): Method B", indent=2)
    block("•  KI Lifecycle ($) supply-side columns (Start INV, Forward Inflows, Ending INV): Method C", indent=2)
    block("•  Miss Summary by KI / Miss by Customer x KI / Miss by KI x Month / Customer Miss Detail: Method B", indent=2)
    block("•  Lift Summary by KI / Lift by Customer x KI / Lift by KI x Month: Method B (where customer dimension applies) or Method A (where item dimension applies)", indent=2)
    block("•  Lift by Customer x Item: Method A (item-level prices)", indent=2)
    block("•  Excess by KI (Excess at Farm $): Method C (no customer assignment for unallocated pool)", indent=2)
    block("•  Over-Plan by KI / Customer Over-Plan Detail: Method B", indent=2)
    block("•  Production Recovery: Method B (recovery $ uses Cust×KI avg)", indent=2)
    block("•  Channel Summary: Method B at customer level", indent=2)
    block("•  Exec Summary walks (Original Plan $, Miss $, Over-Plan $, Projected $): Method B", indent=2)
    block("•  YTD Performance: Method A (item-level)", indent=2)
    block("•  Pricing Comparison: Method A (item-level prices, no aggregation)", indent=2)
    r += 1

    # Section 16 — Glossary (renumbered from 15)
    block("16. Key Terms — Glossary", style="section")
    block("•  Original Plan: the 2026 plan QTY committed by Sales / Planning. Stage 1 defends this. NEVER augmented by synthesis.", indent=1)
    block("•  Lifted Plan: Original Plan × lift factor (where smoothed history exceeds plan). For Phase C synthesized rows, Lifted Plan = synthesized smoothed history (no further lift).", indent=1)
    block("•  Plan_Source: tagged on every plan_lifted row. 'Original' = planner-authored. 'Synthesized' = added by Phase C from history where no original plan existed.", indent=1)
    block("•  Plan Filled (S1): qty fulfilled from inventory pool against Original Plan during Stage 1.", indent=1)
    block("•  Plan Miss (S1): Original Plan demand we cannot fulfill due to insufficient pool. Measured against Original, not Lifted. Production Recovery zeroes recoverable miss.", indent=1)
    block("•  Lift Filled (S2): qty fulfilled from surplus pool against the additional lift demand during Stage 2.", indent=1)
    block("•  Forward Sales Projected: engine forecast for May–Dec = Stage 1 fills + Stage 2 fills + Production Recovery. Distinct from YTD Sales Actual which is real Jan–Apr data.", indent=1)
    block("•  Customer Over-Plan: when a customer's YTD + Forward Sales Projected exceeds their ORIGINAL Plan QTY, the surplus is Customer Over-Plan. Uses original plan as denominator (sacrosanct).", indent=1)
    block("•  Excess at Farm: pool inventory unused at end of Dec — physical stock left over. Distinct from Customer Over-Plan.", indent=1)
    block("•  Stranded KIs: KIs with inventory in pools dataset but no plan and no demand of any kind. Truly orphan.", indent=1)
    block("•  KIs with zero forward fulfillment: broader — any KI with FwdPlan=0 AND FwdProjected=0. Includes stranded.", indent=1)
    block("•  BB / NBB: Big Box (HD/Lowes/Walmart/Other) vs Non-Big-Box (West Coast/Midwest/Southeast/MLC). NBB has access to the SS/GS restricted pre-ready pool; BB does not.", indent=1)
    block("•  Pool Conservation Identity: Start (R+NR) + Inflows (Pool + Recovery) = Forward Projected + Ending. Verified on Build Health Section 1.", indent=1)
    block("•  Walks Reconciliation: YE Net = YTD Net + Forward Net per Customer × KI. Verified on Build Health Section 1.", indent=1)
    r += 1

    block("Worksheet generated by build_norcal_workbook.py — for methodology questions, see Section 2 above.", style="note")

    freeze_at_str = "A4"
    ws.freeze_panes = freeze_at_str
    log("  Tab 1 written.")


def write_tab_changes(wb):
    """Tab: Changes — true running version history backed by changes_history.json.

    Each rebuild appends a NEW version block to the JSON file (only NET changes
    from the prior version). Older versions are preserved. The tab is rendered
    fresh from the full JSON history every build.

    Per-version interpretation_notes are optional and only appear when there's
    something interpretive to flag (e.g., methodology shifts that affect
    comparability across versions).
    """
    import json
    log("Writing Tab: Changes (running version history)...")
    ws = wb.create_sheet("Changes")
    ws["A1"] = "Changes — Version History"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Each rebuild appends a new version block at the top with only NET changes "
                "from the prior version. Older versions preserved below. Optional interpretation notes "
                "appear only when a change affects how to read the numbers vs prior versions.")
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 40
    ws.merge_cells("A2:G2")

    history_path = PROJECT_DIR / "changes_history.json"

    # Load existing history; tolerate missing/corrupt files
    try:
        if history_path.exists():
            history = json.loads(history_path.read_text())
            if not isinstance(history, list):
                history = []
        else:
            history = []
    except Exception as e:
        log(f"  WARN: could not load {history_path.name}: {e}; starting fresh")
        history = []

    # Render JSON history as-is. Updates to changes_history.json are the source of truth.
    # The hardcoded new_block generation below is left as a fallback / template only,
    # but since changes_history.json is now manually curated post each phase, we skip
    # the auto-insert and just render existing history.
    today_str = datetime.now().strftime("%Y-%m-%d")
    snap_str = M.SNAP_DATE.strftime("%Y-%m-%d")
    build_id = f"v-{today_str}-recovery"

    # Auto-generated block (fallback / template — not inserted unless history is empty)
    new_block_changes = [
        ("Tab Added", "Production Recovery (Tab 9) — action list of plan miss recoverable via production. Each row = (Customer, KI, Month) action with target item to grow, recoverable QTY/$, grow days, latest start date, slack days, and items being substituted for."),
        ("Methodology", "Plan miss now reflects only the LOCKED-IN portion. Recoverable miss (where MIN grow days fits within days-from-snap-to-month-end) is removed from miss columns workbook-wide and surfaced on the new Production Recovery tab."),
        ("Methodology", "KI-level recovery: if any item in a customer's KI has fitting grow days for a given month, the FULL KI-level miss for that (Customer, KI, Month) is recoverable via within-KI substitution. Production target = item with shortest grow days."),
        ("Methodology", "Grow days cascade fallback: Tier 1 = exact item match; Tier 2 = same KI MIN; Tier 3 = same Genus+Size MIN; Tier 4 = same Size MIN. Items with no match anywhere stay locked in."),
        ("Source", "Added load_grow_times() to engine; reads WIN and BRA tabs from Shared/Misc Look Ups/Prod lookups ALL 091925.xlsx. NOR CAL plan rows have Org Code = WIN or BRA per row, used to attribute grow days."),
        ("Cosmetic", "Tab color-coding: Production Recovery uses orange (#ED7D31) — distinct from red Miss tabs and green Lift tabs. Reflects 'action-oriented recovery' nature."),
        ("Methodology", "Outlier rule fixed: smoothed history now drops a year when its value > 10x the mean of the OTHER years (previously compared vs raw_mean inclusive of the year itself, which was mathematically impossible to trigger). Outlier Yrs column on Customer Over-Plan Detail now populates for ~9 KI x customer rows. Smoothed history values shift slightly for those rows; downstream Lift demand and Plan Miss adjust accordingly."),
        ("Tab Added", "Three new Miss tabs at KI granularity: 'Miss Summary by KI' (KI-only summary), 'Miss by Customer x KI' (one row per Customer x KI), 'Miss by KI x Month' (one row per KI x Month)."),
        ("Tab Added", "Three new Lift tabs at KI granularity: 'Lift Summary by KI' (KI-only summary), 'Lift by KI x Month' (one row per KI x Month, no customer dimension)."),
        ("Tab Renamed", "'Short Item Detail' -> 'Customer Miss Detail' (Customer x Item granularity preserved; clearer that miss is the metric)."),
        ("Tab Renamed", "'Lift Summary by KI' (was Customer x KI) -> 'Lift by Customer x KI' to match the new symmetric structure. The new 'Lift Summary by KI' is now the KI-only summary."),
        ("Tab Renamed", "'Lift Detail by Month' (Customer x Item x Month) -> 'Lift by Customer x Item' (Customer x Item, Month dimension dropped). Item-level lift detail now lives at Customer x Item granularity."),
        ("Tab Renamed", "'Customer Over-Plan by KI' -> 'Over-Plan by KI' ('Customer' prefix dropped on KI-only tabs since customer dimension isn't visible there)."),
        ("Cosmetic", "Customer Over-Plan Detail tab columns finished renaming: 'Excess QTY/$' -> 'Over-Plan QTY/$ YE'; year history columns -> longer-form like '2023 Sales QTY (full year)'; price columns -> 'Avg Sell Price 2026 YTD' / 'Avg Sell Price 2025'."),
        ("Format", "SUBTOTAL row improvements: per-unit Avg Price columns now show weighted-average formulas (Plan $ / Plan QTY) instead of meaningless sum-of-prices, where a $-and-QTY pair is available. Where no pair is available, leave blank rather than show wrong number."),
        ("Format", "SUBTOTAL row also covers all newly-added Miss/Lift tabs with proper % column ratios."),
        ("Cosmetic", "Naming convention: 'Customer' prefix kept only where customer dimension is actually visible in the rows (Customer x KI, Customer x Item, per-customer walks). Dropped on KI-only tabs and KI x Month tabs."),
        ("Cosmetic", "Tab color-coding applied: same-domain tabs share a color (Miss=red, Lift=green, Excess at Farm=gold, Customer Over-Plan=blue, Cross-cut=purple, Headline=navy, Reference=grey). Verification step at build time warns if any tab is uncolored."),
    ]

    new_block_notes = [
        "Plan miss numbers are NOT directly comparable to prior versions. Miss now reflects only LOCKED-IN portion (grow days don't fit). The recoverable portion (~$481K in this build) moved to the new Production Recovery tab as a separate action list.",
        "Tab count: 19 (added Production Recovery between Miss tabs and Lift tabs).",
        "Production Recovery uses MIN grow days per item, no fall-down inflation, no liner availability check, unlimited production capacity assumed (per CEO 2026-05-07).",
        "Sum of Production Recovery $ + remaining Miss $ may differ slightly from prior total miss due to KI weighted-average pricing on the recovery tab vs item-level pricing on the miss tabs.",
    ]

    new_block = {
        "version": today_str,
        "snapshot": snap_str,
        "build_id": build_id,
        "changes": new_block_changes,
        "interpretation_notes": new_block_notes,
    }

    # Avoid duplicate blocks if user runs the script twice in one day with no changes.
    # Compare as lists (JSON round-trips tuples to lists, so direct == fails on tuples)
    new_changes_norm = [list(c) for c in new_block_changes]
    existing_changes_norm = [list(c) for c in (history[0].get("changes") if history else [])]
    # Only auto-insert if history is empty (initial build). Otherwise trust the curated JSON.
    if not history:
        new_block["changes"] = new_changes_norm
        history.insert(0, new_block)

    # Save back to disk
    try:
        history_path.write_text(json.dumps(history, indent=2))
    except Exception as e:
        log(f"  WARN: could not write {history_path.name}: {e}")

    # Render the FULL history into the tab, newest first
    r = 4
    for v_idx, block in enumerate(history):
        sec_header = f"— Version: {block.get('version','?')} | Snapshot: {block.get('snapshot','?')} | Build: {block.get('build_id','?')} —"
        ws.cell(row=r, column=1, value=sec_header)
        ws.cell(row=r, column=1).data_type = "s"
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=("DCEAF7" if v_idx == 0 else "EDEDED"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

        ws.cell(row=r, column=1, value="Changes in this version:").font = Font(bold=True)
        r += 1
        for i, ch in enumerate(block.get("changes", []), start=1):
            cat = ch[0] if len(ch) > 0 else ""
            desc = ch[1] if len(ch) > 1 else ""
            ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="right", vertical="top")
            ws.cell(row=r, column=2, value=f"[{cat}]").font = Font(bold=True, color="245A8A")
            ws.cell(row=r, column=2).alignment = Alignment(vertical="top")
            cell = ws.cell(row=r, column=3, value=desc)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
            r += 1

        notes = block.get("interpretation_notes", []) or []
        if notes:
            r += 1
            ws.cell(row=r, column=1, value="Interpretation notes:").font = Font(bold=True)
            r += 1
            for note in notes:
                cell = ws.cell(row=r, column=2, value=note)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.row_dimensions[r].height = 22
                r += 1
        r += 2  # blank row between version blocks

    set_col_widths(ws, [5, 16, 60, 18, 18, 18])
    ws.freeze_panes = "A4"
    log(f"  Tab Changes written; {len(history)} version block(s) total.")




def write_tab2_exec(wb, m, d):
    """Tab 2: Exec Summary."""
    log("Writing Tab 2: Exec Summary...")
    ws = wb.create_sheet("Exec Summary")
    set_col_widths(ws, [4, 35, 14, 14, 14, 12, 12, 14, 14, 14])

    ki_total = m["ki_total"]
    ki_panel = m["ki_panel"]

    r = 1
    ws.cell(row=r, column=1, value="NOR CAL Forward Looking — Executive Summary").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.row_dimensions[r].height = 24
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    ws.cell(row=r, column=1, value=f"Snapshot: 2026-04-27  |  Plan vs original 2026 NOR CAL Sales Plan").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    # ---- THREE PLAN WALK SECTIONS (per Customer x KI capped) ----
    # Math reconciles by construction: Plan - Down + Up = Achieved; Net = Achieved - Plan.
    # Plan customers (HD, Lowes, Walmart, West Coast, Midwest) — from ki_panel
    walk_per_cust = ki_panel.groupby("Customer").agg(
        OrigYE_dlr=("OrigYE_dlr", "sum"),
        Miss_YE=("Miss_dlr_KI", "sum"),
        OverPlan_YE=("Customer_OverPlan_$", "sum"),
        OrigMD_dlr=("OrigMD_$", "sum"),
        Miss_MD=("Miss_MD_$", "sum"),
        OverPlan_MD=("OverPlan_MD_$", "sum"),
        FwdFill_dlr=("FwdFill_$", "sum"),
        OrigJAD_dlr=("OrigJAD_$", "sum"),
        Miss_YTD=("Miss_YTD_$", "sum"),
        OverPlan_YTD=("OverPlan_YTD_$", "sum"),
        YTD_dlr=("YTD_$", "sum"),
    )

    # QA #65: Add off-plan customers (Southeast, MLC, Other) to walks.
    # Off-plan have no plan, so Plan=Miss=0; everything they ship is over-plan.
    plan_custs_set = set(walk_per_cust.index)
    ytd_df = d.get("ytd")
    offplan_alloc_df = d.get("offplan_alloc", pd.DataFrame())
    items_df = d.get("items")
    item_to_ki = items_df.drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict() if items_df is not None else {}

    if ytd_df is not None and not ytd_df.empty:
        ytd_offplan = ytd_df[~ytd_df["Customer"].isin(plan_custs_set) & ytd_df["Customer"].notna()].copy()
        # YTD has Revenue column; if not, use Qty × KI avg price
        ytd_offplan["KI"] = ytd_offplan["Item Num"].map(item_to_ki)
        # Filter to items with KI mapping (consistent with Lifecycle aggregation; orphan items excluded both places)
        ytd_offplan = ytd_offplan.dropna(subset=["KI"])
        if "Revenue" in ytd_offplan.columns:
            ytd_op_dlr = ytd_offplan.groupby("Customer")["Revenue"].sum().to_dict()
        else:
            avg_price_per_ki = m["ki_total"].set_index("KI")["AvgPrice"].to_dict() if "ki_total" in m else {}
            ytd_offplan["AvgPrice"] = ytd_offplan["KI"].map(avg_price_per_ki).fillna(0)
            ytd_offplan["YTD_$"] = ytd_offplan["Qty"] * ytd_offplan["AvgPrice"]
            ytd_op_dlr = ytd_offplan.groupby("Customer")["YTD_$"].sum().to_dict()
    else:
        ytd_op_dlr = {}

    if not offplan_alloc_df.empty:
        avg_price_per_ki = m["ki_total"].set_index("KI")["AvgPrice"].to_dict() if "ki_total" in m else {}
        op_with_price = offplan_alloc_df.copy()
        op_with_price["AvgPrice"] = op_with_price["KI"].map(avg_price_per_ki).fillna(0)
        op_with_price["FwdFill_$"] = op_with_price["Stage2Filled"] * op_with_price["AvgPrice"]
        fwd_op_dlr = op_with_price.groupby("Customer")["FwdFill_$"].sum().to_dict()
    else:
        fwd_op_dlr = {}

    # Add off-plan rows to walk_per_cust
    offplan_custs = sorted(set(ytd_op_dlr.keys()) | set(fwd_op_dlr.keys()))
    for op_cust in offplan_custs:
        ytd_d = float(ytd_op_dlr.get(op_cust, 0))
        fwd_d = float(fwd_op_dlr.get(op_cust, 0))
        # Off-plan customers: Plan=0, Miss=0, OverPlan=full ship $
        walk_per_cust.loc[op_cust] = {
            "OrigYE_dlr": 0.0,
            "Miss_YE": 0.0,
            "OverPlan_YE": ytd_d + fwd_d,
            "OrigMD_dlr": 0.0,
            "Miss_MD": 0.0,
            "OverPlan_MD": fwd_d,
            "FwdFill_dlr": fwd_d,
            "OrigJAD_dlr": 0.0,
            "Miss_YTD": 0.0,
            "OverPlan_YTD": ytd_d,
            "YTD_dlr": ytd_d,
        }

    # Customer iteration order: plan customers first, off-plan after
    walk_customer_order = list(CUSTOMERS) + offplan_custs

    def _write_walk(r, section_title, plan_key, miss_key, over_key, period_lbl):
        # Pre-clear any leftover styling on this row's columns 1-10 (defensive — the
        # very first walk's row 4 sometimes inherits HEADER styling from a prior
        # workbook-level title block, which leaked through merge_cells)
        for _col_idx in range(1, 11):
            ws.cell(row=r, column=_col_idx).fill = SECTION_FILL
            ws.cell(row=r, column=_col_idx).font = SECTION_FONT
        ws.cell(row=r, column=1, value=section_title).font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1
        # YTD is real actuals (Jan-Apr); YE and May-Dec involve forecast → "Projected"
        achieved_label = f"Actual $ {period_lbl}" if period_lbl == "YTD" else f"Projected $ {period_lbl}"
        hdr = ["", "Customer",
               f"Original Plan $ {period_lbl}",
               f"Plan Miss $ {period_lbl}",
               f"Customer Over-Plan $ {period_lbl}",
               achieved_label,
               f"Net $ {period_lbl}",
               f"Net % {period_lbl}"]
        for j, h in enumerate(hdr, start=1):
            ws.cell(row=r, column=j, value=h)
        apply_header(ws, r, 8, height=42)
        r += 1
        data_start = r
        for c in walk_customer_order:
            if c not in walk_per_cust.index: continue
            row = walk_per_cust.loc[c]
            plan_v = float(row[plan_key])
            miss_v = float(row[miss_key])
            over_v = float(row[over_key])
            ach_v  = plan_v - miss_v + over_v
            net_v  = ach_v - plan_v
            net_pct= net_v / plan_v if plan_v else 0.0
            ws.cell(row=r, column=2, value=c).font = NORMAL_FONT
            ws.cell(row=r, column=3, value=plan_v).number_format = FMT_DLR
            ws.cell(row=r, column=4, value=miss_v).number_format = FMT_DLR
            if miss_v > 0: ws.cell(row=r, column=4).fill = NEG_FILL
            ws.cell(row=r, column=5, value=over_v).number_format = FMT_DLR
            if over_v > 0: ws.cell(row=r, column=5).fill = POS_FILL
            ws.cell(row=r, column=6, value=ach_v).number_format = FMT_DLR
            ws.cell(row=r, column=7, value=net_v).number_format = FMT_DLR
            if net_v < -1: ws.cell(row=r, column=7).fill = NEG_FILL
            elif net_v > 1: ws.cell(row=r, column=7).fill = POS_FILL
            ws.cell(row=r, column=8, value=net_pct).number_format = FMT_PCT
            r += 1
        # TOTAL row with SUM formulas (so row stays dynamic across customer changes)
        ws.cell(row=r, column=2, value="TOTAL").font = BOLD_FONT
        ws.cell(row=r, column=2).fill = TOTAL_FILL
        for col_idx in [3, 4, 5, 6, 7]:
            cl = get_column_letter(col_idx)
            ws.cell(row=r, column=col_idx, value=f"=SUM({cl}{data_start}:{cl}{r-1})").number_format = FMT_DLR
            ws.cell(row=r, column=col_idx).font = BOLD_FONT
            ws.cell(row=r, column=col_idx).fill = TOTAL_FILL
        ws.cell(row=r, column=8, value=f"=G{r}/C{r}").number_format = FMT_PCT
        ws.cell(row=r, column=8).font = BOLD_FONT
        ws.cell(row=r, column=8).fill = TOTAL_FILL
        return r + 2

    # Walk 1 — YE (Jan-Dec)
    r = _write_walk(r,
        "Walk 1 — Year-End (Jan-Dec): Original Plan -> Plan Miss [Down] -> Customer Over-Plan [Up] -> Net YE Position",
        "OrigYE_dlr", "Miss_YE", "OverPlan_YE", "YE")

    # Walk 2 — Forward May-Dec
    r = _write_walk(r,
        "Walk 2 — Forward (May-Dec): Forward Plan -> Forward Miss [Down] -> Forward Over-Plan [Up] -> Net Forward Position",
        "OrigMD_dlr", "Miss_MD", "OverPlan_MD", "May-Dec")

    # Walk 3 — YTD (Jan-Apr)
    r = _write_walk(r,
        "Walk 3 — Year-to-Date (Jan-Apr): YTD Plan -> YTD Miss [Down] -> YTD Over-Plan [Up] -> Net YTD Position",
        "OrigJAD_dlr", "Miss_YTD", "OverPlan_YTD", "YTD")

    # Top 20 KI by Miss $
    ws.cell(row=r, column=1, value="Top 20 Key Item Groupings by Miss $").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    headers = ["#", "Key Item Grouping", "Original Plan $ YE", "Plan Miss $ YE", "YE Miss %", "Plan QTY YE", "Miss QTY YE",
                "Avg Price $", "YTD %", "Inflow Total"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 10)
    r += 1

    # Use Miss_dlr_KI_total (customer-aware, sum of per-customer Miss × per-customer AvgPrice)
    # for consistency with Plan by KI and Miss Summary by KI.
    top20_miss = ki_total.sort_values("Miss_dlr_KI_total", ascending=False).head(20)
    avg_price_per_ki = ki_panel.groupby("KI").apply(
        lambda g: (g["AvgPrice"]*g["OrigYE_QTY"]).sum()/g["OrigYE_QTY"].sum() if g["OrigYE_QTY"].sum()>0 else 0,
        include_groups=False,
    ).to_dict()
    for i, rr in enumerate(top20_miss.itertuples(), start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=rr.KI)
        ws.cell(row=r, column=3, value=float(rr.OrigYE_dlr)).number_format = FMT_DLR
        miss_dlr = float(rr.Miss_dlr_KI_total)
        ws.cell(row=r, column=4, value=miss_dlr).number_format = FMT_DLR
        ws.cell(row=r, column=4).fill = NEG_FILL
        miss_pct_recalc = miss_dlr / float(rr.OrigYE_dlr) if float(rr.OrigYE_dlr) > 0 else 0.0
        ws.cell(row=r, column=5, value=miss_pct_recalc).number_format = FMT_PCT
        ws.cell(row=r, column=6, value=float(rr.OrigYE_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr.Miss_QTY_KI)).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(avg_price_per_ki.get(rr.KI, 0))).number_format = FMT_DLR2
        ytd_pct = rr.YTD_QTY / rr.OrigYE_QTY if rr.OrigYE_QTY else 0
        ws.cell(row=r, column=9, value=float(ytd_pct)).number_format = FMT_PCT
        ws.cell(row=r, column=10, value=float(rr.FwdFill_QTY)).number_format = FMT_INT
        r += 1
    r += 1

    # Top 20 KI by Excess $
    ws.cell(row=r, column=1, value="Top 20 Key Item Groupings by Over-Plan $ (KIs where customers ship above plan)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    headers = ["#", "Key Item Grouping", "Original Plan $ YE", "Over-Plan $ YE", "Over-Plan QTY YE", "Original Plan QTY YE",
                "Lifted YE Plan QTY", "Smoothed History QTY (3-yr, full year)", "Avg Price $", "Lift Applied $ (May-Dec)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 10)
    r += 1

    top20_exc = ki_total.sort_values("Excess_dlr_vsPlan", ascending=False).head(20)
    for i, rr in enumerate(top20_exc.itertuples(), start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=rr.KI)
        ws.cell(row=r, column=3, value=float(rr.OrigYE_dlr)).number_format = FMT_DLR
        ws.cell(row=r, column=4, value=float(rr.Excess_dlr_vsPlan)).number_format = FMT_DLR
        ws.cell(row=r, column=4).fill = POS_FILL
        ws.cell(row=r, column=5, value=float(rr.Excess_QTY_vsPlan)).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr.OrigYE_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr.LiftedYE_QTY) if hasattr(rr, "LiftedYE_QTY") else float(rr.LiftedMD_QTY + rr.OrigJAD_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(rr.Smoothed_Hist)).number_format = FMT_INT
        ws.cell(row=r, column=9, value=float(avg_price_per_ki.get(rr.KI, 0))).number_format = FMT_DLR2
        ws.cell(row=r, column=10, value=float(rr.Lift_dlr)).number_format = FMT_DLR
        r += 1
    r += 1

    # ---- NEW: Top 20 KIs by Excess at Farm $ (pool unused at YE) ----
    ws.cell(row=r, column=1, value="Top 20 Key Item Groupings by Excess at Farm $ (inventory pool unused at end of December)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    headers3 = ["#", "Key Item Grouping", "Excess at Farm QTY (end of Dec)", "Avg Price $", "Excess at Farm $ (end of Dec)",
                 "% of Total Excess at Farm $", "", "", "", ""]
    for j, h in enumerate(headers3, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 6)
    r += 1
    farm = m.get("excess_at_farm")
    if farm is not None and not farm.empty:
        total_farm_dlr = float(farm["ExcessAtFarm_$"].sum())
        top20_farm = farm.head(20).reset_index(drop=True)
        rank = 1
        for _, frow in top20_farm.iterrows():
            qty = float(frow["ExcessAtFarm_QTY"])
            if qty <= 0:
                continue
            farm_dlr = float(frow["ExcessAtFarm_$"])
            ws.cell(row=r, column=1, value=rank)
            ws.cell(row=r, column=2, value=frow["KI"])
            ws.cell(row=r, column=3, value=qty).number_format = FMT_INT
            ws.cell(row=r, column=4, value=float(frow["AvgPrice"])).number_format = FMT_DLR2
            ws.cell(row=r, column=5, value=farm_dlr).number_format = FMT_DLR
            pct = (farm_dlr / total_farm_dlr) if total_farm_dlr > 0 else 0
            ws.cell(row=r, column=6, value=float(pct)).number_format = FMT_PCT
            r += 1
            rank += 1

    freeze_at_str = "A5"
    ws.freeze_panes = freeze_at_str
    log("  Tab 2 written.")


def write_tab_plan_by_ki(wb, m):
    """Tab: Plan by KI — full universe, walk format with Below/Above/Net columns."""
    log("Writing Tab: Plan by KI...")
    ws = wb.create_sheet("Plan by KI")
    set_col_widths(ws, [4, 35, 15, 15, 15, 13, 11, 11, 11, 11, 11, 11, 11, 13])

    ki_panel = m["ki_panel"]
    ki_total = m["ki_total"]

    # Build per-customer Net pivot
    ki_panel["NetByCust"] = (ki_panel["YTD_QTY"] + ki_panel["FwdFill_QTY"] - ki_panel["OrigYE_QTY"])
    ki_panel["NetByCust_dlr"] = ki_panel["NetByCust"] * ki_panel["AvgPrice"]
    cust_pivot = ki_panel.pivot(index="KI", columns="Customer", values="NetByCust_dlr").fillna(0)
    for c in CUSTOMERS:
        if c not in cust_pivot.columns:
            cust_pivot[c] = 0
    # QA #69: include off-plan customers (Southeast/MLC/Other) — Net = YTD+FwdFill (since plan = 0)
    OFFPLAN_CUSTS = ["Southeast", "MLC", "Other"]
    op_net_pivot = m.get("_offplan_net_pivot")
    for op_c in OFFPLAN_CUSTS:
        if op_net_pivot is not None and not op_net_pivot.empty and op_c in op_net_pivot.columns:
            cust_pivot[op_c] = op_net_pivot[op_c].reindex(cust_pivot.index).fillna(0)
        else:
            cust_pivot[op_c] = 0

    # Merge with ki_total which now has Miss_dlr_KI_total + OverPlan_dlr_KI_total
    full = ki_total.merge(cust_pivot.reset_index(), on="KI", how="left")
    # Compute net per KI
    full["Net_dlr"] = full.get("OverPlan_dlr_KI_total", 0).fillna(0) - full.get("Miss_dlr_KI_total", 0).fillna(0)
    full["Net_pct"] = np.where(full["OrigYE_dlr"]>0, full["Net_dlr"]/full["OrigYE_dlr"], 0)
    # Sort by absolute Net descending so biggest movers (either direction) appear first
    full["abs_net"] = full["Net_dlr"].abs()
    full = full.sort_values("abs_net", ascending=False).drop(columns=["abs_net"])

    r = 1
    ws.cell(row=r, column=1, value="Plan by Key Item Grouping (full universe — Plan, Plan Miss [Down], Customer Over-Plan [Up], Net)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Each KI either has Plan Miss > 0 (down) OR Customer Over-Plan > 0 (up). Sorted by absolute size of Net $. Net = Over-Plan - Plan Miss (signed). Customer columns (8 channels: 5 plan + 3 off-plan) show net Up/Down per customer × KI in $ (signed).").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
    ws.row_dimensions[r].height = 30
    r += 2

    headers = ["#", "Key Item Grouping",
               "Original Plan $ YE", "Plan Miss $ YE", "Over-Plan $ YE",
               "Net $ YE", "Net % YE", "Avg Price $",
               "HD", "Lowes", "Walmart", "West Coast", "Midwest",
               "Southeast", "MLC", "Other",
               "Total Net $"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 17, height=42)
    header_row = r
    r += 1
    data_start = r

    avg_price_per_ki = ki_panel.groupby("KI").apply(
        lambda g: (g["AvgPrice"]*g["OrigYE_QTY"]).sum()/g["OrigYE_QTY"].sum() if g["OrigYE_QTY"].sum()>0 else 0,
        include_groups=False,
    ).to_dict()

    for i, ki_row in full.iterrows():
        ws.cell(row=r, column=1, value=int(i)+1 if False else None)
    # Use itertuples for clean iteration
    for idx, ki_row in enumerate(full.itertuples(), start=1):
        miss_v = float(getattr(ki_row, "Miss_dlr_KI_total", 0) or 0)
        over_v = float(getattr(ki_row, "OverPlan_dlr_KI_total", 0) or 0)
        plan_v = float(ki_row.OrigYE_dlr)
        net_v  = over_v - miss_v
        net_pct = net_v / plan_v if plan_v else 0.0

        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=ki_row.KI)
        ws.cell(row=r, column=3, value=plan_v).number_format = FMT_DLR
        ws.cell(row=r, column=4, value=miss_v).number_format = FMT_DLR
        if miss_v > 0: ws.cell(row=r, column=4).fill = NEG_FILL
        ws.cell(row=r, column=5, value=over_v).number_format = FMT_DLR
        if over_v > 0: ws.cell(row=r, column=5).fill = POS_FILL
        ws.cell(row=r, column=6, value=net_v).number_format = FMT_DLR
        if net_v < -1: ws.cell(row=r, column=6).fill = NEG_FILL
        elif net_v > 1: ws.cell(row=r, column=6).fill = POS_FILL
        ws.cell(row=r, column=7, value=net_pct).number_format = FMT_PCT
        ws.cell(row=r, column=8, value=float(avg_price_per_ki.get(ki_row.KI, 0))).number_format = FMT_DLR2

        # Customer columns 9-16 (5 plan + 3 off-plan; QA #69)
        all_custs_for_cols = list(CUSTOMERS) + ["Southeast", "MLC", "Other"]
        col_map = {c: 9+i for i, c in enumerate(all_custs_for_cols)}
        for c in all_custs_for_cols:
            rval = full[full["KI"]==ki_row.KI][c].values if c in full.columns else [0]
            v = float(rval[0]) if len(rval) > 0 else 0
            cell = ws.cell(row=r, column=col_map[c], value=v)
            cell.number_format = FMT_DLR
            if v < -1: cell.fill = NEG_FILL
            elif v > 1: cell.fill = POS_FILL
        # Total Net column 17 = sum of customer cells
        cells = [get_column_letter(col_map[c]) for c in all_custs_for_cols]
        ws.cell(row=r, column=17, value=f"={'+'.join(cl+str(r) for cl in cells)}").number_format = FMT_DLR
        r += 1
    last_data_row = r - 1

    # TOTAL row — sum aggregate cols (3-6), customer cols (9-16), Total Net (17)
    ws.cell(row=r, column=2, value="TOTAL").font = BOLD_FONT
    ws.cell(row=r, column=2).fill = TOTAL_FILL
    for col_idx in [3, 4, 5, 6, 9, 10, 11, 12, 13, 14, 15, 16, 17]:
        cl = get_column_letter(col_idx)
        ws.cell(row=r, column=col_idx, value=f"=SUM({cl}{data_start}:{cl}{last_data_row})").number_format = FMT_DLR
        ws.cell(row=r, column=col_idx).font = BOLD_FONT
        ws.cell(row=r, column=col_idx).fill = TOTAL_FILL
    ws.cell(row=r, column=7, value=f"=F{r}/C{r}").number_format = FMT_PCT
    ws.cell(row=r, column=7).font = BOLD_FONT
    ws.cell(row=r, column=7).fill = TOTAL_FILL
    # (extra TOTAL block removed; already handled above)


    # Autofilter on data range only (not TOTAL row)
    try:
        if last_data_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:Q{last_data_row}"
    except Exception:
        pass
    ws.freeze_panes = f"A{header_row+1}"
    log(f"  Tab Plan by KI written ({last_data_row - data_start + 1} KIs).")


# Stub writers for remaining tabs — written more concisely

def write_tab4_short_items(wb, m, d):
    log("Writing Tab 4: Short Item Detail...")
    ws = wb.create_sheet("Customer Miss Detail")
    set_col_widths(ws, [4, 14, 14, 30, 25, 12, 11, 11, 11, 11, 11, 11, 12, 50])

    panel = m["panel"]
    pool_walk = d["pool_walk"]

    # Compute monthly inflow per KI and demand per KI for "timing mismatch" detection
    inflow_total = pool_walk.groupby("KI")[["InflowGrad","InflowSNGN"]].sum()
    inflow_total["Inflow_Total"] = inflow_total["InflowGrad"] + inflow_total["InflowSNGN"]
    demand_total = pool_walk.groupby("KI")[["Stage1_BB_Demand","Stage1_NBB_Demand"]].sum()
    demand_total["Demand_Total"] = demand_total["Stage1_BB_Demand"] + demand_total["Stage1_NBB_Demand"]

    # Compute reason flags for each item with miss
    short = panel[panel["Miss_QTY"] > 0].copy()

    def reason_flags(row, ytd_plan_jad_qty, inv_total_md, demand_md):
        flags = []
        miss_dlr = row["Miss_$"]
        # YTD shortfall portion
        ytd_short_qty = max(0, ytd_plan_jad_qty - row["YTD_QTY"])
        if ytd_short_qty > 0:
            ytd_short_dlr = ytd_short_qty * row["Price"]
            flags.append(("YTD shortfall", ytd_short_dlr))

        # Forward gap portion
        fwd_gap_qty = max(0, row["OrigMD_QTY"] - row["FwdFill_QTY"])
        fwd_gap_dlr = fwd_gap_qty * row["Price"]
        if fwd_gap_qty > 0:
            # Production gap vs Timing mismatch
            if row["Customer"] in BB_CUSTOMERS and demand_md > 0:
                # If KI inventory total >= demand total but item still short = NBB-restricted
                # If inventory < demand = production gap
                if inv_total_md < demand_md * 0.95:
                    flags.append(("Production gap", fwd_gap_dlr))
                else:
                    flags.append(("Non-Big-Box (West Coast/Midwest only)", fwd_gap_dlr))
            else:
                flags.append(("Production gap", fwd_gap_dlr))

        # History-capped: ytd shortfall but no lift (LiftedMD ≈ OrigMD)
        if (row["LiftedMD_QTY"] - row["OrigMD_QTY"]) < 1 and ytd_short_qty > 0:
            # already YTD shortfall flagged; we'll add note
            pass

        if not flags:
            return ""
        # Format
        parts = []
        for f, d_ in flags:
            parts.append(f"({f}: ${d_:,.0f})")
        return " ".join(parts)

    # Build YTD plan QTY by Customer × Item
    ytd_plan_qty = panel.set_index(["Customer","Item Num"])["OrigJAD_QTY"].to_dict()
    short["Reason"] = short.apply(
        lambda row: reason_flags(
            row,
            ytd_plan_qty.get((row["Customer"], row["Item Num"]), 0),
            inflow_total.loc[row["KI"], "Inflow_Total"] if row["KI"] in inflow_total.index else 0,
            demand_total.loc[row["KI"], "Demand_Total"] if row["KI"] in demand_total.index else 0,
        ), axis=1
    )

    short = short.sort_values("Miss_$", ascending=False)

    r = 1
    ws.cell(row=r, column=1, value="Customer Miss Detail — Per-Customer x Item Miss (before within-KI substitution)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="One row per (Customer × Item) with miss > 0. 'Why' column has multi-flag with $ impact: (Flag: $X)").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    r += 1
    sub_note = ("⚠ Miss $ on this tab is ITEM-LEVEL gross — does NOT account for intra-KI substitution. "
                "For KI-level miss after substitution (the revenue forecasting view), see the Plan by KI tab "
                "or the Exec Summary headline walks. The two numbers differ when items within a KI offset each other.")
    ws.cell(row=r, column=1, value=sub_note).font = NOTE_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.row_dimensions[r].height = 40
    r += 2

    headers = ["#", "Customer", "Item Num", "Item Description", "Key Item Grouping",
                "Plan QTY YE", "YTD QTY", "Fwd Fill QTY (May-Dec)", "Miss QTY YE",
                "Plan $ YE", "Miss $ YE", "YE Miss %", "Avg Price $", "Why (multi-flag with $ impact)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 14, height=30)
    header_row = r
    r += 1

    for i, rr in enumerate(short.head(500).itertuples(), start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=rr.Customer)
        ws.cell(row=r, column=3, value=rr._6 if hasattr(rr,'_6') else rr[3])  # Item Num
        # Use direct attribute access
        ws.cell(row=r, column=3, value=getattr(rr, "Item_Num", rr[3]) if hasattr(rr, "Item_Num") else getattr(rr, '_3', None))
        # Safer: use the dataframe row
        idx = rr.Index
        srow = short.loc[idx]
        ws.cell(row=r, column=3, value=srow["Item Num"])
        ws.cell(row=r, column=4, value=str(srow.get("Item Desc", ""))[:40])
        ws.cell(row=r, column=5, value=srow["KI"])
        ws.cell(row=r, column=6, value=float(srow["OrigYE_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(srow["YTD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(srow["FwdFill_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=9, value=float(srow["Miss_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=10, value=float(srow["OrigYE_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=11, value=float(srow["Miss_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=11).fill = NEG_FILL
        miss_pct = srow["Miss_$"] / srow["OrigYE_$"] if srow["OrigYE_$"] else 0
        ws.cell(row=r, column=12, value=float(miss_pct)).number_format = FMT_PCT
        ws.cell(row=r, column=13, value=float(srow["Price"])).number_format = FMT_DLR2
        ws.cell(row=r, column=14, value=srow["Reason"])
        r += 1

    ws.freeze_panes = f"A{header_row+1}"
    # Autofilter on data range — span actual data columns
    try:
        from openpyxl.utils import get_column_letter
        end_row = r - 1
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{end_row}"
    except Exception:
        pass
    log("  Tab 4 written.")


def write_tab_lift_summary(wb, m, d):
    """Lift Summary by Customer x Key Item Grouping. Forward window only (May-Dec)."""
    log("Writing Tab: Lift Summary by KI...")
    ws = wb.create_sheet("Lift Summary by KI")
    set_col_widths(ws, [4, 14, 35, 18, 18, 14, 12, 14, 18, 30])

    ki_panel = m["ki_panel"]
    lift_ki = ki_panel[ki_panel["Lift_QTY"] > 0].copy().sort_values("Lift_dlr", ascending=False)

    r = 1
    ws.cell(row=r, column=1, value="Lift Summary by Customer x Key Item Grouping (May-Dec)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Where the model added demand based on smoothed historical pattern (May-Dec only). Lift = max(0, smoothed_history - plan).").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = ["#", "Customer", "Key Item Grouping",
               "Plan QTY May-Dec orig", "Lifted Plan QTY May-Dec",
               "Lift QTY May-Dec", "Lift % May-Dec", "Lift $ May-Dec",
               "Smoothed History QTY (3-yr avg, May-Dec)", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 10, height=42)
    header_row = r
    r += 1
    data_start = r

    for i, rr in enumerate(lift_ki.itertuples(), start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=rr.Customer)
        ws.cell(row=r, column=3, value=rr.KI)
        ws.cell(row=r, column=4, value=float(rr.OrigMD_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr.LiftedMD_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr.Lift_QTY)).number_format = FMT_INT
        lift_pct = rr.Lift_QTY / rr.OrigMD_QTY if rr.OrigMD_QTY else 0
        ws.cell(row=r, column=7, value=float(lift_pct)).number_format = FMT_PCT
        ws.cell(row=r, column=8, value=float(rr.Lift_dlr)).number_format = FMT_DLR
        ws.cell(row=r, column=8).fill = POS_FILL
        ws.cell(row=r, column=9, value=float(rr.Smoothed_Hist)).number_format = FMT_INT
        outliers = rr.Outlier_Years if hasattr(rr, "Outlier_Years") else ""
        notes = f"Smoothed history > plan by {lift_pct:.0%}" + (f"; outliers dropped: {outliers}" if outliers else "")
        ws.cell(row=r, column=10, value=notes)
        r += 1
    last_data_row = r - 1

    # Subtotal row will be added by helper later
    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"

    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Lift Summary by KI written ({last_data_row - data_start + 1} rows).")


def write_tab_lift_detail_by_month(wb, m, d):
    """Lift Detail by Customer x Item x Month (drill-down). Forward window only (May-Dec)."""
    log("Writing Tab: Lift Detail by Month...")
    ws = wb.create_sheet("Lift Detail by Month")
    set_col_widths(ws, [4, 14, 14, 35, 25, 8, 18, 18, 14, 12, 14, 22, 12, 28])

    plan_lifted = d["plan_lifted"]
    items = d["items"]
    panel = m["panel"]

    r = 1
    ws.cell(row=r, column=1, value="Lift Detail — Customer x Item x Month (May-Dec drill-down)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="One row per (Customer, Item Num, Month) where the model lifted plan above the original. Use this tab to audit any specific lift.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = ["#", "Customer", "Item Num", "Item Description", "Key Item Grouping", "Month",
               "Plan QTY May-Dec orig", "Lifted Plan QTY May-Dec",
               "Lift QTY May-Dec", "Lift % May-Dec", "Lift $ May-Dec",
               "Smoothed History QTY (3-yr avg, May-Dec)", "Avg Price $", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 14, height=42)
    header_row = r
    r += 1
    data_start = r

    pl = plan_lifted.copy()
    pl = pl[pl["MonthNum"].isin(FWD_MONTHS)]
    pl = pl[pl["LiftedQty"] - pl["PlanQty"] > 0.01]
    pl = pl.merge(items[["Item Num","Item Desc"]].drop_duplicates(subset=["Item Num"]), on="Item Num", how="left")
    pl = pl.merge(panel[["Customer","Item Num","Price"]].drop_duplicates(subset=["Customer","Item Num"]), on=["Customer","Item Num"], how="left")
    pl["Lift QTY"] = pl["LiftedQty"] - pl["PlanQty"]
    pl["Lift Factor"] = pl["LiftedQty"] / pl["PlanQty"]
    pl = pl.sort_values(["Customer","KI","Item Num","MonthNum"])

    # Pre-build smoothed history per (Customer, LiftGroup, Month) lookup
    sm = m["smoothed"][["Customer","LiftGroup","Month","smoothed"]].copy() if "smoothed" in m else None
    if sm is None:
        # Compute on-the-fly fallback
        sm = d["smoothed"][["Customer","LiftGroup","Month","smoothed"]].copy()
    sm_lookup = {(row["Customer"], row["LiftGroup"], int(row["Month"])): float(row["smoothed"]) for _, row in sm.iterrows()}

    # Pre-build per (Customer, KI) Notes lookup from Lift Summary's notes (same text per group)
    ki_panel_local = m["ki_panel"]
    ki_panel_local["_lift_pct"] = np.where(ki_panel_local["OrigMD_QTY"]>0, ki_panel_local["Lift_QTY"]/ki_panel_local["OrigMD_QTY"], 0)
    notes_lookup = {(row["Customer"], row["KI"]): f"Smoothed history > plan by {row['_lift_pct']:.0%}"
                    for _, row in ki_panel_local.iterrows() if row["Lift_QTY"] > 0}

    # LiftGroup attached to plan_lifted via lift_groups merge — pl needs LiftGroup
    if "LiftGroup" not in pl.columns:
        pl = pl.merge(d["lift_groups"][["Customer","Item Num","LiftGroup"]],
                      on=["Customer","Item Num"], how="left")

    rank = 1
    for _, rr in pl.iterrows():
        plan_q = float(rr["PlanQty"])
        lifted_q = float(rr["LiftedQty"])
        lift_q = float(rr["Lift QTY"])
        lift_pct = (lift_q / plan_q) if plan_q > 0 else 0.0
        price = float(rr.get("Price", 0) or 0)
        lift_dlr = lift_q * price
        sm_q = sm_lookup.get((rr["Customer"], rr.get("LiftGroup"), int(rr["MonthNum"])), 0.0)
        notes = notes_lookup.get((rr["Customer"], rr["KI"]), "")

        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["Item Num"])
        ws.cell(row=r, column=4, value=str(rr.get("Item Desc", ""))[:40])
        ws.cell(row=r, column=5, value=rr["KI"])
        ws.cell(row=r, column=6, value=MONTH_LBL.get(int(rr["MonthNum"]), rr["MonthNum"]))
        ws.cell(row=r, column=7, value=plan_q).number_format = FMT_INT
        ws.cell(row=r, column=8, value=lifted_q).number_format = FMT_INT
        ws.cell(row=r, column=9, value=lift_q).number_format = FMT_INT
        ws.cell(row=r, column=10, value=lift_pct).number_format = FMT_PCT
        ws.cell(row=r, column=11, value=lift_dlr).number_format = FMT_DLR
        ws.cell(row=r, column=11).fill = POS_FILL
        ws.cell(row=r, column=12, value=sm_q).number_format = FMT_INT
        ws.cell(row=r, column=13, value=price).number_format = FMT_DLR2
        ws.cell(row=r, column=14, value=notes)
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"

    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Lift Detail by Month written ({last_data_row - data_start + 1} rows).")


def write_tab6_excess_ki(wb, m):
    log("Writing Tab 6: Excess / Upside by KI...")
    ws = wb.create_sheet("Over-Plan by KI")
    set_col_widths(ws, [4, 35, 12, 12, 12, 12, 12, 12, 12, 12, 30, 30])

    ki_total = m["ki_total"]
    ki_panel = m["ki_panel"]
    # QA #66: standardize to Method B (OverPlan_dlr_KI_total = customer-aware sum)
    # rather than Method C (Excess_dlr_vsPlan = KI×AvgPrice).
    excess_kis = ki_total[ki_total["OverPlan_dlr_KI_total"] > 0].copy().sort_values("OverPlan_dlr_KI_total", ascending=False)

    r = 1
    ws.cell(row=r, column=1, value="Over-Plan by Key Item Grouping (ranked by $ vs Original Plan)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Three over-plan thresholds shown: vs Original Plan (most aggressive), vs Smoothed History, vs Lifted YE Plan (most conservative). 'Why no further lift' explains why model didn't add more demand.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 2

    headers = ["#", "Key Item Grouping", "Original Plan QTY YE", "Lifted YE Plan QTY", "Smoothed Hist QTY (3-yr, full year)",
                "Over-Plan QTY vs Original Plan YE", "Over-Plan $ vs Original Plan YE", "Over-Plan QTY vs Hist", "Over-Plan QTY vs Lifted YE",
                "Avg Price $", "Why no further lift", "Flags"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 12, height=32)
    header_row = r
    r += 1

    avg_price_per_ki = ki_panel.groupby("KI").apply(
        lambda g: (g["AvgPrice"]*g["OrigYE_QTY"]).sum()/g["OrigYE_QTY"].sum() if g["OrigYE_QTY"].sum()>0 else 0,
        include_groups=False,
    ).to_dict()

    for i, rr in enumerate(excess_kis.itertuples(), start=1):
        # Why no further lift
        if rr.Lift_QTY > 0:
            why = "Lift triggered partially (smoothed history > plan, but inventory still exceeds lifted plan)"
        elif rr.Smoothed_Hist <= rr.OrigMD_QTY:
            why = "Plan ≥ smoothed history (no lift trigger)"
        else:
            why = "Lift didn't trigger — possible HD/Lowes SKU-level mismatch"

        # Flags
        ki_data = ki_panel[ki_panel["KI"]==rr.KI]
        ytd_pace = (ki_data["YTD_QTY"].sum() / ki_data["OrigJAD_QTY"].sum()) if ki_data["OrigJAD_QTY"].sum() > 0 else 0
        flags = []
        if ytd_pace < 0.85:
            flags.append("YTD pace below plan")
        if ki_data["YTD_2025_QTY"].sum() > 0:
            yoy = ki_data["YTD_QTY"].sum() / ki_data["YTD_2025_QTY"].sum()
            if yoy < 0.9:
                flags.append("YoY down")
            elif yoy > 1.1:
                flags.append("YoY up")
        if rr.Smoothed_Hist > rr.OrigMD_QTY * 1.2:
            flags.append("History exceeds plan >20%")

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=rr.KI)
        ws.cell(row=r, column=3, value=float(rr.OrigYE_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=4, value=float(rr.LiftedYE_QTY)).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr.Smoothed_Hist)).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr.Excess_QTY_vsPlan)).number_format = FMT_INT
        ws.cell(row=r, column=6).fill = POS_FILL
        ws.cell(row=r, column=7, value=float(rr.OverPlan_dlr_KI_total)).number_format = FMT_DLR
        ws.cell(row=r, column=7).fill = POS_FILL
        ws.cell(row=r, column=8, value=float(max(0, rr.Excess_QTY_vsHist))).number_format = FMT_INT
        ws.cell(row=r, column=9, value=float(max(0, rr.Excess_QTY_vsLifted))).number_format = FMT_INT
        ws.cell(row=r, column=10, value=float(avg_price_per_ki.get(rr.KI, 0))).number_format = FMT_DLR2
        ws.cell(row=r, column=11, value=why)
        ws.cell(row=r, column=12, value="; ".join(flags))
        r += 1

    ws.freeze_panes = f"A{header_row+1}"
    try:
        from openpyxl.utils import get_column_letter
        end_row = r - 1
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{end_row}"
    except Exception:
        pass
    log("  Tab 6 written.")


def write_tab7_excess_cust_ki(wb, m, d):
    log("Writing Tab 7: Excess by Customer × KI...")
    ws = wb.create_sheet("Customer Over-Plan Detail")
    set_col_widths(ws, [4, 14, 35, 12, 12, 12, 11, 11, 11, 11, 11, 11, 12, 11, 11, 11, 12, 30, 35])

    ki_panel = m["ki_panel"]
    panel = m["panel"]

    # Filter to (Customer, KI) where Excess_QTY_vsPlan > 0
    ex = ki_panel.copy()
    ex["Excess_QTY"] = (ex["YTD_QTY"] + ex["FwdFill_QTY"] - ex["OrigYE_QTY"]).clip(lower=0)
    ex["Excess_$"] = ex["Excess_QTY"] * ex["AvgPrice"]
    ex = ex[ex["Excess_$"] > 0].sort_values("Excess_$", ascending=False)

    # QA #69: Add off-plan customer (Southeast/MLC/Other) rows — Plan=0, OverPlan=full ship
    op_net_pivot_copd = m.get("_offplan_net_pivot")
    if op_net_pivot_copd is not None and not op_net_pivot_copd.empty:
        op_records = []
        for ki in op_net_pivot_copd.index:
            for cust in op_net_pivot_copd.columns:
                v = op_net_pivot_copd.loc[ki, cust]
                if v and v > 0.5:
                    op_records.append({
                        "Customer": cust, "KI": ki, "OrigYE_QTY": 0,
                        "Excess_QTY": 0, "Excess_$": float(v),
                        "2023": 0, "2024": 0, "2025": 0,
                        "Smoothed_Hist": 0, "Outlier_Years": "",
                        "YTD_QTY": 0, "YTD_Pace_vs_Plan": float("nan"), "YTD_YoY": float("nan"),
                    })
        if op_records:
            op_df = pd.DataFrame(op_records)
            ex = pd.concat([ex, op_df], ignore_index=True, sort=False)
            ex = ex.sort_values("Excess_$", ascending=False).reset_index(drop=True)

    # Per (Customer, KI) avg sell prices
    ytd_p = panel.groupby(["Customer","KI"]).apply(
        lambda g: (g["AvgPrice_2026YTD"] * g["YTD_QTY"]).sum() / max(g["YTD_QTY"].sum(), 1) if g["YTD_QTY"].sum() > 0 else np.nan,
        include_groups=False,
    ).reset_index().rename(columns={0:"Price_2026"})
    h25_p = panel.groupby(["Customer","KI"]).apply(
        lambda g: (g["AvgPrice_2025"] * g["OrigYE_QTY"]).sum() / max(g["OrigYE_QTY"].sum(), 1) if g["OrigYE_QTY"].sum() > 0 else np.nan,
        include_groups=False,
    ).reset_index().rename(columns={0:"Price_2025"})
    ex = ex.merge(ytd_p, on=["Customer","KI"], how="left")
    ex = ex.merge(h25_p, on=["Customer","KI"], how="left")

    r = 1
    ws.cell(row=r, column=1, value="Customer Over-Plan Detail by Customer × Key Item Grouping").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Per-customer over-plan detail with multi-year history, YTD pace, pricing reference, lift construction explanation, and over-plan attribution flags.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19)
    r += 2

    headers = ["#", "Customer", "Key Item Grouping",
                "Plan QTY YE", "Over-Plan QTY YE", "Over-Plan $ YE",
                "2023 Sales QTY (full year)", "2024 Sales QTY (full year)", "2025 Sales QTY (full year)",
                "Smoothed Hist (3-yr, full year)", "Outlier Yrs",
                "YTD QTY", "YTD vs Plan %", "YTD YoY %",
                "Avg Sell Price 2026 YTD", "Avg Sell Price 2025", "Price YoY % (2026 YTD vs 2025)",
                "Why no further lift", "Flags"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 19, height=32)
    header_row = r
    r += 1

    # Build lookup: which (Customer, KI) have synthesized rows (Phase C synthesis)
    plan_lifted = d["plan_lifted"]
    if "Plan_Source" in plan_lifted.columns:
        synth_rows = plan_lifted[plan_lifted["Plan_Source"] == "Synthesized"]
        synth_cust_ki = set(zip(synth_rows["Customer"], synth_rows["KI"]))
    else:
        synth_cust_ki = set()

    # Build OrigJAD vs YTD (for "YTD over-shipped" flag) and OrigMD (for synthesis detection)
    cust_ki_orig_jad = ki_panel.groupby(["Customer","KI"])["OrigJAD_QTY"].sum().to_dict()
    cust_ki_orig_md = ki_panel.groupby(["Customer","KI"])["OrigMD_QTY"].sum().to_dict()
    cust_ki_orig_ye = ki_panel.groupby(["Customer","KI"])["OrigYE_QTY"].sum().to_dict()
    cust_ki_ytd = ki_panel.groupby(["Customer","KI"])["YTD_QTY"].sum().to_dict()
    cust_ki_fwd = ki_panel.groupby(["Customer","KI"])["FwdFill_QTY"].sum().to_dict()
    cust_ki_smooth = ki_panel.groupby(["Customer","KI"])["Smoothed_Hist"].sum().to_dict()

    def compute_why_and_flags(cust, ki, lift_q, smoothed_h, orig_md):
        """Compute Why-no-further-lift and Flags labels."""
        is_synth = (cust, ki) in synth_cust_ki
        orig_ye = cust_ki_orig_ye.get((cust, ki), 0)
        ytd = cust_ki_ytd.get((cust, ki), 0)
        orig_jad = cust_ki_orig_jad.get((cust, ki), 0)
        fwd = cust_ki_fwd.get((cust, ki), 0)

        # Why no further lift
        if is_synth and orig_md <= 0.5 and orig_ye <= 0.5:
            why = "Synthesized rows: no plan, forecast from history"
        elif is_synth and orig_md <= 0.5 and orig_ye > 0.5:
            why = "Synthesized rows: partial plan, gaps filled from history"
        elif is_synth:
            why = "Synthesized rows: gaps filled from history"
        elif lift_q > 0.5 and smoothed_h > orig_md * 1.05:
            why = "Lift partial — inventory exceeds lifted plan"
        elif smoothed_h <= orig_md * 1.05 and smoothed_h > 0:
            why = "History below plan — no upward lift"
        elif smoothed_h <= 0:
            why = "No history available"
        else:
            why = "Lift fully captured smoothed history"

        # Flags (multi-value, comma-separated)
        flags = []
        # Over-plan attribution
        if orig_jad > 0 and ytd > orig_jad * 1.02:  # YTD ran 2%+ over Q1 plan
            flags.append("Over-plan: YTD over-shipped vs plan")
        if is_synth and orig_md <= 0.5 and orig_ye <= 0.5:
            flags.append("Over-plan: forecast from history (no plan)")
        elif is_synth and orig_md <= 0.5 and orig_ye > 0.5:
            flags.append("Over-plan: forecast from history (partial plan)")
        if fwd > orig_md * 1.02 and orig_md > 0 and not is_synth:
            flags.append("Over-plan: forward over-shipped vs lifted plan")
        # Characteristics
        if smoothed_h > orig_ye * 1.2 and orig_ye > 0:
            flags.append("History exceeds plan >20%")
        return why, "; ".join(flags) if flags else ""

    ex = ex.reset_index(drop=True)
    for i in range(len(ex)):
        rr = ex.iloc[i]
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["KI"])
        ws.cell(row=r, column=4, value=float(rr["OrigYE_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["Excess_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr["Excess_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=6).fill = POS_FILL
        ws.cell(row=r, column=7, value=float(rr["2023"])).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(rr["2024"])).number_format = FMT_INT
        ws.cell(row=r, column=9, value=float(rr["2025"])).number_format = FMT_INT
        ws.cell(row=r, column=10, value=float(rr["Smoothed_Hist"])).number_format = FMT_INT
        ws.cell(row=r, column=11, value=str(rr["Outlier_Years"] or ""))
        ws.cell(row=r, column=12, value=float(rr["YTD_QTY"])).number_format = FMT_INT
        ytd_pace = rr["YTD_Pace_vs_Plan"]
        if pd.notna(ytd_pace):
            ws.cell(row=r, column=13, value=float(ytd_pace)).number_format = FMT_PCT
        yoy = rr["YTD_YoY"]
        if pd.notna(yoy):
            ws.cell(row=r, column=14, value=float(yoy)).number_format = FMT_PCT
        p_2026 = rr["Price_2026"] if pd.notna(rr["Price_2026"]) else None
        p_2025 = rr["Price_2025"] if pd.notna(rr["Price_2025"]) else None
        if p_2026:
            ws.cell(row=r, column=15, value=float(p_2026)).number_format = FMT_DLR2
        if p_2025:
            ws.cell(row=r, column=16, value=float(p_2025)).number_format = FMT_DLR2
        if p_2026 and p_2025 and p_2025 > 0:
            yoy_pct = (p_2026 - p_2025) / p_2025
            ws.cell(row=r, column=17, value=float(yoy_pct)).number_format = FMT_PCT
        # New: Why no further lift + Flags
        cust = rr["Customer"]; ki = rr["KI"]
        lift_q = float(rr.get("Lift_QTY", 0) or 0)
        smoothed_h = float(rr.get("Smoothed_Hist", 0) or 0)
        orig_md = cust_ki_orig_md.get((cust, ki), 0)
        why, flags = compute_why_and_flags(cust, ki, lift_q, smoothed_h, orig_md)
        ws.cell(row=r, column=18, value=why)
        ws.cell(row=r, column=19, value=flags)
        r += 1

    ws.freeze_panes = f"A{header_row+1}"
    try:
        from openpyxl.utils import get_column_letter
        end_row = r - 1
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{end_row}"
    except Exception:
        pass
    log("  Tab 7 written.")


def write_tab8_channel_summary(wb, m, d):
    log("Writing Tab 8: Channel Summary...")
    ws = wb.create_sheet("Channel Summary")
    set_col_widths(ws, [4, 35, 12, 12, 12, 12, 12, 12])

    ki_panel = m["ki_panel"]

    r = 1
    ws.cell(row=r, column=1, value="Channel Summary — Per-Customer Forward View").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 24
    r += 2

    for c in CUSTOMERS:
        ws.cell(row=r, column=1, value=f"{c}").font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        # Per-customer headline
        d_cust = ki_panel[ki_panel["Customer"]==c]
        if d_cust.empty:
            r += 1
            continue
        plan_d = d_cust["OrigYE_dlr"].sum()
        miss_d = d_cust["Miss_dlr_KI"].sum()
        achieved = plan_d - miss_d
        miss_pct = miss_d / plan_d if plan_d else 0
        ws.cell(row=r, column=1, value="    Plan $ YE").font = NORMAL_FONT
        ws.cell(row=r, column=2, value=float(plan_d)).number_format = FMT_DLR
        ws.cell(row=r, column=2).font = BOLD_FONT
        r += 1
        ws.cell(row=r, column=1, value="    Projected Achievement $ YE").font = NORMAL_FONT
        ws.cell(row=r, column=2, value=float(achieved)).number_format = FMT_DLR
        r += 1
        ws.cell(row=r, column=1, value="    Plan Miss $ YE").font = NORMAL_FONT
        ws.cell(row=r, column=2, value=float(miss_d)).number_format = FMT_DLR
        ws.cell(row=r, column=2).fill = NEG_FILL
        r += 1
        ws.cell(row=r, column=1, value="    Miss % YE").font = NORMAL_FONT
        ws.cell(row=r, column=2, value=float(miss_pct)).number_format = FMT_PCT
        r += 2

        # Top KIs for this customer (ranked by miss + excess)
        ws.cell(row=r, column=1, value=f"    Top KIs by Miss $ YE").font = SUBHEADER_FONT
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        headers = ["#", "Key Item Grouping", "Plan $ YE", "Miss $ YE", "Miss % YE", "YTD $", "YTD Pace %", ""]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=r, column=j, value=h)
        apply_header(ws, r, 8)
        r += 1
        d_cust["Excess_$"] = d_cust["Excess_QTY_vsPlan"] * d_cust["AvgPrice"]
        # Top 10 by miss
        top_miss = d_cust[d_cust["Miss_dlr_KI"]>0].sort_values("Miss_dlr_KI", ascending=False).head(10)
        for _miss_rank, (_, rr) in enumerate(top_miss.iterrows(), start=1):
            ws.cell(row=r, column=1, value=_miss_rank)
            ws.cell(row=r, column=2, value=rr["KI"])
            ws.cell(row=r, column=3, value=float(rr["OrigYE_dlr"])).number_format = FMT_DLR
            ws.cell(row=r, column=4, value=float(rr["Miss_dlr_KI"])).number_format = FMT_DLR
            ws.cell(row=r, column=4).fill = NEG_FILL
            ws.cell(row=r, column=5, value=float(rr["Miss_dlr_KI"]/rr["OrigYE_dlr"]) if rr["OrigYE_dlr"] else 0).number_format = FMT_PCT
            ws.cell(row=r, column=6, value=float(rr["YTD_dlr"])).number_format = FMT_DLR
            pace = rr["YTD_QTY"]/rr["OrigJAD_QTY"] if rr["OrigJAD_QTY"]>0 else 0
            ws.cell(row=r, column=7, value=float(pace)).number_format = FMT_PCT
            r += 1
        r += 1
        # Top 10 by excess for this customer
        ws.cell(row=r, column=1, value=f"    Top KIs by Over-Plan $ YE").font = SUBHEADER_FONT
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        headers = ["#", "Key Item Grouping", "Plan $ YE", "Over-Plan $ YE", "Over-Plan QTY YE", "Smoothed Hist (3-yr, full year)", "YTD QTY", "YTD Pace %"]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=r, column=j, value=h)
        apply_header(ws, r, 8)
        r += 1
        top_exc = d_cust[d_cust["Excess_$"]>0].sort_values("Excess_$", ascending=False).head(10)
        for _op_rank, (_, rr) in enumerate(top_exc.iterrows(), start=1):
            ws.cell(row=r, column=1, value=_op_rank)
            ws.cell(row=r, column=2, value=rr["KI"])
            ws.cell(row=r, column=3, value=float(rr["OrigYE_dlr"])).number_format = FMT_DLR
            ws.cell(row=r, column=4, value=float(rr["Excess_$"])).number_format = FMT_DLR
            ws.cell(row=r, column=4).fill = POS_FILL
            ws.cell(row=r, column=5, value=float(rr["Excess_QTY_vsPlan"])).number_format = FMT_INT
            ws.cell(row=r, column=6, value=float(rr["Smoothed_Hist"])).number_format = FMT_INT
            ws.cell(row=r, column=7, value=float(rr["YTD_QTY"])).number_format = FMT_INT
            pace = rr["YTD_QTY"]/rr["OrigJAD_QTY"] if rr["OrigJAD_QTY"]>0 else 0
            ws.cell(row=r, column=8, value=float(pace)).number_format = FMT_PCT
            r += 1
        r += 2


    # QA #69: Off-plan customer summary sections (Southeast/MLC/Other)
    op_net_pivot_ch = m.get("_offplan_net_pivot")
    OFFPLAN_CUSTS_CH = ["Southeast", "MLC", "Other"]
    if op_net_pivot_ch is not None and not op_net_pivot_ch.empty:
        for op_c in OFFPLAN_CUSTS_CH:
            if op_c not in op_net_pivot_ch.columns:
                continue
            cust_total = float(op_net_pivot_ch[op_c].sum())
            if cust_total <= 0.5:
                continue
            ws.cell(row=r, column=1, value=f"{op_c} (off-plan — no 2026 plan, history-driven)").font = SECTION_FONT
            ws.cell(row=r, column=1).fill = SECTION_FILL
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
            ws.cell(row=r, column=1, value="    Plan $ YE").font = NORMAL_FONT
            ws.cell(row=r, column=2, value=0.0).number_format = FMT_DLR
            r += 1
            ws.cell(row=r, column=1, value="    Projected Achievement $ YE").font = NORMAL_FONT
            ws.cell(row=r, column=2, value=cust_total).number_format = FMT_DLR
            ws.cell(row=r, column=2).fill = POS_FILL
            r += 1
            ws.cell(row=r, column=1, value="    Plan Miss $ YE").font = NORMAL_FONT
            ws.cell(row=r, column=2, value=0.0).number_format = FMT_DLR
            r += 1
            ws.cell(row=r, column=1, value="    Customer Over-Plan $ YE (= achievement, no plan baseline)").font = NORMAL_FONT
            ws.cell(row=r, column=2, value=cust_total).number_format = FMT_DLR
            ws.cell(row=r, column=2).fill = POS_FILL
            r += 2
            # Top KIs by ship $
            top_kis = op_net_pivot_ch[op_c].sort_values(ascending=False).head(10)
            ws.cell(row=r, column=1, value="    Top KIs by Over-Plan $ YE (off-plan customer fills)").font = BOLD_FONT
            r += 1
            ws.cell(row=r, column=1, value="#").font = HEADER_FONT
            ws.cell(row=r, column=2, value="Key Item Grouping").font = HEADER_FONT
            ws.cell(row=r, column=3, value="Over-Plan $ YE").font = HEADER_FONT
            r += 1
            rank = 1
            for ki, val in top_kis.items():
                if val <= 0.5: break
                ws.cell(row=r, column=1, value=rank)
                ws.cell(row=r, column=2, value=ki)
                ws.cell(row=r, column=3, value=float(val)).number_format = FMT_DLR
                r += 1
                rank += 1
            r += 1

    log("  Tab 8 written.")


def write_tab9_ytd(wb, m):
    log("Writing Tab 9: YTD Performance...")
    ws = wb.create_sheet("YTD Performance")
    set_col_widths(ws, [4, 14, 35, 12, 12, 12, 12, 11])

    ki_panel = m["ki_panel"]

    r = 1
    ws.cell(row=r, column=1, value="YTD Jan-Apr Performance — Customer × KI").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Plan vs Actual for Jan-Apr 2026 by Customer × KI. Sets context for YE projection — running ahead/behind plan?").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    headers = ["#", "Customer", "Key Item Grouping", "YTD Plan QTY", "YTD Actual QTY", "YTD Pace %",
                "YTD Plan $", "YTD Actual $"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 8, height=28)
    header_row = r
    r += 1

    ki_panel_ytd = ki_panel.copy()
    ki_panel_ytd["YTD_Plan_$"] = ki_panel_ytd["OrigJAD_QTY"] * ki_panel_ytd["AvgPrice"]
    ki_panel_ytd["YTD_Var_$"] = ki_panel_ytd["YTD_dlr"] - ki_panel_ytd["YTD_Plan_$"]
    ki_panel_ytd = ki_panel_ytd.sort_values("YTD_Var_$")  # most negative first

    ki_panel_ytd = ki_panel_ytd.reset_index(drop=True)
    for i in range(len(ki_panel_ytd)):
        rr = ki_panel_ytd.iloc[i]
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["KI"])
        ws.cell(row=r, column=4, value=float(rr["OrigJAD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["YTD_QTY"])).number_format = FMT_INT
        pace = rr["YTD_QTY"] / rr["OrigJAD_QTY"] if rr["OrigJAD_QTY"] > 0 else 0
        ws.cell(row=r, column=6, value=float(pace)).number_format = FMT_PCT
        if pace < 0.85:
            ws.cell(row=r, column=6).fill = NEG_FILL
        elif pace > 1.15:
            ws.cell(row=r, column=6).fill = POS_FILL
        ws.cell(row=r, column=7, value=float(rr["YTD_Plan_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=8, value=float(rr["YTD_dlr"])).number_format = FMT_DLR
        r += 1

    ws.freeze_panes = f"A{header_row+1}"
    try:
        from openpyxl.utils import get_column_letter
        end_row = r - 1
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{end_row}"
    except Exception:
        pass
    log("  Tab 9 written.")



def write_tab_miss_summary_by_ki(wb, m):
    """Tab: Miss Summary by KI — KI-only level, summed across customers."""
    log("Writing Tab: Miss Summary by KI...")
    ws = wb.create_sheet("Miss Summary by KI")
    set_col_widths(ws, [4, 35, 14, 12, 14, 12, 14, 14, 11, 12, 22])
    ki_total = m["ki_total"]

    r = 1
    ws.cell(row=r, column=1, value="Miss Summary by Key Item Grouping (one row per KI, summed across customers)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Miss = Plan - YTD - Forward Fulfillable (capped at 0). KI-level — same numbers as Plan by KI's Plan Miss column. For per-customer or per-item drill-down see other Miss tabs.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 30
    r += 2

    headers = ["#", "Key Item Grouping",
               "Plan QTY YE", "YTD QTY", "Fwd Fill QTY (May-Dec)",
               "Miss QTY YE", "Plan $ YE", "Miss $ YE", "YE Miss %",
               "Avg Price $", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 11, height=42)
    header_row = r
    r += 1
    data_start = r

    miss_kis = ki_total[ki_total["Miss_dlr_KI_total"] > 0].copy().sort_values("Miss_dlr_KI_total", ascending=False)
    rank = 1
    for _, rr in miss_kis.iterrows():
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["KI"])
        ws.cell(row=r, column=3, value=float(rr["OrigYE_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=4, value=float(rr["YTD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["FwdFill_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr["Miss_QTY_KI"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr["OrigYE_dlr"])).number_format = FMT_DLR
        ws.cell(row=r, column=8, value=float(rr["Miss_dlr_KI_total"])).number_format = FMT_DLR
        ws.cell(row=r, column=8).fill = NEG_FILL
        miss_pct_local = float(rr["Miss_dlr_KI_total"]) / float(rr["OrigYE_dlr"]) if float(rr["OrigYE_dlr"]) > 0 else 0.0
        ws.cell(row=r, column=9, value=miss_pct_local).number_format = FMT_PCT
        ws.cell(row=r, column=10, value=float(rr["AvgPrice"])).number_format = FMT_DLR2
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Miss Summary by KI written ({last_data_row - data_start + 1} KIs).")


def write_tab_miss_by_cust_ki(wb, m):
    """Tab: Miss by Customer x KI — one row per (Customer, KI)."""
    log("Writing Tab: Miss by Customer x KI...")
    ws = wb.create_sheet("Miss by Customer x KI")
    set_col_widths(ws, [4, 14, 35, 14, 12, 14, 12, 14, 14, 11, 12, 22])
    ki_panel = m["ki_panel"]

    r = 1
    ws.cell(row=r, column=1, value="Miss by Customer x Key Item Grouping (one row per Customer x KI)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Per (Customer, KI) miss vs Original Plan. Miss is capped per (Customer, KI) — accounts for intra-KI substitution within a customer.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 30
    r += 2

    headers = ["#", "Customer", "Key Item Grouping",
               "Plan QTY YE", "YTD QTY", "Fwd Fill QTY (May-Dec)",
               "Miss QTY YE", "Plan $ YE", "Miss $ YE", "YE Miss %",
               "Avg Price $", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 12, height=42)
    header_row = r
    r += 1
    data_start = r

    miss_rows = ki_panel[ki_panel["Miss_dlr_KI"] > 0].copy().sort_values("Miss_dlr_KI", ascending=False)
    rank = 1
    for _, rr in miss_rows.iterrows():
        miss_pct = float(rr["Miss_dlr_KI"]) / float(rr["OrigYE_dlr"]) if rr["OrigYE_dlr"] > 0 else 0
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["KI"])
        ws.cell(row=r, column=4, value=float(rr["OrigYE_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["YTD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr["FwdFill_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr["Miss_QTY_KI"])).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(rr["OrigYE_dlr"])).number_format = FMT_DLR
        ws.cell(row=r, column=9, value=float(rr["Miss_dlr_KI"])).number_format = FMT_DLR
        ws.cell(row=r, column=9).fill = NEG_FILL
        ws.cell(row=r, column=10, value=miss_pct).number_format = FMT_PCT
        ws.cell(row=r, column=11, value=float(rr["AvgPrice"])).number_format = FMT_DLR2
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Miss by Customer x KI written ({last_data_row - data_start + 1} rows).")


def write_tab_miss_by_ki_month(wb, m):
    """Tab: Miss by KI x Month — one row per (KI, Month). No customer dimension."""
    log("Writing Tab: Miss by KI x Month...")
    ws = wb.create_sheet("Miss by KI x Month")
    set_col_widths(ws, [4, 35, 8, 14, 14, 14, 12, 12, 11])

    ki_month = m.get("ki_month")
    if ki_month is None or ki_month.empty:
        log("  (no ki_month data)")
        return

    r = 1
    ws.cell(row=r, column=1, value="Miss by KI x Month (one row per KI x Month, summed across customers)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Per (KI, Month) miss vs Original Plan for that month. Forward months (May-Dec): Miss = Plan - FwdFill. Jan-Apr months not shown (covered by YTD Performance tab).").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 30
    r += 2

    headers = ["#", "Key Item Grouping", "Month",
               "Plan QTY (that month)", "Fwd Fill QTY (that month)", "Miss QTY (that month)",
               "Plan $ (that month)", "Miss $ (that month)", "Miss % (that month)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 9, height=42)
    header_row = r
    r += 1
    data_start = r

    # Filter to forward months only (Jan-Apr is in YTD Performance)
    fwd_only = ki_month[ki_month["MonthNum"].isin(FWD_MONTHS)].copy()
    # Compute Miss = max(0, OrigPlan - FwdFill)
    fwd_only["Miss_QTY"] = (fwd_only["OrigPlan_QTY"] - fwd_only["FwdFill_QTY"]).clip(lower=0)
    fwd_only["Miss_$"] = fwd_only["Miss_QTY"] * fwd_only["AvgPrice"]
    fwd_only["Miss_pct"] = np.where(fwd_only["OrigPlan_QTY"]>0,
                                     fwd_only["Miss_QTY"]/fwd_only["OrigPlan_QTY"], 0)
    # Sort by Miss $ desc
    fwd_only = fwd_only.sort_values("Miss_$", ascending=False)

    rank = 1
    for _, rr in fwd_only.iterrows():
        if rr["Miss_$"] <= 0:
            continue
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["KI"])
        ws.cell(row=r, column=3, value=MONTH_LBL.get(int(rr["MonthNum"]), int(rr["MonthNum"])))
        ws.cell(row=r, column=4, value=float(rr["OrigPlan_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["FwdFill_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr["Miss_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr["OrigPlan_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=8, value=float(rr["Miss_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=8).fill = NEG_FILL
        ws.cell(row=r, column=9, value=float(rr["Miss_pct"])).number_format = FMT_PCT
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Miss by KI x Month written ({last_data_row - data_start + 1} rows).")


def write_tab_lift_summary_by_ki_only(wb, m):
    """Tab: Lift Summary by KI — KI-only level, summed across customers."""
    log("Writing Tab: Lift Summary by KI (KI-only)...")
    ws = wb.create_sheet("Lift Summary by KI")
    set_col_widths(ws, [4, 35, 18, 18, 14, 12, 14, 22, 12, 22])

    ki_total = m["ki_total"]

    r = 1
    ws.cell(row=r, column=1, value="Lift Summary by Key Item Grouping (KI-only, summed across customers, May-Dec)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Lift = max(0, smoothed history - plan) at KI level for May-Dec. Per-customer detail in 'Lift by Customer x KI' tab.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = ["#", "Key Item Grouping",
               "Plan QTY May-Dec orig", "Lifted Plan QTY May-Dec",
               "Lift QTY May-Dec", "Lift % May-Dec", "Lift $ May-Dec",
               "Smoothed History QTY (3-yr avg, May-Dec)", "Avg Price $", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 10, height=42)
    header_row = r
    r += 1
    data_start = r

    lifts = ki_total[ki_total["Lift_QTY"] > 0].copy().sort_values("Lift_dlr", ascending=False)
    rank = 1
    for _, rr in lifts.iterrows():
        lift_pct = float(rr["Lift_QTY"]) / float(rr["OrigMD_QTY"]) if rr["OrigMD_QTY"] > 0 else 0
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["KI"])
        ws.cell(row=r, column=3, value=float(rr["OrigMD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=4, value=float(rr["LiftedMD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["Lift_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=lift_pct).number_format = FMT_PCT
        ws.cell(row=r, column=7, value=float(rr["Lift_dlr"])).number_format = FMT_DLR
        ws.cell(row=r, column=7).fill = POS_FILL
        ws.cell(row=r, column=8, value=float(rr["Smoothed_Hist"])).number_format = FMT_INT
        ws.cell(row=r, column=9, value=float(rr["AvgPrice"])).number_format = FMT_DLR2
        notes = f"Smoothed history > plan by {lift_pct:.0%}"
        ws.cell(row=r, column=10, value=notes)
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Lift Summary by KI written ({last_data_row - data_start + 1} KIs).")


def write_tab_lift_by_cust_ki(wb, m, d):
    """Tab: Lift by Customer x KI — one row per (Customer, KI). Was previously named 'Lift Summary by KI'."""
    log("Writing Tab: Lift by Customer x KI...")
    ws = wb.create_sheet("Lift by Customer x KI")
    set_col_widths(ws, [4, 14, 35, 18, 18, 14, 12, 14, 22, 30])

    ki_panel = m["ki_panel"]
    lift_ki = ki_panel[ki_panel["Lift_QTY"] > 0].copy().sort_values("Lift_dlr", ascending=False)

    r = 1
    ws.cell(row=r, column=1, value="Lift by Customer x Key Item Grouping (May-Dec)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="One row per (Customer, KI). Lift = max(0, smoothed history - plan) at the customer x KI level for May-Dec.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = ["#", "Customer", "Key Item Grouping",
               "Plan QTY May-Dec orig", "Lifted Plan QTY May-Dec",
               "Lift QTY May-Dec", "Lift % May-Dec", "Lift $ May-Dec",
               "Smoothed History QTY (3-yr avg, May-Dec)", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 10, height=42)
    header_row = r
    r += 1
    data_start = r

    rank = 1
    for _, rr in lift_ki.iterrows():
        lift_pct = float(rr["Lift_QTY"]) / float(rr["OrigMD_QTY"]) if rr["OrigMD_QTY"] > 0 else 0
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["KI"])
        ws.cell(row=r, column=4, value=float(rr["OrigMD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["LiftedMD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr["Lift_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=lift_pct).number_format = FMT_PCT
        ws.cell(row=r, column=8, value=float(rr["Lift_dlr"])).number_format = FMT_DLR
        ws.cell(row=r, column=8).fill = POS_FILL
        ws.cell(row=r, column=9, value=float(rr["Smoothed_Hist"])).number_format = FMT_INT
        outliers = rr.get("Outlier_Years", "") if "Outlier_Years" in ki_panel.columns else ""
        notes = f"Smoothed history > plan by {lift_pct:.0%}" + (f"; outliers dropped: {outliers}" if outliers else "")
        ws.cell(row=r, column=10, value=notes)
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Lift by Customer x KI written ({last_data_row - data_start + 1} rows).")


def write_tab_lift_by_ki_month(wb, m, d):
    """Tab: Lift by KI x Month — one row per (KI, Month). No customer dimension."""
    log("Writing Tab: Lift by KI x Month...")
    ws = wb.create_sheet("Lift by KI x Month")
    set_col_widths(ws, [4, 35, 8, 14, 18, 12, 12, 14, 22])

    ki_month = m.get("ki_month")
    if ki_month is None or ki_month.empty:
        log("  (no ki_month data)")
        return

    r = 1
    ws.cell(row=r, column=1, value="Lift by KI x Month (KI x Month, summed across customers, May-Dec)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Per (KI, Month) lift = LiftedPlan - OrigPlan for that month, summed across customers. Forward months only.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2

    headers = ["#", "Key Item Grouping", "Month",
               "Plan QTY (that month)", "Lifted Plan QTY (that month)",
               "Lift QTY (that month)", "Lift % (that month)", "Lift $ (that month)",
               "Smoothed History QTY (that month)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 9, height=42)
    header_row = r
    r += 1
    data_start = r

    fwd_only = ki_month[ki_month["MonthNum"].isin(FWD_MONTHS)].copy()
    fwd_only["Lift_pct"] = np.where(fwd_only["OrigPlan_QTY"]>0,
                                     fwd_only["Lift_QTY"]/fwd_only["OrigPlan_QTY"], 0)
    fwd_only = fwd_only.sort_values("Lift_$", ascending=False)

    # Smoothed history per KI x Month — derive from d["smoothed"] aggregated to KI level
    smoothed = d["smoothed"]
    sm_lookup = {}
    if smoothed is not None and not smoothed.empty:
        sm_ki = smoothed[smoothed["LiftGroup"].astype(str).str.startswith("KI:")].copy()
        sm_ki["KI_clean"] = sm_ki["LiftGroup"].astype(str).str.replace("KI:", "", regex=False)
        agg = sm_ki.groupby(["KI_clean","Month"])["smoothed"].sum().reset_index()
        sm_lookup = {(row["KI_clean"], int(row["Month"])): float(row["smoothed"]) for _, row in agg.iterrows()}

    rank = 1
    for _, rr in fwd_only.iterrows():
        if rr["Lift_$"] <= 0:
            continue
        sm_qty = sm_lookup.get((rr["KI"], int(rr["MonthNum"])), 0)
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["KI"])
        ws.cell(row=r, column=3, value=MONTH_LBL.get(int(rr["MonthNum"]), int(rr["MonthNum"])))
        ws.cell(row=r, column=4, value=float(rr["OrigPlan_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=5, value=float(rr["LiftedPlan_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=6, value=float(rr["Lift_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr["Lift_pct"])).number_format = FMT_PCT
        ws.cell(row=r, column=8, value=float(rr["Lift_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=8).fill = POS_FILL
        ws.cell(row=r, column=9, value=sm_qty).number_format = FMT_INT
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Lift by KI x Month written ({last_data_row - data_start + 1} rows).")


def write_tab_lift_by_cust_item(wb, m):
    """Tab: Lift by Customer x Item — Customer x Item level (no Month). Was 'Lift Detail by Month'."""
    log("Writing Tab: Lift by Customer x Item...")
    ws = wb.create_sheet("Lift by Customer x Item")
    set_col_widths(ws, [4, 14, 14, 35, 25, 18, 18, 14, 12, 14, 12, 22])

    lbci = m.get("lift_by_cust_item")
    if lbci is None or lbci.empty:
        log("  (no lift_by_cust_item data)")
        return

    r = 1
    ws.cell(row=r, column=1, value="Lift by Customer x Item (May-Dec, summed across months)").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="One row per (Customer, Item Num). Lift QTY/$/% computed from May-Dec total for the item-customer combo (months collapsed).").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 2

    headers = ["#", "Customer", "Item Num", "Item Description", "Key Item Grouping",
               "Plan QTY May-Dec orig", "Lifted Plan QTY May-Dec",
               "Lift QTY May-Dec", "Lift % May-Dec", "Lift $ May-Dec",
               "Avg Price $", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 12, height=42)
    header_row = r
    r += 1
    data_start = r

    # Filter to lift > 0 and sort
    lifts = lbci[lbci["Lift_QTY"] > 0].copy().sort_values("Lift_$", ascending=False)
    rank = 1
    for _, rr in lifts.iterrows():
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["Item Num"])
        ws.cell(row=r, column=4, value=str(rr.get("Item Desc", ""))[:40])
        ws.cell(row=r, column=5, value=rr["KI"])
        ws.cell(row=r, column=6, value=float(rr["OrigMD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr["LiftedMD_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(rr["Lift_QTY"])).number_format = FMT_INT
        ws.cell(row=r, column=9, value=float(rr["Lift_pct"])).number_format = FMT_PCT
        ws.cell(row=r, column=10, value=float(rr["Lift_$"])).number_format = FMT_DLR
        ws.cell(row=r, column=10).fill = POS_FILL
        ws.cell(row=r, column=11, value=float(rr.get("Price", 0) or 0)).number_format = FMT_DLR2
        notes = f"Smoothed history > plan by {rr['Lift_pct']:.0%}"
        ws.cell(row=r, column=12, value=notes)
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Lift by Customer x Item written ({last_data_row - data_start + 1} rows).")


def _write_ki_lifecycle(wb, m, in_dollars: bool):
    """Render the KI Lifecycle tab. in_dollars=True for $ tab; False for Units tab.
    
    Columns: # | KI | YTD Plan | YTD Actual | Start INV (R) | Start INV (NR) |
             Fwd Plan | Fwd Actual | Fwd Inflows | Ending INV
    """
    suffix = "$" if in_dollars else "Units"
    fmt = FMT_DLR if in_dollars else FMT_INT
    tab_name = f"KI Lifecycle ({suffix})"
    log(f"Writing Tab: {tab_name}...")
    ws = wb.create_sheet(tab_name)
    set_col_widths(ws, [4, 35, 12, 12, 14, 16, 12, 14, 14, 14, 14, 13, 13, 13])

    kl = m.get("ki_lifecycle")
    if kl is None or kl.empty:
        ws["A1"] = f"{tab_name} — no data"
        ws["A1"].font = TITLE_FONT
        return

    r = 1
    title = f"KI Lifecycle ({suffix}) — Full picture per Key Item Grouping (no customer dimension)"
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.row_dimensions[r].height = 24
    r += 1
    note = ("YTD = Jan-Apr 2026 (445 calendar). Forward = May-Dec 2026 (snap 4/27 = May 1). "
            "Forward Sales Projected = engine FORECAST (not actuals). "
            "Pool conservation: Start INV (R + NR) + Fwd Inflows (Pool + Recovery) = Forward Sales Projected + Ending INV.")
    ws.cell(row=r, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.row_dimensions[r].height = 36
    r += 2

    period_label = f"({suffix})"
    headers = ["#", "Key Item Grouping",
               f"YTD Sales Plan {period_label}",
               f"YTD Sales Actual {period_label}",
               f"Start INV - Restricted {period_label}",
               f"Start INV - Non-Restricted {period_label}",
               f"Forward Sales Plan {period_label}",
               f"Forward Sales Projected {period_label}",
               f"Forward Inflows - Pool {period_label}",
               f"Forward Inflows - Recovery {period_label}",
               f"Ending INV Available {period_label}",
               f"YE Sales Plan {period_label}",
               f"YE Projected {period_label}",
               f"YE Variance {period_label}"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 14, height=42)
    header_row = r
    r += 1
    data_start = r

    rank = 1
    for _, row in kl.iterrows():
        avg_price = float(row["AvgPrice"])
        mult = avg_price if in_dollars else 1.0
        # Customer-attributable cols (Method B in $ mode, units in unit mode)
        if in_dollars:
            # Method B (QA #64): use customer-aware $ pre-aggregated in ki_lifecycle
            ytd_plan    = float(row.get("YTDPlan_dlr_B", row["YTDPlan_QTY"] * avg_price))
            ytd_actual  = float(row.get("YTDActual_dlr_B", row["YTDActual_QTY"] * avg_price))
            fwd_plan    = float(row.get("FwdPlan_dlr_B", row["FwdPlan_QTY"] * avg_price))
            fwd_actual  = float(row.get("FwdActual_dlr_B", row["FwdActual_QTY"] * avg_price))
        else:
            ytd_plan    = float(row["YTDPlan_QTY"])
            ytd_actual  = float(row["YTDActual_QTY"])
            fwd_plan    = float(row["FwdPlan_QTY"])
            fwd_actual  = float(row["FwdActual_QTY"])
        # Supply-side cols (use KI AvgPrice — no customer dimension), except Recovery $ uses Method B
        start_r       = float(row["StartINV_Restricted"]) * mult
        start_nr      = float(row["StartINV_NonRestricted"]) * mult
        fwd_inflows_pool = float(row.get("FwdInflows_Pool", row.get("FwdInflows", 0))) * mult
        if in_dollars:
            fwd_inflows_rec = float(row.get("Inflow_Recovery_dlr_B", row.get("FwdInflows_Recovery", 0) * avg_price))
        else:
            fwd_inflows_rec = float(row.get("FwdInflows_Recovery", 0))
        ending_inv    = float(row["EndingINV"]) * mult
        # YE summary cols (sum of customer-attributable, so consistent with Method B)
        ye_plan       = ytd_plan + fwd_plan
        ye_projected  = ytd_actual + fwd_actual

        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=row["KI"])
        ws.cell(row=r, column=3, value=ytd_plan).number_format = fmt
        ws.cell(row=r, column=4, value=ytd_actual).number_format = fmt
        ws.cell(row=r, column=5, value=start_r).number_format = fmt
        ws.cell(row=r, column=6, value=start_nr).number_format = fmt
        ws.cell(row=r, column=7, value=fwd_plan).number_format = fmt
        ws.cell(row=r, column=8, value=fwd_actual).number_format = fmt
        ws.cell(row=r, column=9, value=fwd_inflows_pool).number_format = fmt
        ws.cell(row=r, column=10, value=fwd_inflows_rec).number_format = fmt
        ws.cell(row=r, column=11, value=ending_inv).number_format = fmt
        ws.cell(row=r, column=12, value=ye_plan).number_format = fmt
        ws.cell(row=r, column=13, value=ye_projected).number_format = fmt
        ye_variance = ye_projected - ye_plan
        var_cell = ws.cell(row=r, column=14, value=ye_variance)
        var_cell.number_format = fmt
        if ye_variance < -0.5: var_cell.fill = NEG_FILL
        elif ye_variance > 0.5: var_cell.fill = POS_FILL
        r += 1
        rank += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab {tab_name} written ({last_data_row - data_start + 1} KIs)")


def write_tab_ki_lifecycle_units(wb, m):
    _write_ki_lifecycle(wb, m, in_dollars=False)


def write_tab_ki_lifecycle_dollars(wb, m):
    _write_ki_lifecycle(wb, m, in_dollars=True)


def write_tab_production_recovery(wb, m):
    """Tab: Production Recovery — list of (Customer, KI, Month) miss recovery actions.
    
    Each row = produce target item to cover KI-level miss. Within-KI substitution
    means producing the target item satisfies all items in that KI for that month.
    """
    log("Writing Tab: Production Recovery...")
    ws = wb.create_sheet("Production Recovery")
    set_col_widths(ws, [4, 14, 30, 8, 14, 30, 14, 14, 12, 14, 14, 12, 8, 28, 30])

    pr = m.get("production_recovery")
    if pr is None or pr.empty:
        ws["A1"] = "Production Recovery — no recoverable miss this build"
        ws["A1"].font = TITLE_FONT
        return

    r = 1
    ws.cell(row=r, column=1, value="Production Recovery — Plan Miss Recoverable via Production").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Each row = Production action: produce 'Target Item' to cover the KI-level miss for that customer x month. Within-KI substitution covers other items in the same KI. Sort: by $ desc.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
    ws.row_dimensions[r].height = 32
    r += 2

    headers = ["#", "Customer", "Key Item Grouping", "Recovery Month",
               "Production Target Item Num", "Production Target Item Description",
               "Recoverable QTY", "Recoverable $",
               "Min Grow Days", "Days to Month End", "Latest Production Start", "Slack Days",
               "Org Code", "Grow Days Source", "Substitutes For"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 15, height=42)
    header_row = r
    r += 1
    data_start = r

    import datetime
    snap = M.SNAP_DATE.date()
    for i, row in pr.iterrows():
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=row["Customer"])
        ws.cell(row=r, column=3, value=row["KI"])
        ws.cell(row=r, column=4, value=MONTH_LBL.get(int(row["MonthNum"]), int(row["MonthNum"])))
        ws.cell(row=r, column=5, value=row["ProductionTargetItem"])
        ws.cell(row=r, column=6, value=str(row.get("ProductionTargetItemDesc",""))[:40])
        ws.cell(row=r, column=7, value=float(row["RecoverableQTY"])).number_format = FMT_INT
        ws.cell(row=r, column=8, value=float(row["RecoverableDlr"])).number_format = FMT_DLR
        ws.cell(row=r, column=8).fill = POS_FILL
        ws.cell(row=r, column=9, value=float(row["MinGrowDays"])).number_format = FMT_INT
        ws.cell(row=r, column=10, value=float(row["DaysToMonthEnd"])).number_format = FMT_INT
        # Latest production start: month_end - grow_days, formatted as date
        last_day_map = {5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
        month_end = datetime.date(2026, int(row["MonthNum"]), last_day_map.get(int(row["MonthNum"]), 31))
        latest_start = month_end - datetime.timedelta(days=int(row["MinGrowDays"]))
        ws.cell(row=r, column=11, value=latest_start.strftime("%Y-%m-%d"))
        ws.cell(row=r, column=12, value=float(row["SlackDays"])).number_format = FMT_INT
        if row["SlackDays"] >= 30:
            ws.cell(row=r, column=12).fill = POS_FILL
        elif row["SlackDays"] < 7:
            ws.cell(row=r, column=12).fill = NEG_FILL
        ws.cell(row=r, column=13, value=row.get("OrgCode",""))
        ws.cell(row=r, column=14, value=row.get("GrowDaysSource",""))
        ws.cell(row=r, column=15, value=row.get("SubstitutesFor",""))
        r += 1
    last_data_row = r - 1

    if last_data_row >= data_start:
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Production Recovery written ({last_data_row - data_start + 1} actions; total ${pr['RecoverableDlr'].sum():,.0f})")


def write_tab_excess_at_farm(wb, m):
    """Tab: Excess by KI — pool inventory unused at end of December.

    This is the TRUE "excess at the farm" view: how much of the inventory pool
    is left over after Stage 1 (plan defense) and Stage 2 (lift surplus) for
    each Key Item Grouping. Distinct from the Customer Over-Plan tabs, which
    show CUSTOMER-level over-shipping vs their plan.
    """
    log("Writing Tab: Excess by KI...")
    ws = wb.create_sheet("Excess by KI")
    farm = m.get("excess_at_farm")
    if farm is None or farm.empty:
        ws["A1"] = "Excess by KI (Pool Leftover at YE)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "(no data)"
        return

    ws["A1"] = "Excess by KI — Inventory Pool Leftover at End of December"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Per Key Item Grouping, the inventory pool unused after Stage 1 "
                "plan defense and Stage 2 lift fulfillment. Sorted by $ value.")
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 32
    ws.merge_cells("A2:F2")

    headers = ["#", "Key Item Grouping", "Excess at Farm QTY (end of Dec)", "Avg Price $",
               "Excess at Farm $ (end of Dec)", "% of Total Excess at Farm $",
               "Reason for Excess"]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    has_reason = "ExcessReason" in farm.columns
    total_dollars = float(farm["ExcessAtFarm_$"].sum())
    r = header_row + 1
    rank = 1
    for _, row in farm.iterrows():
        if row["ExcessAtFarm_QTY"] <= 0:
            continue
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=row["KI"])
        ws.cell(row=r, column=3, value=float(row["ExcessAtFarm_QTY"])).number_format = '#,##0'
        ws.cell(row=r, column=4, value=float(row["AvgPrice"])).number_format = '$#,##0.00'
        ws.cell(row=r, column=5, value=float(row["ExcessAtFarm_$"])).number_format = '$#,##0;($#,##0);-'
        pct = (float(row["ExcessAtFarm_$"]) / total_dollars) if total_dollars > 0 else 0
        ws.cell(row=r, column=6, value=pct).number_format = '0.0%'
        if has_reason:
            ws.cell(row=r, column=7, value=str(row.get("ExcessReason", "")))
        r += 1
        rank += 1

    last_data_row = r - 1
    if last_data_row >= header_row + 1:
        # Total row
        ws.cell(row=r, column=2, value="TOTAL").font = Font(bold=True)
        cell = ws.cell(row=r, column=3, value=float(farm["ExcessAtFarm_QTY"].sum()))
        cell.number_format = '#,##0'; cell.font = Font(bold=True)
        cell = ws.cell(row=r, column=5, value=total_dollars)
        cell.number_format = '$#,##0;($#,##0);-'; cell.font = Font(bold=True)

    set_col_widths(ws, [5, 50, 18, 14, 18, 18, 48])

    # Autofilter on data rows only (exclude TOTAL)
    if last_data_row >= header_row:
        last_col = "G"
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"

    ws.freeze_panes = f"A{header_row + 1}"
    log(f"  Tab Excess by KI written. Total Excess at Farm $: ${total_dollars:,.0f}")


def write_tab_pricing_comparison(wb, m, d):
    """Tab 10: Pricing Comparison — my cascade price vs Pricing Look up file price,
    per (Customer, Item Num), sorted by absolute dollar impact biggest to smallest.
    Cascade pricing is unchanged in the model — this is purely a reference tab."""
    log("Writing Tab 10: Pricing Comparison...")
    import nor_cal_forward as _M
    pricing_path = Path(_M.PATH_GROW_TIMES).parent / "Pricing look up.xlsx"
    try:
        pf = pd.read_excel(pricing_path)
    except Exception as exc:
        log(f"  Skipping Pricing Comparison tab ({pricing_path}): {exc}")
        ws = wb.create_sheet("Pricing Comparison")
        ws.cell(row=1, column=1, value="Pricing lookup unavailable — tab skipped.")
        return
    pf = pf[pf["Region"] == "NOR CAL"].copy()
    pf["Total Qty"] = pd.to_numeric(pf["Total Qty"], errors="coerce")
    pf["Tot REV"]   = pd.to_numeric(pf["Tot REV"], errors="coerce")
    pf = pf.rename(columns={"Planning Customer": "Customer"})
    # Aggregate to (Customer, Item Num): qty × rev across all months/orgs
    agg = pf.groupby(["Customer", "Item Num"]).agg(
        plan_qty=("Total Qty", "sum"),
        plan_rev=("Tot REV", "sum"),
    ).reset_index()
    agg["plan_price"] = np.where(agg["plan_qty"] > 0, agg["plan_rev"] / agg["plan_qty"], np.nan)
    agg = agg[agg["Customer"].isin(CUSTOMERS)]

    panel = m["panel"]
    item_meta = m["item_meta"]
    pricing = panel[["Customer", "Item Num", "Price"]].drop_duplicates(subset=["Customer","Item Num"]).rename(columns={"Price": "cascade_price"})
    plan_qty_panel = panel[["Customer", "Item Num", "OrigYE_QTY"]].drop_duplicates(subset=["Customer","Item Num"]).rename(columns={"OrigYE_QTY": "model_qty"})

    cmp = agg.merge(pricing, on=["Customer", "Item Num"], how="outer")
    cmp = cmp.merge(plan_qty_panel, on=["Customer","Item Num"], how="outer")
    cmp = cmp.merge(item_meta[["Item Num", "Item Desc", "KI"]], on="Item Num", how="left")
    # Use plan QTY from agg (= official) where present, else model
    cmp["qty_yr"] = cmp["plan_qty"].fillna(cmp["model_qty"]).fillna(0)
    cmp["cascade_price"] = cmp["cascade_price"].fillna(0)
    cmp["plan_price"] = cmp["plan_price"].fillna(0)
    cmp["delta_$"] = cmp["plan_price"] - cmp["cascade_price"]
    cmp["delta_pct"] = np.where(cmp["cascade_price"] > 0, cmp["delta_$"] / cmp["cascade_price"], np.nan)
    cmp["dollar_impact"] = cmp["qty_yr"] * cmp["delta_$"]
    cmp["abs_impact"] = cmp["dollar_impact"].abs()
    cmp["my_plan_dlr"] = cmp["qty_yr"] * cmp["cascade_price"]
    cmp["lookup_plan_dlr"] = cmp["qty_yr"] * cmp["plan_price"]
    cmp = cmp.sort_values("abs_impact", ascending=False).reset_index(drop=True)

    ws = wb.create_sheet("Pricing Comparison")
    set_col_widths(ws, [4, 12, 14, 35, 28, 12, 11, 11, 11, 10, 13, 13, 13])

    r = 1
    ws.cell(row=r, column=1, value="Pricing Comparison — My Cascade Price vs Pricing Look up File Price").font = TITLE_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.row_dimensions[r].height = 24
    r += 1
    ws.cell(row=r, column=1, value="Per (Customer × Item) comparison. Sorted by absolute dollar impact (Plan QTY × price delta) biggest to smallest. Cascade pricing is the model's source of truth; this tab is a reference for understanding where official plan prices differ.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    r += 2

    headers = ["#", "Customer", "Item Num", "Item Description", "Key Item Grouping",
                "Plan QTY YE", "Cascade Price (mine)", "Lookup File Price",
                "Price Delta $", "Price Delta %", "Dollar Impact $ YE",
                "My Plan $ (cascade)", "Plan $ (lookup file)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    apply_header(ws, r, 13, height=32)
    header_row = r
    r += 1
    first_data_row = r

    for i in range(len(cmp)):
        rr = cmp.iloc[i]
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=rr["Customer"])
        ws.cell(row=r, column=3, value=rr["Item Num"])
        ws.cell(row=r, column=4, value=str(rr.get("Item Desc",""))[:42] if pd.notna(rr.get("Item Desc")) else "")
        ws.cell(row=r, column=5, value=str(rr.get("KI","")) if pd.notna(rr.get("KI")) else "")
        ws.cell(row=r, column=6, value=float(rr["qty_yr"])).number_format = FMT_INT
        ws.cell(row=r, column=7, value=float(rr["cascade_price"])).number_format = FMT_DLR2
        ws.cell(row=r, column=8, value=float(rr["plan_price"])).number_format = FMT_DLR2
        ws.cell(row=r, column=9, value=float(rr["delta_$"])).number_format = FMT_DLR2
        if rr["delta_$"] > 0:
            ws.cell(row=r, column=9).fill = POS_FILL
        elif rr["delta_$"] < 0:
            ws.cell(row=r, column=9).fill = NEG_FILL
        if pd.notna(rr["delta_pct"]):
            ws.cell(row=r, column=10, value=float(rr["delta_pct"])).number_format = FMT_PCT
        ws.cell(row=r, column=11, value=float(rr["dollar_impact"])).number_format = FMT_DLR
        if rr["dollar_impact"] > 0:
            ws.cell(row=r, column=11).fill = POS_FILL
        elif rr["dollar_impact"] < 0:
            ws.cell(row=r, column=11).fill = NEG_FILL
        ws.cell(row=r, column=12, value=float(rr["my_plan_dlr"])).number_format = FMT_DLR
        ws.cell(row=r, column=13, value=float(rr["lookup_plan_dlr"])).number_format = FMT_DLR
        r += 1
    last_data_row = r - 1

    # TOTAL row
    ws.cell(row=r, column=2, value="TOTAL").font = BOLD_FONT
    ws.cell(row=r, column=2).fill = TOTAL_FILL
    ws.cell(row=r, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})").number_format = FMT_INT
    ws.cell(row=r, column=11, value=f"=SUM(K{first_data_row}:K{last_data_row})").number_format = FMT_DLR
    ws.cell(row=r, column=12, value=f"=SUM(L{first_data_row}:L{last_data_row})").number_format = FMT_DLR
    ws.cell(row=r, column=13, value=f"=SUM(M{first_data_row}:M{last_data_row})").number_format = FMT_DLR
    for col_idx in [6, 11, 12, 13]:
        ws.cell(row=r, column=col_idx).font = BOLD_FONT
        ws.cell(row=r, column=col_idx).fill = TOTAL_FILL

    # Apply autofilter (per the autofilter rule)
    last_col = "M"  # 13 columns
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"

    ws.freeze_panes = f"A{first_data_row}"
    log("  Tab 10 (Pricing Comparison) written.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# SUBTOTAL row helper config — for percentage columns, instead of summing the
# percentages (which is meaningless), compute the ratio of two subtotaled
# numeric columns from the SAME row. Map: tab_name -> { %_col_letter: (num_col, denom_col) }
# All references are to the column letters AFTER the row insert (data shifts down by 1).
PCT_COL_RATIOS = {
    "Plan by KI":                {"G": ("F", "C")},
    "Customer Miss Detail":      {"L": ("K", "J")},   # was Short Item Detail
    "Miss Summary by KI":        {"I": ("H", "G")},   # YE Miss % = Miss $ / Plan $
    "Miss by Customer x KI":     {"J": ("I", "H")},   # YE Miss % = Miss $ / Plan $
    "Miss by KI x Month":        {"I": ("H", "G")},   # Miss % that month
    "Lift Summary by KI":        {"F": ("E", "C")},   # Lift % May-Dec (KI-only tab)
    "Lift by Customer x KI":     {"G": ("F", "D")},   # Lift % May-Dec
    "Lift by KI x Month":        {"G": ("F", "D")},   # Lift % that month
    "Lift by Customer x Item":   {"I": ("H", "F")},   # Lift % May-Dec
    "Excess by KI":              {"F": None},
    "Over-Plan by KI":           {},
    "Customer Over-Plan Detail": {"M": ("L", "D"), "N": None, "Q": None},
    "YTD Performance":           {"F": ("E", "D")},
    "Pricing Comparison":        {"J": None},
}

# Per-unit Price columns where SUBTOTAL(9) gives meaningless sum-of-prices.
# Provide a (numerator, denominator) ratio for weighted avg, or None to leave blank.
AVG_COL_RATIOS = {
    "Plan by KI":                {"H": None},
    "Miss Summary by KI":        {"J": ("G", "C")},   # Plan $ / Plan QTY
    "Miss by Customer x KI":     {"K": ("H", "D")},
    "Customer Miss Detail":      {"M": ("J", "F")},
    "Lift Summary by KI":        {"I": None},
    "Lift by Customer x Item":   {"K": None},
    "Excess by KI":              {"D": None},
    "Over-Plan by KI":           {"J": None},
    "Customer Over-Plan Detail": {"O": None, "P": None},
    "Pricing Comparison":        {"G": None, "H": None, "I": None},
}



def write_tab_build_health(wb, m, d):
    """Tab 2: Build Health — full dashboard surface for verification, methodology stats,
    inventory snapshot, data quality issues, and comparison to last build.
    Auto-populated by build script. Grey color. Stays visible to surface any failures.
    """
    import datetime, pickle, os
    log("Writing Tab: Build Health...")
    ws = wb.create_sheet("Build Health")
    set_col_widths(ws, [44, 28, 60])

    # Helpers
    GREEN = "70AD47"; RED = "C00000"; AMBER = "BF8F00"; NAVY = "1F4E78"; GREY_FILL_HEX = "EDEDED"
    SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
    LABEL_FONT = Font(bold=True, size=10)
    BODY_FONT = Font(size=10)
    PASS_FONT = Font(bold=True, color=GREEN, size=10)
    FAIL_FONT = Font(bold=True, color=RED, size=10)
    WARN_FONT = Font(bold=True, color=AMBER, size=10)
    TITLE_FONT_BH = Font(bold=True, size=14, color=NAVY)

    r = 1
    ws.cell(row=r, column=1, value="Build Health Dashboard").font = TITLE_FONT_BH
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Auto-populated each build. Surfaces verification status, methodology stats, inventory snapshot, data-quality issues, and changes vs prior build.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 28
    r += 2

    def write_section(title):
        nonlocal r
        c = ws.cell(row=r, column=1, value=title)
        c.font = SECTION_FONT
        c.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 22
        r += 1

    def write_kv(label, value, status_font=None, note=""):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = status_font if status_font else BODY_FONT
        if note:
            ws.cell(row=r, column=3, value=note).font = NOTE_FONT
        r += 1

    # ===== SECTION 1: VERIFICATION =====
    write_section("Section 1 — Verification")
    write_kv("Build timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    ki_panel = m.get("ki_panel")
    walk_check = ki_panel.groupby("Customer").agg(
        OrigYE=("OrigYE_QTY", "sum"),
        OrigJAD=("OrigJAD_QTY", "sum"),
        OrigMD=("OrigMD_QTY", "sum"),
        YTD=("YTD_QTY", "sum"),
        FwdFill=("FwdFill_QTY", "sum"),
    ).reset_index()
    walk_check["YE_Net"] = (walk_check["YTD"] + walk_check["FwdFill"]) - walk_check["OrigYE"]
    walk_check["YTD_Net"] = walk_check["YTD"] - walk_check["OrigJAD"]
    walk_check["Forward_Net"] = walk_check["FwdFill"] - walk_check["OrigMD"]
    walk_check["Diff"] = walk_check["YE_Net"] - (walk_check["YTD_Net"] + walk_check["Forward_Net"])
    walks_max = float(walk_check["Diff"].abs().max()) if not walk_check.empty else 0.0
    walks_pass = walks_max < 1.0
    write_kv(
        "Walks reconciliation: max |YE Net - (YTD Net + Forward Net)|",
        f"{walks_max:.2f} units",
        PASS_FONT if walks_pass else FAIL_FONT,
        "PASS - reconciles within 1 unit per Customer x KI" if walks_pass else "FAIL - investigate",
    )

    # Pool conservation: re-aggregate from ki_lifecycle if available
    kl = m.get("ki_lifecycle")
    if kl is not None and not kl.empty:
        kl_check = kl.copy()
        kl_check["LHS"] = kl_check["StartINV_Restricted"] + kl_check["StartINV_NonRestricted"] + kl_check.get("FwdInflows_Pool", kl_check.get("FwdInflows", 0)) + kl_check.get("FwdInflows_Recovery", 0)
        kl_check["RHS"] = kl_check["FwdActual_QTY"] + kl_check["EndingINV"]
        kl_check["Diff"] = kl_check["LHS"] - kl_check["RHS"]
        failing = kl_check[kl_check["Diff"].abs() > 0.5]
        n_failing = len(failing)
        total_gap_qty = float(failing["Diff"].abs().sum())
        avg_price_overall = float(kl_check["AvgPrice"].mean()) if not kl_check.empty else 0
        total_gap_dlr = total_gap_qty * avg_price_overall
        cons_pass = n_failing == 0
        write_kv(
            "Pool conservation: # KIs failing",
            f"{n_failing}",
            PASS_FONT if cons_pass else WARN_FONT,
            "PASS - all KIs reconcile" if cons_pass else f"WARN - phantom inflow / Source 3 still active",
        )
        write_kv(
            "Pool conservation: total gap units",
            f"{total_gap_qty:,.0f} units",
            BODY_FONT,
            f"~${total_gap_dlr:,.0f} at avg price",
        )
        # Top 10 by absolute diff
        if n_failing > 0:
            ws.cell(row=r, column=1, value="Top KIs by conservation gap (Source 3 / phantom inflow):").font = LABEL_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1
            top10 = failing.assign(AbsDiff=failing["Diff"].abs()).nlargest(10, "AbsDiff")
            for _, row in top10.iterrows():
                ws.cell(row=r, column=1, value=f"   {row['KI']}").font = BODY_FONT
                ws.cell(row=r, column=2, value=f"{row['Diff']:,.0f} units").font = BODY_FONT
                r += 1
    r += 1

    # ===== SECTION 2: METHODOLOGY STATS =====
    write_section("Section 2 — Methodology Stats")
    pr = m.get("production_recovery")
    if pr is not None and not pr.empty:
        write_kv("Production Recovery actions (Customer x KI x Month)", f"{len(pr):,}")
        write_kv("Production Recovery total $ recoverable", f"${float(pr['RecoverableDlr'].sum()):,.0f}")
        write_kv("Production Recovery total units recoverable", f"{float(pr['RecoverableQTY'].sum()):,.0f}")
    else:
        write_kv("Production Recovery actions", "0", BODY_FONT, "no recoverable miss this build")

    # Synthesis stats (post-Task 52 only)
    synth_count = m.get("synthesis_row_count", 0)
    if synth_count > 0:
        write_kv("Synthesized plan rows (Option B)", f"{synth_count:,}")
        write_kv("Synthesized plan $ total", f"${m.get('synthesis_dollars', 0):,.0f}")
    else:
        write_kv("Synthesized plan rows (Option B)", "(not yet implemented - Phase C)")

    # Stranded KIs (truly orphan: in pools but not in pool walk)
    stranded_count = m.get("stranded_ki_count", 0)
    stranded_qty = m.get("stranded_ki_units", 0)
    stranded_dlr = m.get("stranded_ki_dollars", 0)
    write_kv("Stranded KIs (in pools but not pool walk)", f"{stranded_count}", note="Truly orphan — see Read Me for definition")
    write_kv("Stranded inventory units", f"{stranded_qty:,.0f}")
    write_kv("Stranded inventory $", f"${stranded_dlr:,.0f}")

    # Broader: KIs with zero forward fulfillment (includes stranded + plan-but-no-fill)
    zero_fwd_count = 0
    if kl is not None and not kl.empty:
        zero_fwd_mask = (kl["FwdPlan_QTY"].fillna(0) == 0) & (kl["FwdActual_QTY"].fillna(0) == 0)
        zero_fwd_count = int(zero_fwd_mask.sum())
    write_kv("KIs with zero forward fulfillment", f"{zero_fwd_count}", note="Includes stranded + KIs whose plan/forecast resolved to zero forward — see Read Me")
    r += 1

    # ===== SECTION 3: INVENTORY SNAPSHOT =====
    write_section("Section 3 — Inventory Snapshot ($)")
    if kl is not None and not kl.empty:
        avg_price = kl["AvgPrice"]
        total_start_dlr = float((kl["StartINV_Restricted"] + kl["StartINV_NonRestricted"]) @ avg_price)
        total_pool_inflow_dlr = float(kl.get("FwdInflows_Pool", kl.get("FwdInflows", 0)) @ avg_price)
        total_rec_inflow_dlr = float(kl.get("FwdInflows_Recovery", 0) @ avg_price)
        total_fwd_plan_dlr = float(kl["FwdPlan_QTY"] @ avg_price)
        total_fwd_actual_dlr = float(kl["FwdActual_QTY"] @ avg_price)
        total_ending_dlr = float(kl["EndingINV"] @ avg_price)
        ki_total = m.get("ki_total")
        total_lifted_plan = float(ki_total["LiftedYE_QTY"].fillna(0) @ ki_total["AvgPrice"].fillna(0)) if "LiftedYE_QTY" in ki_total.columns else 0
        write_kv("Total Start INV ($)", f"${total_start_dlr:,.0f}")
        write_kv("Total Forward Inflows - Pool ($)", f"${total_pool_inflow_dlr:,.0f}")
        write_kv("Total Forward Inflows - Recovery ($)", f"${total_rec_inflow_dlr:,.0f}")
        write_kv("Total Forward Plan ($)", f"${total_fwd_plan_dlr:,.0f}")
        write_kv("Total Lifted YE Plan ($)", f"${total_lifted_plan:,.0f}")
        write_kv("Total Forward Forecast / FwdFill ($)", f"${total_fwd_actual_dlr:,.0f}")
        write_kv("Total Ending INV ($)", f"${total_ending_dlr:,.0f}")
    r += 1

    # ===== SECTION 4: DATA QUALITY ISSUES =====
    write_section("Section 4 — Data Quality Issues")
    issues = []

    plan = d["plan"]
    items = d["items"]
    prices = d.get("prices", {})
    pools = d["pools"]
    hist = d["hist"]

    plan_items = set(plan["Item Num"].dropna().unique())
    # Fix #62: prices is a DataFrame; DataFrame.keys() returns column names not item numbers.
    if hasattr(prices, "columns") and "Item Num" in list(getattr(prices, "columns", [])):
        priced_items = set(prices["Item Num"].dropna().unique())
    else:
        priced_items = set(prices.keys()) if hasattr(prices, "keys") else set()
    no_price = plan_items - priced_items
    if no_price:
        issues.append(("Items in plan with no pricing", str(len(no_price)), f"sample: {list(no_price)[:3]}"))

    # Items in pool with no plan and no history (true orphans)
    item_to_ki = items[["Item Num", "KI"]].drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict()
    pool_items = set(pools["Item Num"].dropna().unique())
    plan_kis = set(plan["KI"].dropna().unique())
    hist_with_ki = hist.copy()
    hist_with_ki["KI"] = hist_with_ki["Item Num"].map(item_to_ki)
    hist_kis = set(hist_with_ki.dropna(subset=["KI"])["KI"].unique())
    orphan_items = []
    for it in pool_items:
        ki = item_to_ki.get(it)
        if ki and ki not in plan_kis and ki not in hist_kis:
            orphan_items.append(it)
    if orphan_items:
        issues.append(("Items in pool with no plan AND no history (orphans)", str(len(orphan_items)), f"sample: {orphan_items[:3]}"))

    # YTD items not in item-master
    ytd = d["ytd"]
    ytd_no_master = set(ytd["Item Num"].dropna().unique()) - set(items["Item Num"].dropna().unique())
    if ytd_no_master:
        issues.append(("YTD items not in item-master", str(len(ytd_no_master)), f"sample: {list(ytd_no_master)[:3]}"))

    # Plan items not in item-master
    plan_no_master = plan_items - set(items["Item Num"].dropna().unique())
    if plan_no_master:
        issues.append(("Plan items not in item-master", str(len(plan_no_master)), f"sample: {list(plan_no_master)[:3]}"))

    # YTD items without KI mapping (data quality / coverage gap, QA #68)
    if d.get("ytd") is not None and not d["ytd"].empty:
        items_for_yidx = d.get("items")
        if items_for_yidx is not None:
            item_ki_map_yidx = items_for_yidx.drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict()
            ytd_data = d["ytd"].copy()
            ytd_data["KI"] = ytd_data["Item Num"].map(item_ki_map_yidx)
            no_ki_ytd = ytd_data[ytd_data["KI"].isna()]
            if not no_ki_ytd.empty:
                qty_total = float(no_ki_ytd["Qty"].sum())
                rev_col = "Revenue" if "Revenue" in no_ki_ytd.columns else None
                rev_total = float(no_ki_ytd[rev_col].sum()) if rev_col else 0
                issues.append(("YTD items with no KI mapping (excluded from per-KI tabs)", str(len(no_ki_ytd)),
                              f"{qty_total:,.0f} units / ${rev_total:,.0f} — see Read Me Section 15 'Coverage scope' for explanation"))

    # Negative YTD Actual values (credits/returns in source — real, but worth flagging)
    if kl is not None and not kl.empty:
        neg_ytd = kl[kl["YTDActual_QTY"] < -0.01]
        if not neg_ytd.empty:
            sample = ", ".join(f"{row['KI']}({row['YTDActual_QTY']:.0f})" for _, row in neg_ytd.head(3).iterrows())
            issues.append(("KIs with negative YTD Actual (credits/returns)", str(len(neg_ytd)), f"sample: {sample}"))

    if not issues:
        write_kv("Data quality issues detected", "None", PASS_FONT, "all checks pass")
    else:
        for label, value, note in issues:
            write_kv(label, value, WARN_FONT, note)
    r += 1

    # ===== SECTION 5: COMPARISON TO PRIOR BUILD =====
    write_section("Section 5 — Comparison to Prior Build")
    PRIOR_PATH = "/sessions/optimistic-beautiful-ramanujan/tmp/prior_build_metrics.pkl"
    current_metrics = {
        "plan_miss_dlr": float(m["ki_panel"]["Miss_dlr_KI"].sum()) if "Miss_dlr_KI" in m["ki_panel"].columns else 0,
        "fwd_forecast_dlr": float(total_fwd_actual_dlr) if kl is not None and not kl.empty else 0,
        "excess_at_farm_dlr": float(m["excess_at_farm"]["ExcessAtFarm_$"].sum()) if "excess_at_farm" in m and not m["excess_at_farm"].empty else 0,
        "synth_count": synth_count,
        "cons_failing": n_failing if kl is not None and not kl.empty else 0,
    }
    if os.path.exists(PRIOR_PATH):
        try:
            with open(PRIOR_PATH, "rb") as f:
                prior = pickle.load(f)
            for label, key, fmt in [
                ("Plan Miss $ delta vs last build", "plan_miss_dlr", "${:+,.0f}"),
                ("Forward Forecast $ delta vs last build", "fwd_forecast_dlr", "${:+,.0f}"),
                ("Excess at Farm $ delta vs last build", "excess_at_farm_dlr", "${:+,.0f}"),
                ("Synthesized rows delta", "synth_count", "{:+,}"),
                ("Pool conservation # failing KIs delta", "cons_failing", "{:+,}"),
            ]:
                cur = current_metrics.get(key, 0)
                prv = prior.get(key, 0)
                delta = cur - prv
                pct = (delta / prv * 100) if prv else 0
                write_kv(label, fmt.format(delta) + (f" ({pct:+.1f}%)" if prv else ""))
        except Exception as e:
            write_kv("Prior build metrics", "Could not load", BODY_FONT, str(e))
    else:
        write_kv("Prior build metrics", "(no prior build yet — first run)")

    # Persist current metrics for next build comparison
    try:
        os.makedirs(os.path.dirname(PRIOR_PATH), exist_ok=True)
        with open(PRIOR_PATH, "wb") as f:
            pickle.dump(current_metrics, f)
    except Exception as e:
        log(f"  WARN: could not persist prior_build_metrics: {e}")

    log("  Tab Build Health written.")


# ---------------------------------------------------------------------------
# Tab color-coding (per CEO 2026-05-07): same-domain tabs share a color.
# See memory: project_tab_color_scheme.md
TAB_COLORS = {
    # Reference / Utility (grey)
    "Read Me & Methodology":        "BFBFBF",
    "Build Health":                 "BFBFBF",
    "Changes":                      "BFBFBF",
    # Top-level / Headline tabs — each unique to differentiate (CEO 2026-05-08)
    "Exec Summary":                 "1F4E78",  # Navy — highest-level summary
    "Plan by KI":                   "117A65",  # Teal — operational plan view
    "KI Lifecycle (Units)":         "595959",  # Slate — lifecycle pair (Units)
    "KI Lifecycle ($)":             "595959",  # Slate — lifecycle pair ($)
    # Miss (red)
    "Miss Summary by KI":           "C00000",
    "Miss by Customer x KI":        "C00000",
    "Miss by KI x Month":           "C00000",
    "Customer Miss Detail":         "C00000",
    # Lift (green)
    "Lift Summary by KI":           "548235",
    "Lift by Customer x KI":        "548235",
    "Lift by KI x Month":           "548235",
    "Lift by Customer x Item":      "548235",
    # Production Recovery (orange — action-oriented, distinct from Miss/Lift)
    "Production Recovery":          "ED7D31",
    # Excess at Farm (gold)
    "Excess by KI":                 "BF8F00",
    # Customer Over-Plan (blue)
    "Over-Plan by KI":              "2E75B6",
    "Customer Over-Plan Detail":    "2E75B6",
    # Cross-cut (purple)
    "Channel Summary":              "7030A0",
    "YTD Performance":              "7030A0",
    "Pricing Comparison":           "7030A0",
}


def apply_tab_colors_and_verify(wb):
    """Set tab colors per TAB_COLORS dict; warn on any uncolored tab.

    Verification: every tab on the workbook must have an entry in TAB_COLORS.
    Tabs in the same domain group must share a color (color-by-name lookup
    enforces this — no risk of drift).
    """
    log("Applying tab colors...")
    uncolored = []
    color_groups = {}
    for tab_name in wb.sheetnames:
        color = TAB_COLORS.get(tab_name)
        if color:
            wb[tab_name].sheet_properties.tabColor = color
            color_groups.setdefault(color, []).append(tab_name)
        else:
            uncolored.append(tab_name)
    if uncolored:
        log(f"  WARN: tabs without TAB_COLORS entry: {uncolored}")
    log(f"  applied {len(color_groups)} distinct colors across {sum(len(v) for v in color_groups.values())} tabs")


def add_subtotal_rows_to_workbook(wb):
    """Post-processing pass: for every tab with an autofilter, insert a SUBTOTAL row
    just above the header (within filter visibility). Uses SUBTOTAL(9, range) which
    dynamically sums only visible (filtered) rows.

    For percentage columns: instead of summing percentages (meaningless), compute
    the ratio of two other subtotaled cells via PCT_COL_RATIOS lookup.

    For formula columns (cells containing =A+B style formulas): SUBTOTAL is still
    written so the column gets summed. After insert_rows shifts cells down, all
    formula references in the shifted range are translated forward by one row
    using openpyxl's Translator (otherwise references become stale and point to
    the header row instead of data).

    Implementation detail: ranges are sized to the actual last_data_row at the time
    of insertion, so the subtotal captures every populated row regardless of build size.
    """
    import re
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formula.translate import Translator

    SUBTOTAL_FILL = PatternFill("solid", start_color="FFE699")
    SUBTOTAL_FONT = Font(bold=True, color="000000")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not ws.auto_filter.ref:
            continue
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ws.auto_filter.ref)
        if not m:
            continue
        start_col_letter = m.group(1)
        header_row = int(m.group(2))
        end_col_letter = m.group(3)
        last_data_row = int(m.group(4))
        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        # Insert a blank row at the header position; shifts header + data down by 1
        ws.insert_rows(header_row)
        new_header_row = header_row + 1
        new_last_data_row = last_data_row + 1
        subtotal_row = header_row  # the freshly inserted row

        # ---- Translate every formula in the shifted region (header_row+1 onwards) ----
        # insert_rows moves cell objects but NOT their formula text — so we must re-anchor.
        # Walk every cell from new_header_row through ws.max_row; for any cell containing a
        # formula, translate it from its old coordinate (one row up) to the new coordinate.
        for row_idx in range(new_header_row, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    old_coord = f"{get_column_letter(col_idx)}{row_idx - 1}"
                    new_coord = cell.coordinate
                    try:
                        cell.value = Translator(cell.value, origin=old_coord).translate_formula(new_coord)
                    except Exception:
                        # If a formula can't be translated (rare — e.g., array formulas), skip silently
                        pass

        # ---- Detect summable columns: numeric values OR arithmetic formulas ----
        sum_cols = []
        for col in range(start_col, end_col + 1):
            v = ws.cell(row=new_header_row + 1, column=col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                sum_cols.append(col)
            elif isinstance(v, str) and v.startswith("=") and any(op in v for op in "+-*/"):
                sum_cols.append(col)

        # Label cell: leftmost text column
        label_col = start_col + 1 if (start_col + 1) <= end_col else start_col
        ws.cell(row=subtotal_row, column=label_col, value="VISIBLE TOTAL").font = SUBTOTAL_FONT
        ws.cell(row=subtotal_row, column=label_col).fill = SUBTOTAL_FILL

        # Lookup table for % columns -> ratio formulas (same as before)
        pct_map = PCT_COL_RATIOS.get(sheet_name, {})

        def _is_pct_col(col_letter):
            hdr = ws.cell(row=new_header_row, column=column_index_from_string(col_letter)).value
            return isinstance(hdr, str) and "%" in hdr

        avg_map = AVG_COL_RATIOS.get(sheet_name, {})

        def _is_avg_price_col(col_letter):
            hdr = ws.cell(row=new_header_row, column=column_index_from_string(col_letter)).value
            if not isinstance(hdr, str):
                return False
            low = hdr.lower()
            # "Price" in header but skip "Price Delta $" (it IS a delta dollar, summable)
            return ("price" in low) and ("delta" not in low or "%" in low)

        # Write SUBTOTAL formulas (or ratios for % columns; weighted avgs for price cols) ----
        for col in sum_cols:
            if col == label_col:
                continue
            col_l = get_column_letter(col)
            data_first = new_header_row + 1
            data_last = new_last_data_row

            if _is_pct_col(col_l):
                spec = pct_map.get(col_l)
                if spec is None:
                    continue  # leave blank
                num_col, denom_col = spec
                formula = f"=IFERROR({num_col}{subtotal_row}/{denom_col}{subtotal_row}, 0)"
                cell = ws.cell(row=subtotal_row, column=col, value=formula)
            elif _is_avg_price_col(col_l):
                spec = avg_map.get(col_l)
                if spec is None:
                    continue  # leave blank
                num_col, denom_col = spec
                formula = f"=IFERROR({num_col}{subtotal_row}/{denom_col}{subtotal_row}, 0)"
                cell = ws.cell(row=subtotal_row, column=col, value=formula)
            else:
                cell = ws.cell(row=subtotal_row, column=col,
                                value=f"=SUBTOTAL(9, {col_l}{data_first}:{col_l}{data_last})")

            cell.font = SUBTOTAL_FONT
            cell.fill = SUBTOTAL_FILL
            sample_fmt = ws.cell(row=new_header_row + 1, column=col).number_format
            if sample_fmt and sample_fmt != "General":
                cell.number_format = sample_fmt
            else:
                cell.number_format = FMT_DLR if "$" in str(ws.cell(row=new_header_row, column=col).value or "") else FMT_INT

        # Fill rest of the subtotal row with same yellow band
        for col in range(start_col, end_col + 1):
            if ws.cell(row=subtotal_row, column=col).value is None:
                ws.cell(row=subtotal_row, column=col).fill = SUBTOTAL_FILL

        # Update autofilter ref to the new shifted positions
        ws.auto_filter.ref = f"{start_col_letter}{new_header_row}:{end_col_letter}{new_last_data_row}"
        ws.freeze_panes = f"A{new_header_row + 1}"

        log(f"  SUBTOTAL row added to {sheet_name}; filter now {ws.auto_filter.ref}; sum cols: {len(sum_cols)}")


def main():
    log("=" * 80)
    log("Building NOR CAL Forward Looking INV vs Sales Plan workbook")
    log("=" * 80)

    # Archive any prior reports before generating new (project rule per CEO)
    archive_previous_reports()

    # Run model
    log("Step 1/3: Running model...")
    plan = M.load_plan()
    v158 = M.load_v158_region_dataset()
    dmnd = M.load_v158_demand_data()
    inv = M.load_inventory()
    hist = M.load_history()
    ytd  = M.load_ytd_actuals()
    hd_xref    = M.load_hd_xref()
    lowes_xref = M.load_lowes_xref()
    items = M.build_item_universe(plan, inv, dmnd)
    items = M.attach_ki_grouping(items, v158)
    plan = plan.merge(items[["Item Num", "KI"]].drop_duplicates(), on="Item Num", how="left")
    lift_groups = M.build_lift_groups(items, hist, plan, hd_xref, lowes_xref)
    smoothed = M.compute_smoothed_history(hist, lift_groups)
    plan_lifted = M.apply_lift(plan, smoothed, lift_groups)
    pools = M.build_pools(inv, items)
    offplan_demand = M.build_offplan_stage2_demand(smoothed)
    pool_walk, cust_alloc, offplan_alloc = M.run_walk(plan_lifted, pools, offplan_demand)
    prices = M.build_price_cascade(items, hist, ytd, lift_groups)

    d = {
        "plan": plan, "v158": v158, "items": items,
        "ytd": ytd, "hist": hist,
        "lift_groups": lift_groups,
        "smoothed": smoothed, "plan_lifted": plan_lifted,
        "pools": pools, "pool_walk": pool_walk,
        "cust_alloc": cust_alloc, "offplan_alloc": offplan_alloc,
        "offplan_demand": offplan_demand, "prices": prices,
        "inv": inv,
    }

    # Compute metrics
    log("Step 2/3: Computing metrics...")
    m = compute_metrics(d)

    # Build workbook
    log("Step 3/3: Building workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    write_tab1_readme(wb)                       # Tab 1
    write_tab_build_health(wb, m, d)            # Tab 2: NEW Build Health dashboard
    write_tab_changes(wb)                       # Tab 3: version history (was Tab 2)
    write_tab2_exec(wb, m, d)                   # Tab 3: Exec Summary
    write_tab_plan_by_ki(wb, m)                 # Tab 4: Plan by KI
    write_tab_ki_lifecycle_units(wb, m)         # Tab 5: KI Lifecycle (Units) — NEW
    write_tab_ki_lifecycle_dollars(wb, m)       # Tab 6: KI Lifecycle ($) — NEW
    write_tab_miss_summary_by_ki(wb, m)         # Tab 7: Miss Summary by KI
    write_tab_miss_by_cust_ki(wb, m)            # Tab 6: Miss by Customer x KI (NEW)
    write_tab_miss_by_ki_month(wb, m)           # Tab 7: Miss by KI x Month (NEW)
    write_tab4_short_items(wb, m, d)            # Tab 8: Customer Miss Detail (RENAMED)
    write_tab_production_recovery(wb, m)        # Tab 9: Production Recovery (NEW — recoverable miss action list)
    write_tab_lift_summary_by_ki_only(wb, m)    # Tab 10: Lift Summary by KI (NEW, KI-only)
    write_tab_lift_by_cust_ki(wb, m, d)         # Tab 10: Lift by Customer x KI (RENAMED)
    write_tab_lift_by_ki_month(wb, m, d)        # Tab 11: Lift by KI x Month (NEW)
    write_tab_lift_by_cust_item(wb, m)          # Tab 12: Lift by Customer x Item (was Lift Detail by Month)
    write_tab_excess_at_farm(wb, m)             # Tab 13: Excess by KI (pool leftover)
    write_tab6_excess_ki(wb, m)                 # Tab 14: Over-Plan by KI (renamed from Customer Over-Plan by KI)
    write_tab7_excess_cust_ki(wb, m, d)         # Tab 15: Customer Over-Plan Detail
    write_tab8_channel_summary(wb, m, d)           # Tab 16
    write_tab9_ytd(wb, m)                       # Tab 17
    write_tab_pricing_comparison(wb, m, d)      # Tab 18

    log("Adding SUBTOTAL rows to detail tabs...")
    add_subtotal_rows_to_workbook(wb)

    apply_tab_colors_and_verify(wb)

    log(f"Saving to {OUT_PATH}...")
    wb.save(OUT_PATH)
    log(f"DONE. Saved: {OUT_PATH.name}")


if __name__ == "__main__":
    main()
