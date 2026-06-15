"""
build_or_workbook_patched.py
────────────────────────────
Everde Growers — Oregon (OR) Forward-Looking Workbook Builder
Cloned and adapted from build_norcal_workbook_patched.py (May 2026)

Runs the OR forward fulfillment model and produces:
  OR Forward Looking YTD Miss vs Inventory MMDDYY.xlsx

Output location (network share):
  \\192.168.190.10\\Claude Sandbox\\DataDrops\\Sales Plan Review\

Same 8-tab layout as NOR CAL workbook so portal extract_sales_plan.py
can read it with the same sheet name conventions.

Tabs produced:
  1. Exec Summary
  2. YTD Performance
  3. Miss by KI
  4. Miss by Customer
  5. Plan by KI
  6. Excess at Farm
  7. Historical Lift
  8. Channel Summary

Usage:
    python build_or_workbook_patched.py
    python build_or_workbook_patched.py --inv "path/to/Inventory.xlsx" --ytd "path/to/Sales.xlsx"

Dependencies:
    pip install pandas openpyxl pyxlsb polars fastexcel pyarrow
"""

import sys
import os
import re
import argparse
from pathlib import Path
from datetime import datetime, date
import warnings
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning)

# ─────────────────────────────────────────────────────
# PATH CONFIG — matches the network share layout
# ─────────────────────────────────────────────────────

SHARE_BASE   = Path(r"\\\192.168.190.10\\Claude Sandbox")
DATADROPS    = SHARE_BASE / "DataDrops" / "Sales Plan Review"
JS_FILES     = SHARE_BASE / "JS Files"
SCRIPTS_DIR  = JS_FILES / "Sales Plan Review"

# Source files (same shared files used by NOR CAL)
SHARED_DIR   = JS_FILES / "Shared"
XREF_DIR     = SHARED_DIR / "Inventory Cross References"

# ─────────────────────────────────────────────────────
# ARGUMENT PARSING
# ─────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Everde OR Workbook Builder")
    p.add_argument("--inv",   default=None, help="Inventory Transform xlsx path")
    p.add_argument("--ytd",   default=None, help="2026 Sales by Item xlsx path")
    p.add_argument("--plan",  default=None, help="2026 Sales Plan by Item xlsx path (optional)")
    p.add_argument("--v158",  default=None, help="Key Item Report V158 xlsx path (optional)")
    p.add_argument("--out",   default=None, help="Output xlsx path (default: auto-named)")
    p.add_argument("--no-share", action="store_true",
                   help="Don't copy output to DataDrops share (local run only)")
    return p.parse_args()


def resolve_paths(args):
    """Find all required files, searching standard locations."""
    base = Path(__file__).parent

    def find_latest(folder: Path, patterns: list) -> Path | None:
        if not folder.exists():
            return None
        candidates = []
        for pat in patterns:
            candidates.extend(folder.glob(pat))
        return max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None

    def find(arg_val, *search_dirs_and_patterns):
        if arg_val and Path(arg_val).exists():
            return Path(arg_val)
        for item in search_dirs_and_patterns:
            if isinstance(item, (str, Path)) and Path(item).exists():
                return Path(item)
        return None

    # Weekly drop folder
    weekly_drop = SHARE_BASE / "DataDrops" / "SalesOpportunity"

    # Inventory Transform — check weekly drop first, then local
    inv = find(args.inv) or find_latest(weekly_drop, [
        "Inventory_Transform_*.xlsx", "Inventory Transform*.xlsx"
    ]) or find_latest(base, ["Inventory_Transform_*.xlsx", "Inventory Transform*.xlsx"])

    # YTD Sales — check weekly drop first
    ytd = find(args.ytd) or find_latest(weekly_drop, [
        "2026_Sales_by_Item_*.xlsx", "2026 Sales by Item*.xlsx"
    ]) or find_latest(base, ["2026_Sales_by_Item_*.xlsx", "2026 Sales by Item*.xlsx"])

    # Stable files — check scripts dir, then base
    plan_candidates = [
        SCRIPTS_DIR / "2026_Sales_Plan_by_Item.xlsx",
        SCRIPTS_DIR / "2026 Sales Plan by Item.xlsx",
        base / "2026_Sales_Plan_by_Item.xlsx",
        base / "2026 Sales Plan by Item.xlsx",
    ]
    plan = find(args.plan, *[str(c) for c in plan_candidates])

    v158_candidates = [
        SCRIPTS_DIR / "Key_Item_Report_V158.xlsx",
        SCRIPTS_DIR / "Key Item Report V158.xlsx",
        base / "Key_Item_Report_V158.xlsx",
    ]
    v158 = find(args.v158, *[str(c) for c in v158_candidates])

    # xrefs
    hd    = find_latest(XREF_DIR, ["Home_Depot_Corp-VN_PO_xref*.xlsb",
                                    "Home Depot Corp-VN*PO xref*.xlsb"]) or \
            find_latest(base, ["Home_Depot_Corp-VN*.xlsb"])
    lowes = find_latest(XREF_DIR, ["LOWE_S_xref*.xlsb", "LOWE'S xref*.xlsb"]) or \
            find_latest(base, ["LOWE*xref*.xlsb"])

    # Output filename
    today = datetime.now().strftime("%m%d%y")
    out_filename = f"OR_Forward_Looking_YTD_Miss_vs_Inventory_{today}.xlsx"
    if args.out:
        out_path = Path(args.out)
    else:
        out_path = base / out_filename

    return {
        "inv":    inv,
        "ytd":    ytd,
        "plan":   plan,
        "v158":   v158,
        "hd":     hd,
        "lowes":  lowes,
        "out":    out_path,
        "cache":  base / "cache_or",
    }


# ─────────────────────────────────────────────────────
# EXCEL STYLING
# (matches NOR CAL workbook aesthetic)
# ─────────────────────────────────────────────────────

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Color palette (matches Everde portal)
C_DARK_BG    = "FF11161E"  # Sidebar dark
C_GREEN      = "FF2F5233"  # Everde green
C_GOLD       = "FFC49B3F"  # Gold accent
C_RED        = "FFC0392B"  # Red / miss
C_NAVY       = "FF1F3A5F"  # Navy
C_GREY       = "FF404040"  # Grey
C_PURPLE     = "FF5B4F8A"  # Purple
C_WHITE      = "FFFFFFFF"
C_LIGHT_GREY = "FFF5F5F5"
C_LIGHT_GREEN= "FFE8F0E9"
C_LIGHT_RED  = "FFFDECEA"
C_HEADER_BG  = "FF1F3A5F"  # Navy header

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="FF000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin():
    thin = Side(style="thin", color="FFD0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _style_header_row(ws, row_num: int, num_cols: int,
                      bg=C_HEADER_BG, fg=C_WHITE, height=22):
    ws.row_dimensions[row_num].height = height
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill  = _fill(bg)
        cell.font  = _font(bold=True, color=fg, size=10)
        cell.alignment = _align("center")

def _auto_col_width(ws, min_w=8, max_w=35):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def _write_df_to_sheet(ws, df: pd.DataFrame, start_row=1, header=True,
                        header_bg=C_HEADER_BG, zebra=True):
    """Write a DataFrame to a worksheet with Everde styling."""
    if header:
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.fill  = _fill(header_bg)
            cell.font  = _font(bold=True, color=C_WHITE, size=10)
            cell.alignment = _align("center")
        start_row += 1

    for row_idx, row in enumerate(df.itertuples(index=False), start_row):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Convert numpy types
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif isinstance(val, float) and np.isnan(val):
                val = None
            cell.value = val
            cell.alignment = _align("right" if isinstance(val, (int, float)) else "left")
            if zebra and (row_idx % 2 == 0):
                cell.fill = _fill("FFF8F8F8")

    return start_row + len(df)


# ─────────────────────────────────────────────────────
# TAB BUILDERS
# ─────────────────────────────────────────────────────

def _build_exec_summary(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Exec Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    row = 1

    # Title banner
    ws.merge_cells(f"A{row}:E{row}")
    title = ws.cell(row=row, column=1,
                    value=f"Oregon (OR) Forward-Looking YTD Miss vs Inventory")
    title.fill = _fill(C_GREEN)
    title.font = _font(bold=True, color=C_WHITE, size=14)
    title.alignment = _align("center")
    ws.row_dimensions[row].height = 30
    row += 1

    # Subtitle
    ws.merge_cells(f"A{row}:E{row}")
    snap = getattr(d, "snapshot_date", datetime.now().strftime("%m/%d/%Y"))
    sub = ws.cell(row=row, column=1,
                  value=f"Snapshot: {snap}  |  YTD: Jan–May 2026  |  Forward: Jun–Dec 2026")
    sub.fill = _fill(C_NAVY)
    sub.font = _font(color=C_WHITE, size=10)
    sub.alignment = _align("center")
    ws.row_dimensions[row].height = 18
    row += 2

    # ── KPI boxes ──
    kpis = getattr(d, "kpis", {})
    kpi_data = [
        ("Total OR Key Items",        kpis.get("total_ki_count", 0),    ""),
        ("Below Plan (YTD)",          kpis.get("ki_below_plan", 0),     ""),
        ("YTD Achievement",           f"{kpis.get('ytd_achievement_pct', 0):.1f}%", ""),
        ("YTD Miss (units)",          f"{kpis.get('ytd_miss_units', 0):,.0f}", ""),
        ("Short Forward (Jun–Dec)",   kpis.get("ki_short_fwd", 0),      ""),
        ("Excess at Farm KIs",        kpis.get("ki_excess_count", 0),   ""),
    ]

    ws.merge_cells(f"A{row}:E{row}")
    hdr = ws.cell(row=row, column=1, value="KEY PERFORMANCE INDICATORS")
    hdr.fill = _fill(C_GOLD)
    hdr.font = _font(bold=True, color="FF000000", size=11)
    hdr.alignment = _align("center")
    ws.row_dimensions[row].height = 20
    row += 1

    for label, value, note in kpi_data:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = _align("right")
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1

    # ── Walk tables ──
    for walk_key, walk_attr in [
        ("Year-End Walk",  "walk_ye"),
        ("Forward Walk",   "walk_fwd"),
        ("YTD Walk",       "walk_ytd"),
    ]:
        walk = getattr(d, walk_attr, None)
        if not walk:
            continue

        ws.merge_cells(f"A{row}:E{row}")
        w_hdr = ws.cell(row=row, column=1, value=walk.get("label", walk_key))
        w_hdr.fill = _fill(C_NAVY)
        w_hdr.font = _font(bold=True, color=C_WHITE)
        w_hdr.alignment = _align("center")
        ws.row_dimensions[row].height = 18
        row += 1

        for k, v in walk.items():
            if k == "label":
                continue
            ws.cell(row=row, column=1, value=k.replace("_", " ").title())
            cell_v = ws.cell(row=row, column=2, value=v)
            cell_v.alignment = _align("right")
            row += 1
        row += 1

    # ── Top 20 KIs by Miss ──
    if d.top_miss is not None and len(d.top_miss) > 0:
        ws.merge_cells(f"A{row}:E{row}")
        tm_hdr = ws.cell(row=row, column=1, value="TOP 20 KEY ITEMS BY MISS (YTD)")
        tm_hdr.fill = _fill(C_RED.replace("FF", ""))
        tm_hdr.fill = _fill(C_RED)
        tm_hdr.font = _font(bold=True, color=C_WHITE)
        tm_hdr.alignment = _align("center")
        ws.row_dimensions[row].height = 18
        row += 1

        tm_cols = ["_item_key", "ytd_plan_units", "ytd_actual_units", "ytd_miss_units", "ytd_miss_pct"]
        tm_df   = d.top_miss[[c for c in tm_cols if c in d.top_miss.columns]].copy()
        tm_df.columns = [c.replace("_", " ").title() for c in tm_df.columns]
        _write_df_to_sheet(ws, tm_df, start_row=row)
        row += len(tm_df) + 2


def _build_ytd_performance(wb: Workbook, d) -> None:
    ws = wb.create_sheet("YTD Performance")
    ws.sheet_view.showGridLines = False

    # Title
    ws.cell(row=1, column=1, value="Oregon YTD Performance — Jan–May 2026").font = \
        _font(bold=True, size=12)
    ws.row_dimensions[1].height = 22

    if d.ytd_miss is not None and len(d.ytd_miss) > 0:
        df = d.ytd_miss.copy()
        df.columns = [c.replace("_", " ").replace("ytd ", "").title() for c in df.columns]
        _write_df_to_sheet(ws, df, start_row=3)

    _auto_col_width(ws)


def _build_miss_by_ki(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Miss by KI")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Oregon — Miss by Key Item (YTD Jan–May 2026)").font = \
        _font(bold=True, size=12)

    if d.ytd_miss is not None:
        df = d.ytd_miss.sort_values("ytd_miss_units", ascending=True).copy()
        df.columns = [c.replace("_", " ").title() for c in df.columns]
        _write_df_to_sheet(ws, df, start_row=3)

    _auto_col_width(ws)


def _build_miss_by_customer(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Miss by Customer")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Oregon — Miss by Customer / Channel").font = \
        _font(bold=True, size=12)

    if d.channel is not None:
        _write_df_to_sheet(ws, d.channel, start_row=3)

    _auto_col_width(ws)


def _build_plan_by_ki(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Plan by KI")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Oregon — Full Plan by Key Item (2026)").font = \
        _font(bold=True, size=12)

    if d.fwd_gap is not None:
        df = d.fwd_gap.copy()
        df.columns = [c.replace("_", " ").title() for c in df.columns]
        _write_df_to_sheet(ws, df, start_row=3)

    _auto_col_width(ws)


def _build_excess_at_farm(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Excess at Farm")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Oregon — Excess Inventory at Farm").font = \
        _font(bold=True, size=12)

    if d.excess is not None and len(d.excess) > 0:
        df = d.excess.copy()
        df.columns = [c.replace("_", " ").title() for c in df.columns]
        _write_df_to_sheet(ws, df, start_row=3)
    else:
        ws.cell(row=3, column=1, value="No excess inventory identified for OR.").font = \
            _font(italic=True, color="FF888888")

    _auto_col_width(ws)


def _build_historical_lift(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Historical Lift")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Oregon — Historical Lift (3-yr Smoothed History > Plan)").font = \
        _font(bold=True, size=12)

    if d.hist_lift is not None and len(d.hist_lift) > 0:
        df = d.hist_lift.copy()
        df.columns = [c.replace("_", " ").title() for c in df.columns]
        _write_df_to_sheet(ws, df, start_row=3)
    else:
        ws.cell(row=3, column=1,
                value="Historical lift data requires hist_or_YYYY.parquet cache files.").font = \
            _font(italic=True, color="FF888888")
        ws.cell(row=4, column=1,
                value="Place cache files in: cache_or/ subdirectory alongside this script.").font = \
            _font(italic=True, color="FF888888")

    _auto_col_width(ws)


def _build_channel_summary(wb: Workbook, d) -> None:
    ws = wb.create_sheet("Channel Summary")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Oregon — Channel / Customer Summary").font = \
        _font(bold=True, size=12)

    if d.channel is not None:
        _write_df_to_sheet(ws, d.channel, start_row=3)

    _auto_col_width(ws)


# ─────────────────────────────────────────────────────
# WORKBOOK BUILDER
# ─────────────────────────────────────────────────────

def build_workbook(d, out_path: Path) -> Path:
    """Build the 8-tab OR workbook from the data object."""
    print(f"\nBuilding OR workbook: {out_path.name}")

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("  Tab 1: Exec Summary...")
    _build_exec_summary(wb, d)

    print("  Tab 2: YTD Performance...")
    _build_ytd_performance(wb, d)

    print("  Tab 3: Miss by KI...")
    _build_miss_by_ki(wb, d)

    print("  Tab 4: Miss by Customer...")
    _build_miss_by_customer(wb, d)

    print("  Tab 5: Plan by KI...")
    _build_plan_by_ki(wb, d)

    print("  Tab 6: Excess at Farm...")
    _build_excess_at_farm(wb, d)

    print("  Tab 7: Historical Lift...")
    _build_historical_lift(wb, d)

    print("  Tab 8: Channel Summary...")
    _build_channel_summary(wb, d)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    size_kb = out_path.stat().st_size / 1024
    print(f"\n  ✓ Saved: {out_path}")
    print(f"    Size: {size_kb:.1f} KB")
    print(f"    Tabs: {len(wb.sheetnames)}")
    return out_path


# ─────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────

def main():
    args = parse_args()
    paths = resolve_paths(args)

    print("=" * 60)
    print("Everde OR Forward-Looking Workbook Builder")
    print(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    print()

    # Report resolved paths
    for k, v in paths.items():
        if k == "cache":
            continue
        status = "✓" if v and Path(v).exists() else "✗ NOT FOUND"
        print(f"  {k:<10} {status}  {v}")
    print()

    # Validate required files
    for req in ["inv", "ytd"]:
        if not paths[req] or not paths[req].exists():
            print(f"ERROR: Required file not found: {req} = {paths[req]}")
            print("Pass --inv and --ytd arguments, or place files in the standard locations.")
            sys.exit(1)

    # Import and run the OR model
    import importlib.util
    base = Path(__file__).parent
    model_path = base / "or_forward_patched.py"

    if not model_path.exists():
        print(f"ERROR: or_forward_patched.py not found at {model_path}")
        sys.exit(1)

    spec = importlib.util.spec_from_file_location("or_forward_patched", model_path)
    M = importlib.util.module_from_spec(spec)

    # Inject paths before running
    M.PATH_INV        = paths["inv"]
    M.PATH_YTD        = paths["ytd"]
    M.PATH_PLAN       = paths["plan"] or (base / "2026_Sales_Plan_by_Item.xlsx")
    M.PATH_V158       = paths["v158"] or (base / "Key_Item_Report_V158.xlsx")
    M.PATH_HD_XREF    = paths["hd"]   or (base / "Home_Depot_Corp-VN_PO_xref_rev_04222026.xlsb")
    M.PATH_LOWES_XREF = paths["lowes"] or (base / "LOWE_S_xref_rev_04292026.xlsb")
    M.CACHE_DIR       = paths["cache"]

    spec.loader.exec_module(M)
    M.run()
    d = M.d

    # Build workbook
    out_path = build_workbook(d, paths["out"])

    # Copy to DataDrops share
    if not args.no_share:
        try:
            import shutil
            share_out = DATADROPS / out_path.name
            DATADROPS.mkdir(parents=True, exist_ok=True)
            shutil.copy2(out_path, share_out)
            print(f"\n  ✓ Copied to share: {share_out}")
        except Exception as e:
            print(f"\n  WARNING: Could not copy to share: {e}")
            print(f"  Manual copy: {out_path} → {DATADROPS}")

    print(f"\n{'='*60}")
    print(f"  Output workbook: {out_path.name}")
    print(f"  Tabs: {8}")
    print("  Drop to portal: DataDrops\\Sales Plan Review\\")
    print(f"{'='*60}")
    print("\nDONE.")
    return out_path


if __name__ == "__main__":
    main()
