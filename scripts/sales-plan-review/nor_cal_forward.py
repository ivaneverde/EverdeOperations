"""
nor_cal_forward.py
==================

Everde Growers — Forward Fulfillment & Sales Plan Analysis (NOR CAL pilot)
Owner: Jonathan Saperstein (CEO)
Audience: ABOD + Sales/Ops leaders

Reproduces the methodology documented in NEW_PROJECT_REFERENCE.md and the
2-tab validation output in NOR CAL Forward Validation 050226.xlsx (Summary +
BUXUS Detail). Validation simplification is in effect: all customers route
through Key Item Grouping pool for substitution (HD/Lowes SKU specifics OFF).

USAGE
-----
    python3 nor_cal_forward.py
    python3 nor_cal_forward.py --rebuild-cache   # force re-build per-year cache

OUTPUTS
-------
    NOR CAL Forward Reproduction <MMDDYY>.xlsx  (Summary + BUXUS Detail)

LOCKED-IN CONVENTIONS (see NEW_PROJECT_REFERENCE.md for full discussion)
- SNAP_DATE = 2026-04-27, FWD_MONTHS = May..Dec, YTD_MONTHS = Jan..Apr
- Customer buckets: HD, Lowes, Walmart, West Coast, Midwest (5)
- BB = HD/Lowes/Walmart  ; NBB = West Coast/Midwest
- Inventory pool eligibility per grade × ready-date (see build_pools)
- Demand Window filter: 2026 H1, 2026 H2, 2025, 2024, blank only
- Cross-farm fulfillment: allowed within same Region
- Smoothing: drop years > 10x the 3-yr baseline mean, re-average remaining
- Validation simplification: KI-Grouping aggregation for ALL channels (lift
  + substitution), HD/Lowes SKU files NOT consulted
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

PROJECT_DIR = Path("/sessions/optimistic-beautiful-ramanujan/mnt/Sales Plan Review")
KEY_ITEM_DIR = Path("/sessions/optimistic-beautiful-ramanujan/mnt/Key Item Review")
SHARED_DIR = Path("/sessions/optimistic-beautiful-ramanujan/mnt")  # 'Shared' subfolder lives below

CACHE_DIR = PROJECT_DIR / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# Source files
PATH_PLAN     = KEY_ITEM_DIR / "2026 Sales Plan by Item.xlsx"
PATH_V158     = KEY_ITEM_DIR / "Key Item Report V158.xlsx"
PATH_INV      = SHARED_DIR / "Shared" / "INV" / "Inventory Transform 042726.xlsx"  # moved 2026-05-08
PATH_YTD      = SHARED_DIR / "Shared" / "Sales Data" / "2026 Sales by Item 042726.xlsx"
PATH_HIST = {
    2023: SHARED_DIR / "Shared" / "Sales Data" / "2023 Sales by Item.xlsx",
    2024: SHARED_DIR / "Shared" / "Sales Data" / "2024 Sales by Item.xlsx",
    2025: SHARED_DIR / "Shared" / "Sales Data" / "2025 Sales by Item.xlsx",
}
PATH_HD_XREF    = SHARED_DIR / "Shared" / "Inventory Cross References" / "Home Depot Corp-VN=PO xref rev.04222026.xlsb"
PATH_LOWES_XREF = SHARED_DIR / "Shared" / "Inventory Cross References" / "LOWE'S xref rev.04292029.xlsb"
PATH_GROW_TIMES = SHARED_DIR / "Shared" / "Misc Look Ups" / "Prod lookups ALL 091925.xlsx"

# Per-region production org codes (from plan file). NOR CAL ships from WIN and BRA.
# Used by load_grow_times() to read only the relevant tabs.
REGION_ORG_CODES = {
    "NOR CAL": ["WIN", "BRA"],
    "FL":      ["BNL"],
    "OR":      ["FOR", "PAU"],   # placeholder; verify when scaling to OR
    "SO CAL":  ["MIR", "ESC"],   # placeholder; verify when scaling to SO CAL
    "TX":      ["HUN", "STE"],   # placeholder; verify when scaling to TX
}

# Constants
REGION = "NOR CAL"
SNAP_DATE = pd.Timestamp("2026-04-27")
PRIOR_SNAP_DATE = pd.Timestamp("2026-04-20")
FWD_MONTHS = list(range(5, 13))   # May..Dec
YTD_MONTHS = [1, 2, 3, 4]
HIST_YEARS = [2023, 2024, 2025]
OUTLIER_MULT = 10.0

# Canonical company-wide channel universe (per CEO 2026-05-06)
# CHANNEL_TYPE keys are the 7 Planning Customer buckets that appear across the
# company's 5 regions in 2026 Sales Plan by Item.xlsx, plus an "Other" catch-all.
# BB = Big Box (no pre-ready pool access); NBB = independent/regional (gets
# first dibs on the SS/GS restricted pre-ready pool).
CHANNEL_TYPE = {
    "HD":         "BB",
    "Lowes":      "BB",
    "Walmart":    "BB",
    "West Coast": "NBB",
    "Midwest":    "NBB",
    "Southeast":  "NBB",
    "MLC":        "NBB",
    "Other":      "BB",   # catch-all; unknown -> safe BB-like (no pre-ready access)
}

# ALL_CUSTOMERS / BB_CUSTOMERS / NBB_CUSTOMERS are CANONICAL company-wide sets
# derived from CHANNEL_TYPE — not region-specific. This is the universe used
# for filtering history/YTD loaders and for Stage 2 (lift surplus) demand,
# which can include off-plan channels.
ALL_CUSTOMERS = tuple(CHANNEL_TYPE.keys())
BB_CUSTOMERS  = tuple(c for c in ALL_CUSTOMERS if CHANNEL_TYPE[c] == "BB")
NBB_CUSTOMERS = tuple(c for c in ALL_CUSTOMERS if CHANNEL_TYPE[c] == "NBB")

# PLAN_CUSTOMERS is the REGION-SPECIFIC planned-customer subset, loaded
# dynamically by load_plan() from the plan file. Stage 1 (defend plan) only
# allocates to these customers. Default below is the NOR CAL legacy set.
PLAN_CUSTOMERS = ("HD", "Lowes", "Walmart", "West Coast", "Midwest")


def _set_region_plan_customers(planned_customers) -> None:
    """Pin module-level PLAN_CUSTOMERS to the planned-customer set for the
    current REGION. Called from load_plan() once the plan file is read."""
    global PLAN_CUSTOMERS
    PLAN_CUSTOMERS = tuple(planned_customers)


# Backward-compat shim retained so any caller importing this still works.
_set_region_customer_sets = _set_region_plan_customers


DW_KEEP = {"2026 HALF 1", "2026 HALF 2", "2025", "2024", None, ""}

MONTH_NAME_TO_NUM = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
MONTH_NUM_TO_NAME = {v: k for k, v in MONTH_NAME_TO_NUM.items()}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def map_demand_channel(ch) -> Optional[str]:
    """Map a 'Demand Channel' string to one of the 7 canonical Planning Customer
    buckets, or to 'Other' as a catch-all. Returns None ONLY for null inputs.

    Per CEO confirmation 2026-05-06: every actual sales channel maps to one of
    the 7 canonical buckets (HD, Lowes, Walmart, West Coast, Midwest, Southeast,
    MLC) so the same model can run for any region; anything else folds into
    'Other' rather than being dropped.

    Sub-channel rollups:
      - COSTCO PNW, SITEONE MIDWEST, GREEN ACRES -> Midwest (regional distributors)
      - HD FL - SOUTHEAST and other "HD ... SOUTHEAST" hybrids -> Southeast
        (the Southeast-region distribution arm, not big-box HD demand)
    """
    if ch is None or pd.isna(ch):
        return None
    s = str(ch).strip().upper()
    if not s:
        return None

    # Hybrid HD-Southeast variants are Southeast distribution, not big-box HD.
    if "SOUTHEAST" in s and s.startswith("HD"):
        return "Southeast"
    if s.startswith("HD "):
        return "HD"
    if s.startswith("LOWES"):
        return "Lowes"
    if s.startswith("WM ") or s.startswith("WALMART"):
        return "Walmart"
    if s.startswith("WEST COAST"):
        return "West Coast"
    # Midwest + sub-channels that roll up to Midwest
    if s == "MIDWEST" or s.startswith("MIDWEST"):
        return "Midwest"
    if s == "COSTCO PNW" or s.startswith("COSTCO"):
        return "Midwest"
    if s.startswith("SITEONE MIDWEST"):
        return "Midwest"
    if s.startswith("GREEN ACRES"):
        return "Midwest"
    # Southeast variants
    if s.startswith("SOUTHEAST") or "SOUTHEAST" in s:
        return "Southeast"
    # MLC variants (e.g. "MLC", "MLC SO CAL", etc.)
    if s == "MLC" or s.startswith("MLC"):
        return "MLC"
    # Anything else -> Other catch-all
    return "Other"


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_plan() -> pd.DataFrame:
    """2026 Sales Plan, current REGION × planned-customer subset. Returns long
    format (one row per item × customer × month).

    Derives the planned-customer set DYNAMICALLY from the plan file rather than
    hard-coding it. This makes the model region-agnostic: NOR CAL gets the
    HD/Lowes/Walmart/West Coast/Midwest set; SO CAL adds Southeast and MLC; FL
    has only HD/Midwest/Southeast; etc. The set is restricted to canonical
    company-wide buckets defined in CHANNEL_TYPE.

    Side-effect: calls _set_region_customer_sets() to pin module-level
    BB_CUSTOMERS / NBB_CUSTOMERS / ALL_CUSTOMERS for downstream functions.
    """
    log("Loading 2026 Sales Plan...")
    df = pd.read_excel(PATH_PLAN, sheet_name="2026 Sales Plan")
    df = df[df["Region"] == REGION].copy()
    # Derive the planned-customer set for this region from the plan itself.
    canonical = set(CHANNEL_TYPE.keys())
    region_customers = sorted(set(df["Planning Customer"].dropna().unique()) & canonical)
    if not region_customers:
        raise ValueError(f"No planned customers found in plan for region={REGION}; "
                         f"check that the region name matches the plan file.")
    _set_region_plan_customers(region_customers)
    log(f"  region={REGION} planned customers: {region_customers}")
    df = df[df["Planning Customer"].isin(region_customers)].copy()
    df["MonthNum"] = df["Month"].map(MONTH_NAME_TO_NUM)
    df = df.rename(columns={"Planning Customer": "Customer", "Total Qty": "PlanQty"})
    keep = ["Org Code", "Region", "Item Num", "Customer", "MonthNum", "PlanQty"]
    df = df[keep]
    df["PlanQty"] = pd.to_numeric(df["PlanQty"], errors="coerce").fillna(0.0)
    log(f"  plan rows: {len(df):,}  total qty: {df['PlanQty'].sum():,.0f}")
    return df


def load_v158_region_dataset() -> pd.DataFrame:
    """V158 Region Dataset: one row per (Region, Item or Grouping). Provides the
    canonical Key Item Grouping label and Demand Price per Item."""
    log("Loading V158 Region Dataset...")
    df = pd.read_excel(PATH_V158, sheet_name="Region Dataset", header=7)
    keep = ["Region Key Item", "Key Item", "Type", "Region", "Size", "Description", "Demand Price"]
    df = df[keep].copy()
    df = df.dropna(subset=["Region", "Region Key Item"])
    log(f"  V158 region rows: {len(df):,}")
    return df


def load_v158_demand_data() -> pd.DataFrame:
    """V158 Demand Data: provides Item -> Brand, Genus, Size, Region, KI Grouping."""
    log("Loading V158 Demand Data...")
    df = pd.read_excel(PATH_V158, sheet_name="Demand Data")
    keep = ["Region Item", "REGION REGION", "Genus", "Oracle Size", "Brand",
            "Part Code", "Description", "Key Item/Grouping"]
    df = df[keep].copy()
    df = df.rename(columns={
        "REGION REGION": "Region",
        "Oracle Size": "Size",
        "Part Code": "Item Num",
        "Description": "Item Desc",
        "Key Item/Grouping": "KI Tagged",
    })
    df = df.dropna(subset=["Item Num"])
    df = df.drop_duplicates(subset=["Region", "Item Num"], keep="first").reset_index(drop=True)
    log(f"  V158 demand-data rows: {len(df):,}")
    return df


def load_inventory() -> pd.DataFrame:
    """Inventory snapshot, NOR CAL only, with Demand Window filter applied and
    grade-based inclusions per CEO confirmation:

      INCLUDE:  A, B, SS, GS, SN, GN, S2N
      EXCLUDE:  C, D, and any P-prefixed grade (PN, P2N, P, etc.)

    The xlsx 'Type' column is always 'FG' for NOR CAL — useless for filtering.
    Filtering happens purely on the 'Grade' column.
    """
    log("Loading Inventory Transform...")
    df = pd.read_excel(PATH_INV, sheet_name="Inventory Dataset")
    df = df[df["Region"] == REGION].copy()
    # Demand Window filter
    df["Demand Window"] = df["Demand Window"].fillna("")
    df = df[df["Demand Window"].isin(DW_KEEP)].copy()
    # Grade filter — exclude C/D and any P-prefixed grade
    df["Grade"] = df["Grade"].astype(str).str.strip()
    df = df[~df["Grade"].isin(["C", "D"])]
    df = df[~df["Grade"].str.startswith("P")]
    # Numeric coercion
    for col in ["Total QTY", "Available QTY", "Soft SO Reserve QTY", "SO Reserve QTY"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["Ready Date Final"] = pd.to_datetime(df["Ready Date Final"], errors="coerce")
    log(f"  inventory rows: {len(df):,}")
    return df


def load_ytd_actuals() -> pd.DataFrame:
    """2026 YTD actuals, NOR CAL × 5 customers, Jan-Apr only."""
    log("Loading 2026 YTD actuals (polars+fastexcel)...")
    import polars as pl
    pdf = pl.read_excel(
        PATH_YTD,
        columns=["Tree", "Region", "445 Year", "445 Month", "Qty Inv SUM",
                 "Demand Channel", "Revenue Amt Sum"],
    ).filter(pl.col("Region") == REGION)
    df = pdf.to_pandas()
    df["Customer"] = df["Demand Channel"].map(map_demand_channel)
    df = df[df["Customer"].notna()].copy()
    # Restrict to region's planned customer set (legacy behavior). When Stage 2
    # extends to off-plan customers, this filter loosens.
    df = df[df["Customer"].isin(ALL_CUSTOMERS)].copy()
    df["Month"] = pd.to_numeric(df["445 Month"], errors="coerce").astype("Int64")
    df["Year"] = pd.to_numeric(df["445 Year"], errors="coerce").astype("Int64")
    # No Year/Month filter — the YTD file is already snapshot-bounded
    # (file dated 042726 = transactions through Sat 4/25/2026 per CEO)
    df = df.rename(columns={"Tree": "Item Num", "Qty Inv SUM": "Qty",
                              "Revenue Amt Sum": "Revenue"})
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0.0)
    df["Revenue"] = pd.to_numeric(df["Revenue"], errors="coerce").fillna(0.0)
    df = df[["Item Num", "Customer", "Month", "Qty", "Revenue"]].copy()
    log(f"  YTD rows: {len(df):,}  total qty: {df['Qty'].sum():,.0f}  total rev: ${df['Revenue'].sum():,.0f}")
    return df


def load_hist_year(year: int, rebuild: bool = False) -> pd.DataFrame:
    """Historical sales for one year, NOR CAL × 5 customers. Cached as parquet.

    Uses openpyxl read_only streaming mode (much faster than pd.read_excel
    on the 200-280MB xlsx files — minutes not tens of minutes).
    """
    cache_path = CACHE_DIR / f"hist_norcal_{year}.parquet"
    if cache_path.exists() and not rebuild:
        log(f"  loading cached hist_norcal_{year}.parquet")
        df = pd.read_parquet(cache_path)
    else:
        log(f"  streaming {year} Sales by Item from xlsx (polars+fastexcel)...")
        import polars as pl
        pdf = pl.read_excel(
            PATH_HIST[year],
            columns=["Tree", "Region", "445 Year", "445 Month", "Qty Inv SUM",
                     "Demand Channel", "Sku"],
        ).filter(pl.col("Region") == REGION)
        pdf.write_parquet(cache_path)
        df = pdf.to_pandas()
    # Normalize columns
    if "Tree" in df.columns and "Item Num" not in df.columns:
        df = df.rename(columns={"Tree": "Item Num"})
    if "445 Year" in df.columns:
        df = df.rename(columns={"445 Year": "Year"})
    if "445 Month" in df.columns:
        df = df.rename(columns={"445 Month": "Month"})
    if "Qty Inv SUM" in df.columns:
        df = df.rename(columns={"Qty Inv SUM": "Qty"})
    if "Customer" not in df.columns:
        df["Customer"] = df["Demand Channel"].map(map_demand_channel)
    df = df[df["Customer"].notna()].copy()
    # Restrict to current region's planned customer set (legacy behavior).
    df = df[df["Customer"].isin(ALL_CUSTOMERS)].copy()
    df["Year"]  = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")
    df["Month"] = pd.to_numeric(df["Month"], errors="coerce").astype("Int64")
    df["Qty"]   = pd.to_numeric(df["Qty"], errors="coerce").fillna(0.0)
    # Normalize Sku: keep as string, drop "NULL" -> NaN
    if "Sku" in df.columns:
        df["Sku"] = df["Sku"].astype(str).str.strip()
        df.loc[df["Sku"].str.upper().isin(["NULL", "NAN", ""]), "Sku"] = pd.NA
    else:
        df["Sku"] = pd.NA
    # Revenue (for pricing cascade); Sell Price (per-row reference)
    if "Revenue Amt Sum" in df.columns:
        df["Revenue"] = pd.to_numeric(df["Revenue Amt Sum"], errors="coerce").fillna(0.0)
    else:
        df["Revenue"] = 0.0
    if "Sell Price" in df.columns:
        df["Sell Price"] = pd.to_numeric(df["Sell Price"], errors="coerce")
    else:
        df["Sell Price"] = pd.NA
    return df[["Item Num", "Customer", "Year", "Month", "Qty", "Sku", "Revenue", "Sell Price"]].copy()


def load_history(rebuild: bool = False) -> pd.DataFrame:
    """All 3 history years stacked."""
    log("Loading historical sales (2023-2025)...")
    parts = [load_hist_year(y, rebuild=rebuild) for y in HIST_YEARS]
    df = pd.concat(parts, ignore_index=True)
    log(f"  total history rows: {len(df):,}")
    return df


# ---------------------------------------------------------------------------
# HD / Lowes cross-reference loaders
# ---------------------------------------------------------------------------

def _load_xlsb_data_tab(path: Path) -> pd.DataFrame:
    """Read the DATA sheet from an HD/Lowes xref xlsb file."""
    from pyxlsb import open_workbook
    rows = []
    with open_workbook(str(path)) as wb:
        with wb.get_sheet("DATA") as ws:
            for r in ws.rows():
                rows.append([c.v for c in r])
    df = pd.DataFrame(rows[1:], columns=rows[0])
    return df


def load_hd_xref() -> pd.DataFrame:
    """HD VN=PO cross-reference. Cols: SKU, Item, Address Category, etc."""
    log("Loading HD xref...")
    df = _load_xlsb_data_tab(PATH_HD_XREF)
    df = df.rename(columns={"Item": "Item Num"})
    df["SKU"] = df["SKU"].astype(str).str.replace(r"\.0$", "", regex=True)
    df["Item Num"] = df["Item Num"].astype(str).str.strip()
    log(f"  HD xref rows: {len(df):,}  unique SKUs: {df['SKU'].nunique()}  unique Items: {df['Item Num'].nunique()}")
    return df[["SKU", "Item Num", "Address Category"]].copy()


def load_lowes_xref() -> pd.DataFrame:
    """Lowe's cross-reference. Same shape as HD xref."""
    log("Loading Lowes xref...")
    df = _load_xlsb_data_tab(PATH_LOWES_XREF)
    df = df.rename(columns={"Item": "Item Num"})
    df["SKU"] = df["SKU"].astype(str).str.replace(r"\.0$", "", regex=True)
    df["Item Num"] = df["Item Num"].astype(str).str.strip()
    log(f"  Lowes xref rows: {len(df):,}  unique SKUs: {df['SKU'].nunique()}  unique Items: {df['Item Num'].nunique()}")
    return df[["SKU", "Item Num", "Address Category"]].copy()


def load_grow_times(region: str = None) -> pd.DataFrame:
    """Load grow days from Prod lookups ALL 091925.xlsx for the relevant region orgs.

    Returns DataFrame with columns: [Org Code, Item Num, GrowDays].
    For each (Org, Item), takes MIN grow days across all entries (per CEO direction
    2026-05-07 — most aggressive estimate of fastest-possible grow time).

    Headers are on row 3 of each tab (rows 1-2 are spacers).
    """
    log("Loading grow times (Prod lookups ALL)...")
    if region is None:
        region = REGION
    org_codes = REGION_ORG_CODES.get(region, [])
    if not org_codes:
        log(f"  WARN: no production orgs configured for region={region}; grow times unavailable")
        return pd.DataFrame(columns=["Org Code", "Item Num", "GrowDays"])

    rows = []
    for org in org_codes:
        try:
            df = pd.read_excel(PATH_GROW_TIMES, sheet_name=org, header=2)
            # Columns: ITEM | ITEM DESCRIPTION | ITEM SIZE | DEMAND HALF | DATE ID | DELIVERY MONTH | GROW DAYS | LINER | LINER PPP | FALL DOWN FACTOR | Initial Type | COMMENTS
            df = df[["ITEM", "GROW DAYS"]].copy()
            df = df.rename(columns={"ITEM": "Item Num", "GROW DAYS": "GrowDays"})
            df["Item Num"] = df["Item Num"].astype(str).str.strip()
            df["GrowDays"] = pd.to_numeric(df["GrowDays"], errors="coerce")
            df = df.dropna(subset=["Item Num", "GrowDays"])
            df = df[df["Item Num"] != ""]
            df = df[df["Item Num"].str.upper() != "MAKE"]
            # Aggregate to MIN per item (most aggressive)
            df = df.groupby("Item Num", as_index=False)["GrowDays"].min()
            df["Org Code"] = org
            rows.append(df[["Org Code", "Item Num", "GrowDays"]])
            log(f"  loaded {org}: {len(df):,} unique items with grow days (min={df['GrowDays'].min():.0f}, max={df['GrowDays'].max():.0f})")
        except Exception as e:
            log(f"  WARN: failed to load {org} tab from grow times file: {e}")
    if not rows:
        return pd.DataFrame(columns=["Org Code", "Item Num", "GrowDays"])
    out = pd.concat(rows, ignore_index=True)
    log(f"  total grow times rows: {len(out):,}")
    return out


# ---------------------------------------------------------------------------
# Item universe + Key Item Grouping mapping
# ---------------------------------------------------------------------------

def build_item_universe(plan: pd.DataFrame, inv: pd.DataFrame, dmnd: pd.DataFrame) -> pd.DataFrame:
    """Union of items in plan, inventory, and demand-data (for Genus/Size/Brand
    fallback). Adds Key Item Grouping label per item.

    Validation simplification: for ALL customers, grouping unit = Key Item
    Grouping (the 'Key Item' label from V158 Demand Data with the
    Genus + Size fallback when blank).
    """
    log("Building item universe...")
    # Items from V158 Demand Data, NOR CAL slice
    nc = dmnd[dmnd["Region"] == REGION].copy()
    nc["KI"] = ""  # placeholder; we'll pull KI from V158 Region Dataset table
    return nc


def attach_ki_grouping(items: pd.DataFrame, v158_region: pd.DataFrame) -> pd.DataFrame:
    """Attach Key Item Grouping label per (Region, Item Num).

    Primary source: V158 Demand Data 'Key Item/Grouping' column (already on
    items). Values like '#001 BUXUS' / '#005 BUXUS'. Numeric 0 or blank
    means untagged → fall back to '#{Size} {Genus}' format which mirrors
    V158's tagged convention.

    Also pulls Demand Price per (Region, Item Num) from V158 Region Dataset.
    """
    rd = v158_region[v158_region["Region"] == REGION].copy()
    rd_pc = rd[rd["Type"] == "Part Code"].copy()
    rd_pc["Item Num"] = rd_pc["Region Key Item"].astype(str).str.split(";", n=1).str[1]
    rd_pc = rd_pc[["Item Num", "Demand Price"]].drop_duplicates(subset=["Item Num"], keep="first")
    items = items.merge(rd_pc, on="Item Num", how="left")

    def _ki(row):
        tag = row.get("KI Tagged")
        if isinstance(tag, str) and tag.strip() and tag.strip() != "0":
            return tag.strip()
        # numeric 0 means untagged — fall back to "#{Size} {Genus}"
        g = (row.get("Genus") or "")
        sz = (row.get("Size") or "")
        g = str(g).strip() if pd.notna(g) else ""
        sz = str(sz).strip() if pd.notna(sz) else ""
        if g and sz:
            sz = sz if sz.startswith("#") else f"#{sz}"
            return f"{sz} {g}"
        return None
    items["KI"] = items.apply(_ki, axis=1)
    return items


# ---------------------------------------------------------------------------
# Inventory pools per Key Item Grouping
# ---------------------------------------------------------------------------

def build_pools(inv: pd.DataFrame, items: pd.DataFrame) -> pd.DataFrame:
    """Construct inventory pools per (KI Grouping, Month) for the May..Dec walk.

    Per pool eligibility table:
      A, B  : in all-customer pool from May 1 (use Available QTY)
      SS,GS pre-ready (Ready > current month-end, Available > 0): NBB-only restricted pool
      SS,GS month-of/post-ready (Ready <= current month-end): all-customer pool, full Total QTY
      SN,GN pre-ready: not in any pool
      SN,GN month-of/post-ready: all-customer pool, full Total QTY
      C, D : excluded
      PN, P2N, S2N: excluded

    Returns a DataFrame with one row per (KI, lot) and columns:
      KI, Item Num, Total QTY, Available QTY, Grade, Ready Month, Inflow Month
        (the calendar month the lot becomes part of all-customer pool)
    """
    log("Building inventory pools per KI Grouping...")
    inv = inv.merge(
        items[["Item Num", "KI"]].drop_duplicates(),
        on="Item Num", how="left",
    )
    # Items with no KI mapping are dropped (can't be allocated to a grouping pool).
    inv = inv[inv["KI"].notna()].copy()

    inv["Ready Month"] = inv["Ready Date Final"].dt.month
    inv["Ready Year"]  = inv["Ready Date Final"].dt.year
    # Inflow Month logic:
    #  - For A/B: in all-pool from May 1 of current year — i.e., already in pool at start of walk.
    #  - For SS/GS/SN/GN pre-ready (Ready Date > SNAP_DATE month-end):
    #      SS/GS go into restricted pool immediately; transfer to all-pool in their Ready Month.
    #      SN/GN sit out until Ready Month, then all-pool with Total QTY.
    #  - For SS/GS/SN/GN with Ready Date <= snap month-end: all-pool with Total QTY at start of walk.
    snap_month_end = SNAP_DATE + pd.offsets.MonthEnd(0)

    grade = inv["Grade"].astype(str).str.strip()
    typ = inv["Type"].astype(str).str.strip()

    # Categorize. Per CEO: A/B in pool from May 1; SS/GS pre-ready in restricted
    # NBB-only pool until ready, then full Total to all-pool; SN/GN/S2N pre-ready
    # not in any pool until ready, then full Total to all-pool.
    is_ab = grade.isin(["A", "B"])
    is_ssgs = grade.isin(["SS", "GS"])
    is_sngn = grade.isin(["SN", "GN", "S2N"])  # S2N treated like SN/GN
    is_pre_ready = inv["Ready Date Final"] > snap_month_end

    # Tag pool kind per lot
    def _categorize(row, is_ab_v, is_ssgs_v, is_sngn_v, is_pre_v):
        if is_ab_v:
            # Per CEO 2026-05-04: use Total QTY (not Available) for A/B grade. SO Reserves
            # represent committed forward orders, not already-shipped inventory, so
            # they should remain in the forward fulfillment pool. Aligns with the user's
            # Key Item Report which uses Total in its Ready Now calculation.
            return ("all_initial", row["Total QTY"])
        if is_ssgs_v:
            if is_pre_v and row["Available QTY"] > 0:
                # Restricted pre-ready, will graduate to all-pool in Ready Month
                return ("restricted_then_all", row["Available QTY"])
            else:
                return ("all_initial", row["Total QTY"])
        if is_sngn_v:
            if is_pre_v:
                return ("sngn_inflow_at_ready", row["Total QTY"])
            else:
                return ("all_initial", row["Total QTY"])
        return ("excluded", 0.0)

    inv = inv.assign(
        _is_ab=is_ab, _is_ssgs=is_ssgs, _is_sngn=is_sngn, _is_pre=is_pre_ready,
    )
    cats = inv.apply(
        lambda r: _categorize(r, r["_is_ab"], r["_is_ssgs"], r["_is_sngn"], r["_is_pre"]),
        axis=1,
    )
    inv["PoolCat"] = [c[0] for c in cats]
    inv["PoolQty"] = [c[1] for c in cats]
    inv = inv[inv["PoolCat"] != "excluded"].copy()

    log(f"  inv pool lots: {len(inv):,} | by cat: {inv['PoolCat'].value_counts().to_dict()}")
    return inv[["KI", "Item Num", "Grade", "Type", "PoolCat", "PoolQty", "Ready Month", "Ready Year",
                "Total QTY", "Available QTY"]]


# ---------------------------------------------------------------------------
# Per-customer LiftGroup assignment (HD/Lowes use SKU; others use KI)
# ---------------------------------------------------------------------------

def build_lift_groups(items: pd.DataFrame, hist: pd.DataFrame, plan: pd.DataFrame,
                      hd_xref: pd.DataFrame, lowes_xref: pd.DataFrame) -> pd.DataFrame:
    """Per-customer item-to-LiftGroup mapping.

    Logic per the user spec:
      HD:
        - Item has HD SKU(s) in xref:
            * If multiple SKUs: pick SKU with majority HD history (the Sku
              column on history transactions).
            * If only one SKU: use it.
            * If item has no HD history at all but multiple SKUs: pick the
              first SKU listed.
          → LiftGroup = f"HDSKU:{sku}"
        - Item has no HD SKU in xref BUT has HD plan QTY > 0 OR HD history > 0:
          → LiftGroup = f"KI:{KI grouping}"  (fallback)
        - Item has no HD SKU AND no HD plan AND no HD history:
          → excluded entirely (HD has never sold it and isn't planning to).

      Lowes: same logic with Lowes xref.
      Walmart, West Coast, Midwest: LiftGroup = f"KI:{KI grouping}".

    Returns a DataFrame with cols [Customer, Item Num, LiftGroup].
    """
    log("Building per-customer LiftGroup assignments...")
    item_ki = items[["Item Num", "KI"]].drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict()

    out_rows = []

    for cust, xref in [("HD", hd_xref), ("Lowes", lowes_xref)]:
        # Items in HD's xref
        cust_xref = xref.copy()
        cust_xref_skus = cust_xref.groupby("Item Num")["SKU"].apply(lambda s: list(dict.fromkeys(s))).to_dict()

        # HD's history rows + their Sku values
        cust_hist = hist[hist["Customer"] == cust].copy()
        # Build (Item Num, Sku) -> total qty across all years/months
        hist_by_item_sku = (cust_hist.dropna(subset=["Sku"])
                                      .groupby(["Item Num", "Sku"])["Qty"].sum().reset_index())
        # Items with any HD history
        items_with_hist = set(cust_hist["Item Num"].dropna().unique())
        # Items in cust's plan
        items_in_plan = set(plan[plan["Customer"] == cust]["Item Num"].dropna().unique())

        candidate_items = set(cust_xref_skus.keys()) | items_with_hist | items_in_plan
        for item in candidate_items:
            xref_skus = cust_xref_skus.get(item, [])
            if xref_skus:
                # Has SKU(s) in xref → use SKU grouping
                if len(xref_skus) == 1:
                    primary = xref_skus[0]
                else:
                    cands = hist_by_item_sku[(hist_by_item_sku["Item Num"] == item) &
                                              (hist_by_item_sku["Sku"].isin(xref_skus))]
                    if not cands.empty:
                        primary = cands.loc[cands["Qty"].idxmax(), "Sku"]
                    else:
                        primary = xref_skus[0]
                lg = f"{cust.upper()}SKU:{primary}"
            else:
                # No SKU in xref. Fallback to KI ONLY if has plan or history.
                if item not in items_in_plan and item not in items_with_hist:
                    continue  # exclude
                ki = item_ki.get(item)
                if ki is None or (isinstance(ki, float) and pd.isna(ki)):
                    continue  # no KI either
                lg = f"KI:{ki}"
            out_rows.append({"Customer": cust, "Item Num": item, "LiftGroup": lg})

    # All non-HD/Lowes canonical customers: KI grouping for all items in plan
    # or history. This includes planned customers (Walmart/WC/Midwest, plus
    # region-specific Southeast/MLC) AND off-plan customers like "Other".
    for cust in ALL_CUSTOMERS:
        if cust in ("HD", "Lowes"):
            continue  # handled above with SKU-level logic
        items_in_plan = set(plan[plan["Customer"] == cust]["Item Num"].dropna().unique())
        items_with_hist = set(hist[hist["Customer"] == cust]["Item Num"].dropna().unique())
        for item in (items_in_plan | items_with_hist):
            ki = item_ki.get(item)
            if ki is None or (isinstance(ki, float) and pd.isna(ki)):
                continue
            out_rows.append({"Customer": cust, "Item Num": item, "LiftGroup": f"KI:{ki}"})

    df = pd.DataFrame(out_rows).drop_duplicates(subset=["Customer", "Item Num"])
    log(f"  LiftGroup rows: {len(df):,}  by customer: {df['Customer'].value_counts().to_dict()}")
    return df


# ---------------------------------------------------------------------------
# Historical lift (smoothed) per (Customer, KI, Month)
# ---------------------------------------------------------------------------

def compute_smoothed_history(hist: pd.DataFrame, lift_groups: pd.DataFrame) -> pd.DataFrame:
    """For each (Customer, LiftGroup, Month) compute the 3-yr smoothed avg
    with the >10x outlier-drop rule.

    Aggregation unit varies per customer:
      HD     -> HD SKU (or KI fallback)  [encoded as 'HDSKU:nnnnn' / 'KI:...']
      Lowes  -> Lowes SKU (or KI fallback)
      Walmart, WC, Midwest -> KI

    `lift_groups` provides the (Customer, Item Num) -> LiftGroup mapping.
    History rows for items not in the customer's lift_groups are excluded.
    """
    log("Computing smoothed historical baseline...")
    h = hist.merge(lift_groups, on=["Customer", "Item Num"], how="inner")
    agg = h.groupby(["Customer", "LiftGroup", "Year", "Month"], dropna=False)["Qty"].sum().reset_index()
    pv = agg.pivot_table(
        index=["Customer", "LiftGroup", "Month"], columns="Year", values="Qty",
        fill_value=0.0,
    ).reset_index()
    for y in HIST_YEARS:
        if y not in pv.columns:
            pv[y] = 0.0
    pv = pv.rename(columns={2023: "y2023", 2024: "y2024", 2025: "y2025"})
    pv["raw_mean"] = pv[["y2023", "y2024", "y2025"]].mean(axis=1)
    # Outlier rule: a year is dropped if its value > OUTLIER_MULT × the mean of the OTHER
    # years. Comparing vs raw_mean (which includes the year being checked) is mathematically
    # impossible to trigger; this fix gives the rule its intended behavior of catching one-off
    # spikes vs the typical baseline. (CEO direction 2026-05-07.)
    def _others_mean(row, year):
        others = [row[f"y{y}"] for y in HIST_YEARS if y != year]
        if all(o == 0 for o in others):
            return float("inf")  # never trigger when other years are all zero
        return sum(others) / len(others)

    pv["out_2023"] = pv.apply(lambda r: r["y2023"] > OUTLIER_MULT * _others_mean(r, 2023), axis=1)
    pv["out_2024"] = pv.apply(lambda r: r["y2024"] > OUTLIER_MULT * _others_mean(r, 2024), axis=1)
    pv["out_2025"] = pv.apply(lambda r: r["y2025"] > OUTLIER_MULT * _others_mean(r, 2025), axis=1)
    def _smooth(row):
        vals = []
        if not row["out_2023"]: vals.append(row["y2023"])
        if not row["out_2024"]: vals.append(row["y2024"])
        if not row["out_2025"]: vals.append(row["y2025"])
        return float(np.mean(vals)) if vals else 0.0
    pv["smoothed"] = pv.apply(_smooth, axis=1)
    return pv  # cols: Customer, LiftGroup, Month, y20xx, raw_mean, out flags, smoothed


def build_offplan_stage2_demand(smoothed: pd.DataFrame) -> pd.DataFrame:
    """Stage-2 demand for off-plan customers — those in ALL_CUSTOMERS but not
    PLAN_CUSTOMERS for the current REGION. Their Stage-2 demand equals their
    smoothed monthly history (since they have no plan to defend in Stage 1).

    Returns rows per (Customer, KI, MonthNum) with Stage2Demand column. Rows
    with Stage2Demand <= 0 are dropped.
    """
    offplan = sorted(set(ALL_CUSTOMERS) - set(PLAN_CUSTOMERS))
    if not offplan:
        return pd.DataFrame(columns=["Customer", "KI", "MonthNum", "Stage2Demand"])
    op = smoothed[smoothed["Customer"].isin(offplan)].copy()
    if op.empty:
        return pd.DataFrame(columns=["Customer", "KI", "MonthNum", "Stage2Demand"])
    # LiftGroup is "KI:<ki>" for non-HD/Lowes customers (off-plan never have HD/Lowes
    # SKU mapping, since they're not HD or Lowes themselves).
    mask = op["LiftGroup"].astype(str).str.startswith("KI:")
    op = op[mask].copy()
    op["KI"] = op["LiftGroup"].astype(str).str.replace("^KI:", "", regex=True)
    op = op.rename(columns={"Month": "MonthNum"})
    op = op[op["MonthNum"].isin(FWD_MONTHS)].copy()
    op = op[["Customer", "KI", "MonthNum", "smoothed"]].rename(
        columns={"smoothed": "Stage2Demand"})
    op = op[op["Stage2Demand"] > 0].copy()
    log(f"  off-plan Stage-2 demand rows: {len(op):,}  customers: {sorted(op['Customer'].unique().tolist()) if not op.empty else []}")
    return op


def apply_lift(plan: pd.DataFrame, smoothed: pd.DataFrame, lift_groups: pd.DataFrame) -> pd.DataFrame:
    """Apply historical lift to plan items using per-customer LiftGroup mapping.

    For each (Customer, LiftGroup, Month):
      - grouping_plan = sum plan QTY for items in this LiftGroup × customer × month
      - if smoothed > grouping_plan: scale items by (smoothed/grouping_plan)
      - else: lifted = plan
    Only operates on FWD_MONTHS (May..Dec).

    Plan items not present in lift_groups (e.g. HD/Lowes items with no SKU
    AND no history AND not in plan — but note: plan items are by definition
    in plan, so this only excludes plan items that lift_groups deliberately
    drops, i.e. items missing both KI grouping and SKU mapping). Such items
    receive LiftFactor=1.0 (no lift, but still in the plan).
    """
    log("Applying historical lift (per-customer LiftGroup)...")
    p = plan.copy()
    p = p.merge(lift_groups, on=["Customer", "Item Num"], how="left")

    fwd = p[p["MonthNum"].isin(FWD_MONTHS)].copy()
    jad = p[~p["MonthNum"].isin(FWD_MONTHS)].copy()
    jad["LiftedQty"] = jad["PlanQty"]

    # Items that lack a LiftGroup (no KI, no SKU) → no lift
    has_lg = fwd["LiftGroup"].notna()
    fwd_lift   = fwd[has_lg].copy()
    fwd_nolift = fwd[~has_lg].copy()
    fwd_nolift["LiftedQty"] = fwd_nolift["PlanQty"]

    grp = (fwd_lift.groupby(["Customer", "LiftGroup", "MonthNum"], dropna=False)["PlanQty"]
                .sum().reset_index().rename(columns={"PlanQty": "GroupingPlan"}))
    sm = smoothed.rename(columns={"Month": "MonthNum"})[["Customer", "LiftGroup", "MonthNum", "smoothed"]]
    grp = grp.merge(sm, on=["Customer", "LiftGroup", "MonthNum"], how="left")
    grp["smoothed"] = grp["smoothed"].fillna(0.0)
    grp["LiftFactor"] = np.where(
        (grp["smoothed"] > grp["GroupingPlan"]) & (grp["GroupingPlan"] > 0),
        grp["smoothed"] / grp["GroupingPlan"],
        1.0,
    )
    fwd_lift = fwd_lift.merge(grp[["Customer", "LiftGroup", "MonthNum", "LiftFactor"]],
                              on=["Customer", "LiftGroup", "MonthNum"], how="left")
    fwd_lift["LiftFactor"] = fwd_lift["LiftFactor"].fillna(1.0)
    fwd_lift["LiftedQty"] = fwd_lift["PlanQty"] * fwd_lift["LiftFactor"]

    out = pd.concat([fwd_lift, fwd_nolift, jad], ignore_index=True, sort=False)
    out["Plan_Source"] = "Original"
    return out


def apply_history_synthesis(plan_lifted: pd.DataFrame, smoothed: pd.DataFrame,
                            items: pd.DataFrame, hist: pd.DataFrame,
                            lift_groups: pd.DataFrame) -> pd.DataFrame:
    """Phase C / Option B: synthesize plan rows from smoothed history where missing.

    For each (Customer, LiftGroup, MonthNum) in May-Dec where smoothed > 0 but no
    plan row exists for that combo, synthesize a row with PlanQty = LiftedQty = smoothed.
    Tag synthesized rows with Plan_Source = "Synthesized".

    HD/Lowes use SKU-level LiftGroup ("SKU:<item>"); others use KI-level ("KI:<ki>").
    For KI-level synthesis, the synthesized row is attributed to the customer's
    most-historically-shipped item in that KI.
    Off-plan customers (Southeast/MLC/Other) skipped — they go through offplan_demand path.
    """
    log("Applying history synthesis (Option B — fill plan gaps where history exists)...")

    # Map item to KI and Org Code
    item_to_ki = items[["Item Num", "KI"]].drop_duplicates(subset=["Item Num"]).set_index("Item Num")["KI"].to_dict()

    # Get Org Code per item from plan_lifted (each item is tagged with an Org Code)
    item_to_org = plan_lifted.dropna(subset=["Org Code"]).drop_duplicates(subset=["Item Num"]).set_index("Item Num")["Org Code"].to_dict()

    # Precompute (Customer, KI) -> representative_item dict (vectorized)
    hist_with_ki = hist.copy()
    hist_with_ki["KI"] = hist_with_ki["Item Num"].map(item_to_ki)
    hist_with_ki = hist_with_ki.dropna(subset=["KI"])
    rep_item_lookup = {}
    if not hist_with_ki.empty:
        agg = hist_with_ki.groupby(["Customer", "KI", "Item Num"], as_index=False)["Qty"].sum()
        agg_max = agg.loc[agg.groupby(["Customer", "KI"])["Qty"].idxmax()]
        for _, r in agg_max.iterrows():
            rep_item_lookup[(r["Customer"], r["KI"])] = r["Item Num"]

    # Build (Customer, LiftGroup) -> primary Item Num lookup for HD/Lowes SKU-level synthesis
    # (LiftGroup like "HDSKU:1000038388" is HD's SKU code, not internal Item Num — need reverse lookup)
    # Multiple Item Nums can map to the same SKU. Pick the one with most historical Qty
    # under that customer; fall back to most plan-ed item; else first.
    cust_lg_to_item = {}
    # Build (Customer, Item) -> Qty from history
    cust_item_hist_qty = hist.groupby(["Customer", "Item Num"])["Qty"].sum().to_dict()
    # Build (Customer, Item) -> Qty from plan
    plan_for_lookup = plan_lifted[plan_lifted["Plan_Source"] == "Original"]
    cust_item_plan_qty = plan_for_lookup.groupby(["Customer", "Item Num"])["PlanQty"].sum().to_dict()
    for cust, lg_group in lift_groups.groupby(["Customer", "LiftGroup"]):
        cust_name, lg_name = cust
        candidates = lg_group["Item Num"].tolist()
        # Score candidates: prefer items with plan first, then by historical Qty
        def score(item):
            plan_qty = cust_item_plan_qty.get((cust_name, item), 0)
            hist_qty = cust_item_hist_qty.get((cust_name, item), 0)
            # Use plan_qty as tiebreaker bonus, hist_qty as primary
            return (plan_qty + hist_qty * 0.001, hist_qty)
        best = max(candidates, key=score) if candidates else None
        cust_lg_to_item[(cust_name, lg_name)] = best

    # Build set of existing (Customer, LiftGroup, MonthNum) keys (vectorized)
    pl_fwd = plan_lifted[plan_lifted["MonthNum"].isin(FWD_MONTHS) & plan_lifted["LiftGroup"].notna()]
    existing = set(zip(pl_fwd["Customer"].tolist(), pl_fwd["LiftGroup"].astype(str).tolist(), pl_fwd["MonthNum"].astype(int).tolist()))

    # Iterate smoothed and synthesize where needed
    new_rows = []
    n_skipped_offplan = 0
    n_skipped_no_hist = 0
    n_skipped_existing = 0
    n_synthesized_sku = 0
    n_synthesized_ki = 0

    HD_LOWES = {"HD", "Lowes"}

    # Pre-filter smoothed to plan customers + forward months + smoothed > 0
    sm_filtered = smoothed[
        smoothed["Customer"].isin(PLAN_CUSTOMERS)
        & (smoothed["smoothed"] > 0)
        & smoothed["Month"].isin(FWD_MONTHS)
    ].copy()
    n_skipped_offplan = (~smoothed["Customer"].isin(PLAN_CUSTOMERS)).sum()

    for _, sm in sm_filtered.iterrows():
        cust = sm["Customer"]
        smoothed_qty = float(sm["smoothed"])
        month = int(sm["Month"])
        lg = str(sm["LiftGroup"])

        if (cust, lg, month) in existing:
            n_skipped_existing += 1
            continue

        # HD/Lowes can use HDSKU:/LOWESSKU: (SKU-level) or KI: (fallback)
        if cust in HD_LOWES and (lg.startswith("HDSKU:") or lg.startswith("LOWESSKU:")):
            # SKU-level: look up Item Num from (Customer, LiftGroup)
            item_num = cust_lg_to_item.get((cust, lg))
            if item_num is None:
                n_skipped_no_hist += 1
                continue
            ki = item_to_ki.get(item_num)
            if not ki:
                continue
            org = item_to_org.get(item_num, "")
            new_rows.append({
                "Org Code": org, "Region": "NOR CAL",
                "Item Num": item_num, "Customer": cust, "MonthNum": month,
                "PlanQty": smoothed_qty, "KI": ki,
                "LiftGroup": lg, "LiftFactor": 1.0,
                "LiftedQty": smoothed_qty, "Plan_Source": "Synthesized",
            })
            n_synthesized_sku += 1
        elif lg.startswith("KI:"):
            # KI-level (all non-HD/Lowes customers + HD/Lowes fallback when no SKU mapping)
            ki = lg.replace("KI:", "")
            rep_item = rep_item_lookup.get((cust, ki))
            if rep_item is None:
                n_skipped_no_hist += 1
                continue
            org = item_to_org.get(rep_item, "")
            new_rows.append({
                "Org Code": org, "Region": "NOR CAL",
                "Item Num": rep_item, "Customer": cust, "MonthNum": month,
                "PlanQty": smoothed_qty, "KI": ki,
                "LiftGroup": lg, "LiftFactor": 1.0,
                "LiftedQty": smoothed_qty, "Plan_Source": "Synthesized",
            })
            n_synthesized_ki += 1

    if new_rows:
        new_df = pd.DataFrame(new_rows)
        result = pd.concat([plan_lifted, new_df], ignore_index=True, sort=False)
    else:
        result = plan_lifted.copy()

    log(f"  synthesized: {len(new_rows):,} rows ({n_synthesized_sku} SKU-level HD/Lowes, {n_synthesized_ki} KI-level others)")
    log(f"  skipped: {n_skipped_offplan} off-plan, {n_skipped_existing} existing rows, {n_skipped_no_hist} no representative item")
    log(f"  total synthesized QTY: {sum(r['PlanQty'] for r in new_rows):,.0f}")

    return result


# ---------------------------------------------------------------------------
# Sequential month walk (per KI Grouping)
# ---------------------------------------------------------------------------

def run_walk(plan_lifted: pd.DataFrame, pools: pd.DataFrame,
             offplan_demand: pd.DataFrame = None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Two-stage sequential pool walk per (KI, Month).

    STAGE 1 — DEFEND ORIGINAL PLAN (May → Dec):
      For each KI, sequentially walk months May..Dec. Each month:
        * Add inflows (graduations of SS/GS at Ready Date; new SN/GN ready)
        * NBB consumes restricted pool first; pro-rata across NBB customers if short
        * Then NBB-remaining + BB plan demand both compete in the all-pool
          pro-rata if the pool is short (no BB-vs-NBB priority within all-pool)
        * Track plan_filled / plan_short per customer
        * Record end_pool_S1[M] (split into all-pool and restricted-pool)

    STAGE 2 — APPLY LIFT FROM SURPLUS (May → Dec, layered on Stage-1 end pools):
      For each KI, walk months May..Dec sequentially again. Each month M:
        * Compute lift demand per customer:
            - Planned customers: max(0, LiftedQty - PlanQty) for this KI×M
            - Off-plan customers: their smoothed history at this KI×M
        * Compute max_stage2[M] = min(end_pool_S1[m] for m ∈ [M, Dec]) - cum_stage2_prior
          — protects later months' end-pool from going below zero
        * Apply same allocation logic as Stage 1 to the constrained pool budget:
            - NBB lift consumes any remaining restricted pool first (pro-rata)
            - Combined NBB-remaining + BB lift competes in all-pool pro-rata
        * Update running pool levels for all m ≥ M

    Returns:
      pool_walk: per (KI, Month) row with Stage 1 + Stage 2 detail and
                 final end-of-month pool levels
      cust_alloc: per (Customer, KI, Item Num, Month) row with PlanQty,
                  LiftedQty, PlanFilled, PlanShort, LiftFilled, LiftShort
      offplan_alloc: per (Customer, KI, Month) row for off-plan customers
                     with Stage2Demand, Stage2Filled, Stage2Short
    """
    log("Running two-stage pool walk (Stage 1: defend plan; Stage 2: apply lift)...")

    fwd_plan = plan_lifted[plan_lifted["MonthNum"].isin(FWD_MONTHS)].copy()
    fwd_plan = fwd_plan[fwd_plan["KI"].notna()].copy()

    if offplan_demand is None or offplan_demand.empty:
        offplan_demand = pd.DataFrame(columns=["Customer", "KI", "MonthNum", "Stage2Demand"])

    pool_walk_rows = []
    cust_alloc_rows = []
    offplan_alloc_rows = []

    bb_set = set(BB_CUSTOMERS)
    nbb_set = set(NBB_CUSTOMERS)

    # Process each KI independently. Within a KI, items pool together (substitution allowed).
    all_kis = sorted(set(fwd_plan["KI"].unique()) | set(offplan_demand["KI"].unique()))

    for ki in all_kis:
        ki_pools = pools[pools["KI"] == ki] if not pools.empty else pd.DataFrame()

        # Initial pools at start of May
        if not ki_pools.empty:
            initial_all = float(ki_pools.loc[ki_pools["PoolCat"] == "all_initial", "PoolQty"].sum())
            inflow_sngn = (
                ki_pools[ki_pools["PoolCat"] == "sngn_inflow_at_ready"]
                .groupby("Ready Month")["PoolQty"].sum()
            )
        else:
            initial_all = 0.0
            inflow_sngn = pd.Series(dtype=float)

        # ===== PHASE B FIX: per-lot tracking for restricted pool =====
        # Replace scalar s1_pool_restricted with restricted_lots list.
        # Each lot tracks remaining_avail (what's still in restricted) and consumed_pre_grad
        # (NBB consumption taken from this lot before it graduated). At graduation, all-pool
        # gains (total - consumed_pre_grad) — the actual physical residual — instead of full
        # total. This eliminates phantom inflow that double-counts consumed restricted units.
        restricted_lots = []
        if not ki_pools.empty:
            ss_gs_lots = ki_pools[ki_pools["PoolCat"] == "restricted_then_all"]
            for _, row in ss_gs_lots.iterrows():
                restricted_lots.append({
                    "item_num":         row["Item Num"],
                    "ready_month":      int(row["Ready Month"]),
                    "total_init":       float(row["Total QTY"]),
                    "avail_init":       float(row["Available QTY"]),
                    "remaining_avail":  float(row["Available QTY"]),
                    "consumed_pre_grad": 0.0,
                    "graduated":        False,
                })

        def _restricted_balance():
            return sum(L["remaining_avail"] for L in restricted_lots if not L["graduated"])

        def _consume_restricted_fifo(amount):
            """Consume `amount` from active restricted lots, FIFO by ready_month.
            Returns actually consumed (could be less than amount if pool exhausted)."""
            if amount <= 0:
                return 0.0
            taken = 0.0
            need = amount
            active = sorted(
                [L for L in restricted_lots if not L["graduated"] and L["remaining_avail"] > 0],
                key=lambda L: L["ready_month"]
            )
            for L in active:
                if need <= 0:
                    break
                avail = L["remaining_avail"]
                t = min(need, avail)
                L["remaining_avail"] -= t
                L["consumed_pre_grad"] += t
                taken += t
                need -= t
            return taken

        ki_plan = fwd_plan[fwd_plan["KI"] == ki].copy()
        ki_offplan = offplan_demand[offplan_demand["KI"] == ki].copy()

        # =================================================================
        # PASS 1 — Stage 1 plan defense, sequential May..Dec
        # =================================================================
        s1_pool_all = float(initial_all)

        # Per-month state arrays
        s1_state = {m: {} for m in FWD_MONTHS}

        # Per-customer plan_filled across all months (back-allocated to items below)
        plan_filled_by_cm = {}  # (cust, m) -> filled qty across the KI

        for m in FWD_MONTHS:
            new_sngn = float(inflow_sngn.get(m, 0.0)) if not inflow_sngn.empty else 0.0

            # Process graduations: lots with ready_month == m transition to all-pool
            grad_total_this_month = 0.0  # for reporting (whole physical lot total)
            grad_leave_restricted_this_month = 0.0  # for reporting (units that left restricted at grad)
            for L in restricted_lots:
                if not L["graduated"] and L["ready_month"] == m:
                    # All-pool gains the actual physical remainder (total - already-consumed)
                    actual_residual = L["total_init"] - L["consumed_pre_grad"]
                    s1_pool_all += actual_residual
                    # Restricted loses the lot's still-available portion
                    grad_leave_restricted_this_month += L["remaining_avail"]
                    grad_total_this_month += L["total_init"]
                    L["remaining_avail"] = 0.0
                    L["graduated"] = True

            # Add SN/GN inflow at-ready
            s1_pool_all += new_sngn

            # Plan demand by customer for this KI × month
            mplan = ki_plan[ki_plan["MonthNum"] == m]
            plan_by_cust = mplan.groupby("Customer")["PlanQty"].sum().to_dict()

            nbb_plan = {c: q for c, q in plan_by_cust.items() if c in nbb_set and q > 0}
            bb_plan  = {c: q for c, q in plan_by_cust.items() if c in bb_set and q > 0}

            nbb_total = sum(nbb_plan.values())
            bb_total  = sum(bb_plan.values())

            # Step 1: NBB pulls restricted pool first (FIFO by ready_month across active lots)
            nbb_filled_r = {c: 0.0 for c in nbb_plan}
            current_restricted = _restricted_balance()
            r_consumed_target = min(current_restricted, nbb_total)
            if nbb_total > 0 and r_consumed_target > 0:
                # Allocate per-customer pro-rata
                for c, q in nbb_plan.items():
                    share = q / nbb_total
                    nbb_filled_r[c] = share * r_consumed_target
                # Consume from FIFO lots
                _consume_restricted_fifo(r_consumed_target)
            r_consumed = r_consumed_target  # for compatibility with downstream
            nbb_remaining = {c: max(0.0, nbb_plan[c] - nbb_filled_r[c]) for c in nbb_plan}

            # Step 2: NBB-remaining + BB compete pro-rata in all-pool
            combined_demand = sum(nbb_remaining.values()) + bb_total
            allp_consumed = min(s1_pool_all, combined_demand)
            allp_factor = (allp_consumed / combined_demand) if combined_demand > 0 else 0.0
            s1_pool_all -= allp_consumed

            nbb_filled_a = {c: nbb_remaining[c] * allp_factor for c in nbb_plan}
            bb_filled    = {c: bb_plan[c] * allp_factor       for c in bb_plan}

            # Aggregate per-customer plan-filled for Stage 1
            for c in nbb_plan:
                plan_filled_by_cm[(c, m)] = nbb_filled_r[c] + nbb_filled_a[c]
            for c in bb_plan:
                plan_filled_by_cm[(c, m)] = bb_filled[c]

            s1_state[m] = {
                "InflowGrad": grad_total_this_month,  # full physical lot total (for legacy reporting)
                "InflowSNGN": new_sngn,
                "Stage1_NBB_Demand": nbb_total,
                "Stage1_BB_Demand":  bb_total,
                "Stage1_NBB_Filled": sum(nbb_filled_r.values()) + sum(nbb_filled_a.values()),
                "Stage1_BB_Filled":  sum(bb_filled.values()),
                "Stage1_End_All":        s1_pool_all,
                "Stage1_End_Restricted": _restricted_balance(),
            }
            s1_state[m]["Stage1_NBB_Short"] = max(0.0, nbb_total - s1_state[m]["Stage1_NBB_Filled"])
            s1_state[m]["Stage1_BB_Short"]  = max(0.0, bb_total  - s1_state[m]["Stage1_BB_Filled"])

        # =================================================================
        # PASS 2 — Stage 2 lift surplus, sequential May..Dec on running pools
        # =================================================================
        # Combined-pool view (all + restricted) from Stage 1 end states
        end_pool_S1_combined = {m: s1_state[m]["Stage1_End_All"] + s1_state[m]["Stage1_End_Restricted"]
                                for m in FWD_MONTHS}

        # Running pool levels (start = Stage 1 end). Stage 2 will reduce these.
        running_all = {m: s1_state[m]["Stage1_End_All"]        for m in FWD_MONTHS}
        running_res = {m: s1_state[m]["Stage1_End_Restricted"] for m in FWD_MONTHS}

        cum_stage2_total = 0.0  # cumulative Stage 2 consumption for this KI through end of prior month

        # Stage 2 fulfilled by (cust, m) for back-allocation
        lift_filled_by_cm = {}  # planned customers
        offplan_filled_by_cm = {}  # off-plan customers

        for m in FWD_MONTHS:
            # ----- Compute lift demand per customer -----
            mplan = ki_plan[ki_plan["MonthNum"] == m]
            # Per planned customer, lift demand at KI level = max(0, sum(LiftedQty) - sum(PlanQty))
            cust_plan = mplan.groupby("Customer")[["PlanQty", "LiftedQty"]].sum()
            lift_dem_planned = {}
            for c, row in cust_plan.iterrows():
                lift = max(0.0, float(row["LiftedQty"]) - float(row["PlanQty"]))
                if lift > 0:
                    lift_dem_planned[c] = lift

            # Per off-plan customer, Stage-2 demand = smoothed history at KI×M
            mop = ki_offplan[ki_offplan["MonthNum"] == m]
            lift_dem_offplan = {row["Customer"]: float(row["Stage2Demand"]) for _, row in mop.iterrows() if row["Stage2Demand"] > 0}

            # Combine
            lift_dem = {**lift_dem_planned, **lift_dem_offplan}

            nbb_lift = {c: q for c, q in lift_dem.items() if c in nbb_set and q > 0}
            bb_lift  = {c: q for c, q in lift_dem.items() if c in bb_set and q > 0}
            nbb_total = sum(nbb_lift.values())
            bb_total  = sum(bb_lift.values())
            total_demand = nbb_total + bb_total

            # ----- Compute max_stage2[M] from min running combined pool over [M, Dec] -----
            min_combined = min(running_all[mm] + running_res[mm] for mm in FWD_MONTHS if mm >= m)
            # cum_stage2_total has already been subtracted from running pools, so min_combined
            # is already the headroom available in any future month.
            max_stage2 = max(0.0, min_combined)

            # If demand exceeds headroom, scale uniformly across NBB + BB
            scale_demand = min(total_demand, max_stage2)
            if total_demand > 0:
                scale = scale_demand / total_demand
            else:
                scale = 0.0
            # Scaled demands (this is what Stage 2 will TRY to allocate; sub-allocation by
            # restricted pool eligibility happens next).
            nbb_lift_eff = {c: q * scale for c, q in nbb_lift.items()}
            bb_lift_eff  = {c: q * scale for c, q in bb_lift.items()}
            nbb_total_eff = sum(nbb_lift_eff.values())
            bb_total_eff  = sum(bb_lift_eff.values())

            # Within-month allocation (analogous to Stage 1):
            # Step 1: NBB lift consumes restricted pool first
            r_avail = running_res[m]
            r_consumed = min(r_avail, nbb_total_eff)
            nbb_filled_r = {c: 0.0 for c in nbb_lift_eff}
            if nbb_total_eff > 0 and r_consumed > 0:
                for c, q in nbb_lift_eff.items():
                    share = q / nbb_total_eff
                    nbb_filled_r[c] = share * r_consumed
            nbb_remaining_eff = {c: max(0.0, nbb_lift_eff[c] - nbb_filled_r[c]) for c in nbb_lift_eff}

            # Step 2: NBB-remaining + BB compete pro-rata in all-pool
            combined_remaining = sum(nbb_remaining_eff.values()) + bb_total_eff
            a_avail = running_all[m]
            a_consumed = min(a_avail, combined_remaining)
            a_factor = (a_consumed / combined_remaining) if combined_remaining > 0 else 0.0

            nbb_filled_a = {c: nbb_remaining_eff[c] * a_factor for c in nbb_lift_eff}
            bb_filled    = {c: bb_lift_eff[c] * a_factor       for c in bb_lift_eff}

            # Total Stage 2 consumption this month (all + restricted)
            stage2_consumed = r_consumed + a_consumed

            # Update running pools for THIS month and ALL future months
            for mm in FWD_MONTHS:
                if mm >= m:
                    # Restricted pool: only "this month" loses r_consumed (that consumption
                    # affects this month's restricted balance going forward, since we model
                    # restricted as a pool that can be drawn down whenever there's demand).
                    # In future months, the running_res value already reflects what would
                    # have been there at end of m PLUS any future graduations leaving it.
                    # Simplification: subtract r_consumed from every future running_res too,
                    # because graduations happening between m and mm could reduce restricted
                    # by less than r_consumed. We protect against this via the min_combined
                    # constraint already (which uses running_all + running_res).
                    pass
            # Apply the consumption: simplest correct view — total combined pool drops by
            # stage2_consumed at end of m and every later month.
            for mm in FWD_MONTHS:
                if mm >= m:
                    # Apply r_consumed to restricted side (capped at current value to avoid negative)
                    take_r = min(r_consumed, running_res[mm])
                    running_res[mm] -= take_r
                    # Anything that couldn't come from restricted (because future month already had
                    # less restricted than current) goes off the all-pool side
                    leftover = r_consumed - take_r
                    running_all[mm] = max(0.0, running_all[mm] - (a_consumed + leftover))
            cum_stage2_total += stage2_consumed

            # Aggregate per-customer Stage 2 fulfilled
            for c in nbb_lift_eff:
                f = nbb_filled_r[c] + nbb_filled_a[c]
                if c in PLAN_CUSTOMERS:
                    lift_filled_by_cm[(c, m)] = lift_filled_by_cm.get((c, m), 0.0) + f
                else:
                    offplan_filled_by_cm[(c, m)] = offplan_filled_by_cm.get((c, m), 0.0) + f
            for c in bb_lift_eff:
                f = bb_filled[c]
                if c in PLAN_CUSTOMERS:
                    lift_filled_by_cm[(c, m)] = lift_filled_by_cm.get((c, m), 0.0) + f
                else:
                    offplan_filled_by_cm[(c, m)] = offplan_filled_by_cm.get((c, m), 0.0) + f

            # Per-month pool-walk row (combine Stage 1 + Stage 2 state)
            row = dict(s1_state[m])
            row.update({
                "KI": ki, "Month": m,
                "Stage2_Demand_Total": total_demand,
                "Stage2_Filled":       stage2_consumed,
                "Stage2_Short":        max(0.0, total_demand - stage2_consumed),
                "End_All":             running_all[m],
                "End_Restricted":      running_res[m],
                "End_Combined":        running_all[m] + running_res[m],
            })
            pool_walk_rows.append(row)

        # =================================================================
        # Back-allocate Stage 1 + Stage 2 to (Customer, Item Num, Month)
        # =================================================================
        # For each planned-customer × month, split filled QTY across items.
        for m in FWD_MONTHS:
            mplan = ki_plan[ki_plan["MonthNum"] == m]
            for c in mplan["Customer"].unique():
                sub = mplan[mplan["Customer"] == c]
                cust_plan_total = float(sub["PlanQty"].sum())
                cust_lift_total = float((sub["LiftedQty"] - sub["PlanQty"]).clip(lower=0).sum())
                pfilled = plan_filled_by_cm.get((c, m), 0.0)
                lfilled = lift_filled_by_cm.get((c, m), 0.0)
                for _, r in sub.iterrows():
                    item_plan = float(r["PlanQty"])
                    item_lift_dem = max(0.0, float(r["LiftedQty"]) - item_plan)
                    p_share = (item_plan / cust_plan_total) if cust_plan_total > 0 else 0.0
                    l_share = (item_lift_dem / cust_lift_total) if cust_lift_total > 0 else 0.0
                    item_pfilled = pfilled * p_share
                    item_lfilled = lfilled * l_share
                    cust_alloc_rows.append({
                        "Customer": c, "KI": ki, "Item Num": r["Item Num"],
                        "MonthNum": m,
                        "PlanQty":     item_plan,
                        "LiftedQty":   float(r["LiftedQty"]),
                        "PlanFilled":  item_pfilled,
                        "PlanShort":   max(0.0, item_plan - item_pfilled),
                        "LiftFilled":  item_lfilled,
                        "LiftShort":   max(0.0, item_lift_dem - item_lfilled),
                    })

        # Off-plan customers — track at (Customer, KI, Month) level
        for m in FWD_MONTHS:
            mop = ki_offplan[ki_offplan["MonthNum"] == m]
            for _, r in mop.iterrows():
                c = r["Customer"]
                d = float(r["Stage2Demand"])
                f = offplan_filled_by_cm.get((c, m), 0.0)
                offplan_alloc_rows.append({
                    "Customer": c, "KI": ki, "MonthNum": m,
                    "Stage2Demand": d,
                    "Stage2Filled": f,
                    "Stage2Short":  max(0.0, d - f),
                })

    pool_walk_df = pd.DataFrame(pool_walk_rows)
    cust_alloc_df = pd.DataFrame(cust_alloc_rows)
    offplan_alloc_df = pd.DataFrame(offplan_alloc_rows)

    log(f"  pool walk rows: {len(pool_walk_df):,}  cust alloc rows: {len(cust_alloc_df):,}  "
        f"offplan alloc rows: {len(offplan_alloc_df):,}")
    return pool_walk_df, cust_alloc_df, offplan_alloc_df



# ---------------------------------------------------------------------------
# YE Miss
# ---------------------------------------------------------------------------

def compute_ye_miss(plan: pd.DataFrame, plan_lifted: pd.DataFrame, ytd: pd.DataFrame,
                    cust_alloc: pd.DataFrame, items: pd.DataFrame,
                    prices: pd.DataFrame) -> pd.DataFrame:
    """Per (Customer, Item Num) — updated for two-stage allocation:
       Original YE Plan       = sum(plan, all months)
       Lifted YE Plan         = sum(plan, Jan-Apr) + sum(lifted, May-Dec)
       YTD Actual             = sum(2026 actuals, Jan-Apr)
       Plan Fulfilled (S1)    = sum(PlanFilled, May-Dec)        — Stage 1 result
       Lift Fulfilled (S2)    = sum(LiftFilled, May-Dec)        — Stage 2 result
       Forward Fulfillable    = Plan Fulfilled + Lift Fulfilled — total fwd shippable
       YE Miss QTY            = max(0, Original YE Plan - YTD - Forward Fulfillable)
                                — measured against ORIGINAL plan per CEO rule
       YE Miss $              = QTY × cascade price
       Customer Over-Plan QTY = max(0, YTD + Forward - Original YE Plan)
                                — when fulfillment exceeds plan due to lift
       Customer Over-Plan $   = QTY × cascade price
    """
    log("Computing YE Miss (two-stage)...")
    orig = plan.groupby(["Customer", "Item Num"])["PlanQty"].sum().reset_index().rename(
        columns={"PlanQty": "Original YE Plan"})
    lifted = plan_lifted.copy()
    lifted["Q"] = np.where(lifted["MonthNum"].isin(FWD_MONTHS), lifted["LiftedQty"], lifted["PlanQty"])
    lift_ye = lifted.groupby(["Customer", "Item Num"])["Q"].sum().reset_index().rename(
        columns={"Q": "Lifted YE Plan"})
    ytd_g = ytd.groupby(["Customer", "Item Num"])["Qty"].sum().reset_index().rename(
        columns={"Qty": "YTD Actual"})

    # Two-stage fulfilled aggregations
    if cust_alloc.empty:
        plan_f = pd.DataFrame(columns=["Customer", "Item Num", "Plan Fulfilled (S1)"])
        lift_f = pd.DataFrame(columns=["Customer", "Item Num", "Lift Fulfilled (S2)"])
    else:
        plan_f = cust_alloc.groupby(["Customer", "Item Num"])["PlanFilled"].sum().reset_index().rename(
            columns={"PlanFilled": "Plan Fulfilled (S1)"})
        lift_f = cust_alloc.groupby(["Customer", "Item Num"])["LiftFilled"].sum().reset_index().rename(
            columns={"LiftFilled": "Lift Fulfilled (S2)"})

    yem = orig.merge(lift_ye, on=["Customer", "Item Num"], how="outer") \
              .merge(ytd_g,   on=["Customer", "Item Num"], how="outer") \
              .merge(plan_f,  on=["Customer", "Item Num"], how="outer") \
              .merge(lift_f,  on=["Customer", "Item Num"], how="outer").fillna(0.0)
    yem["Forward Fulfillable"] = yem["Plan Fulfilled (S1)"] + yem["Lift Fulfilled (S2)"]
    # Plan miss measured against ORIGINAL plan (not lifted) per CEO rule.
    yem["YE Miss QTY"]         = (yem["Original YE Plan"] - yem["YTD Actual"] - yem["Forward Fulfillable"]).clip(lower=0)
    yem["Customer Over-Plan QTY"] = (yem["YTD Actual"] + yem["Forward Fulfillable"] - yem["Original YE Plan"]).clip(lower=0)

    # Pricing — per (Customer, Item) from actuals-based cascade
    yem = yem.merge(prices[["Customer", "Item Num", "Price"]],
                     on=["Customer", "Item Num"], how="left")
    yem["Price"] = yem["Price"].fillna(0.0)
    yem["YE Miss $"]         = yem["YE Miss QTY"] * yem["Price"]
    yem["Customer Over-Plan $"] = yem["Customer Over-Plan QTY"] * yem["Price"]
    return yem


def build_price_cascade(items: pd.DataFrame, hist: pd.DataFrame, ytd: pd.DataFrame,
                          lift_groups: pd.DataFrame) -> pd.DataFrame:
    """Per-(Customer, Item) plan-aligned price cascade, built entirely from
    actuals (per the user's spec — V158 prices not used).

    Tier 1: 2026 YTD avg sell price by (Customer, Item)        — what we're charging now
    Tier 2: 2025 avg sell price by (Customer, Item)             — last year's price
    Tier 3: Customer × LiftGroup avg from 2025+2026 actuals     — group consistency
    Tier 4: LiftGroup avg across customers from 2025+2026       — broader fallback
    Tier 5: Genus+Size avg from 2025+2026 actuals               — last resort
    Tier 6: Size avg from 2025+2026 actuals                      — final catch
    """
    log("Building price cascade (actuals-based)...")

    # Build avg sell price helper: weighted = Revenue/Qty
    def avg_price(df, group_cols):
        g = df.groupby(group_cols).agg(_q=("Qty", "sum"), _r=("Revenue", "sum")).reset_index()
        g["price"] = np.where(g["_q"] > 0, g["_r"] / g["_q"], np.nan)
        return g[group_cols + ["price"]]

    hist_25 = hist[hist["Year"] == 2025].copy()
    # 2026 YTD actuals (Jan-Apr) — passed in as ytd

    # Tier 1: per (Customer, Item Num) from 2026 YTD
    t1 = avg_price(ytd, ["Customer", "Item Num"]).rename(columns={"price": "T1"})

    # Tier 2: per (Customer, Item Num) from 2025
    t2 = avg_price(hist_25, ["Customer", "Item Num"]).rename(columns={"price": "T2"})

    # Combined 2025+2026 actuals for tier 3-6
    h_combined = pd.concat([hist_25[["Customer", "Item Num", "Qty", "Revenue"]],
                              ytd   [["Customer", "Item Num", "Qty", "Revenue"]]], ignore_index=True)
    h_combined = h_combined.merge(items[["Item Num", "Genus", "Size"]].drop_duplicates(subset=["Item Num"]),
                                    on="Item Num", how="left")
    h_combined = h_combined.merge(lift_groups[["Customer", "Item Num", "LiftGroup"]],
                                    on=["Customer", "Item Num"], how="left")

    # Tier 3: per (Customer, LiftGroup)
    t3 = avg_price(h_combined.dropna(subset=["LiftGroup"]),
                    ["Customer", "LiftGroup"]).rename(columns={"price": "T3"})

    # Tier 4: per (LiftGroup) across customers
    t4 = avg_price(h_combined.dropna(subset=["LiftGroup"]),
                    ["LiftGroup"]).rename(columns={"price": "T4"})

    # Tier 5: per (Genus, Size)
    t5 = avg_price(h_combined.dropna(subset=["Genus", "Size"]),
                    ["Genus", "Size"]).rename(columns={"price": "T5"})

    # Tier 6: per Size only
    t6 = avg_price(h_combined.dropna(subset=["Size"]),
                    ["Size"]).rename(columns={"price": "T6"})

    # Build per (Customer, Item Num) lookup
    item_meta = items[["Item Num", "Genus", "Size", "KI"]].drop_duplicates(subset=["Item Num"])
    base = lift_groups.merge(item_meta, on="Item Num", how="left")
    base = base.merge(t1, on=["Customer", "Item Num"], how="left")
    base = base.merge(t2, on=["Customer", "Item Num"], how="left")
    base = base.merge(t3, on=["Customer", "LiftGroup"], how="left")
    base = base.merge(t4, on="LiftGroup", how="left")
    base = base.merge(t5, on=["Genus", "Size"], how="left")
    base = base.merge(t6, on="Size", how="left")

    # Tag tier used
    def pick_tier(row):
        for i, c in enumerate(["T1", "T2", "T3", "T4", "T5", "T6"], start=1):
            if pd.notna(row[c]) and row[c] > 0:
                return i, row[c]
        return 0, 0.0
    tiers = base.apply(pick_tier, axis=1, result_type="expand")
    base["Tier"] = tiers[0]
    base["Price"] = tiers[1]

    log(f"  price cascade tiers used: {base['Tier'].value_counts().to_dict()}")
    return base[["Customer", "Item Num", "Price", "Tier"]].drop_duplicates(subset=["Customer", "Item Num"])


# ---------------------------------------------------------------------------
# Output writer (Summary + BUXUS Detail)
# ---------------------------------------------------------------------------

def write_output(out_path: Path, plan: pd.DataFrame, plan_lifted: pd.DataFrame,
                 ytd: pd.DataFrame, cust_alloc: pd.DataFrame,
                 pool_walk: pd.DataFrame, ye_miss: pd.DataFrame, items: pd.DataFrame) -> None:
    log(f"Writing {out_path.name}...")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()

    # ---- Summary tab ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NOR CAL Forward Fulfillment — Validation Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Snapshot: {SNAP_DATE.date()} | "
                "Plan source: 2026 Sales Plan by Item | "
                "History: 2023+24+25 actuals (smoothed)")

    # Annual Plan & Lift Totals (May-Dec)
    fwd_p = plan[plan["MonthNum"].isin(FWD_MONTHS)]
    fwd_l = plan_lifted[plan_lifted["MonthNum"].isin(FWD_MONTHS)]
    orig = fwd_p.groupby("Customer")["PlanQty"].sum()
    lift = fwd_l.groupby("Customer")["LiftedQty"].sum()

    ws["A4"] = "Annual Plan & Lift Totals (NOR CAL May-Dec)"
    ws["A4"].font = Font(bold=True)
    headers_a = ["Customer", "Original Plan QTY", "Lifted Plan QTY", "Lift Δ QTY", "Lift Δ %"]
    for j, h in enumerate(headers_a, start=1):
        ws.cell(row=5, column=j, value=h).font = Font(bold=True)
    rownum = 6
    cust_order = sorted(set(PLAN_CUSTOMERS))
    for c in cust_order:
        o = float(orig.get(c, 0.0))
        l = float(lift.get(c, 0.0))
        ws.cell(row=rownum, column=1, value=c)
        ws.cell(row=rownum, column=2, value=o)
        ws.cell(row=rownum, column=3, value=l)
        ws.cell(row=rownum, column=4, value=l - o)
        ws.cell(row=rownum, column=5, value=(l - o) / o if o else 0.0)
        rownum += 1
    ws.cell(row=rownum, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=rownum, column=2, value=float(orig.sum()))
    ws.cell(row=rownum, column=3, value=float(lift.sum()))
    ws.cell(row=rownum, column=4, value=float(lift.sum() - orig.sum()))
    rownum += 3

    # YE Miss by Customer
    ws.cell(row=rownum, column=1, value="YE Miss by Customer").font = Font(bold=True)
    rownum += 1
    headers_b = ["Customer", "Original YE Plan", "Lifted YE Plan", "YTD Actual",
                 "Forward Fulfillable", "YE Miss QTY", "YE Miss $"]
    for j, h in enumerate(headers_b, start=1):
        ws.cell(row=rownum, column=j, value=h).font = Font(bold=True)
    rownum += 1
    yem_g = ye_miss.groupby("Customer").agg({
        "Original YE Plan": "sum", "Lifted YE Plan": "sum", "YTD Actual": "sum",
        "Forward Fulfillable": "sum", "YE Miss QTY": "sum", "YE Miss $": "sum",
    }).reset_index()
    for c in cust_order:
        row = yem_g[yem_g["Customer"] == c]
        if row.empty:
            continue
        r = row.iloc[0]
        ws.cell(row=rownum, column=1, value=c)
        ws.cell(row=rownum, column=2, value=float(r["Original YE Plan"]))
        ws.cell(row=rownum, column=3, value=float(r["Lifted YE Plan"]))
        ws.cell(row=rownum, column=4, value=float(r["YTD Actual"]))
        ws.cell(row=rownum, column=5, value=float(r["Forward Fulfillable"]))
        ws.cell(row=rownum, column=6, value=float(r["YE Miss QTY"]))
        ws.cell(row=rownum, column=7, value=float(r["YE Miss $"]))
        rownum += 1

    # ---- BUXUS Detail tab ----
    ws2 = wb.create_sheet("BUXUS Detail")
    ws2["A1"] = "#001 BUXUS NOR CAL — Forward Fulfillment Walk"
    ws2["A1"].font = Font(bold=True, size=14)

    bux_ki = "#001 BUXUS"
    bux_walk = pool_walk[pool_walk["KI"] == bux_ki].copy().sort_values("Month")

    ws2["A3"] = "Pool Walk by Month"
    ws2["A3"].font = Font(bold=True)
    hdr = ["Month", "Inflow Grad", "Inflow SN/GN",
           "S1 BB Demand (Plan)", "S1 NBB Demand (Plan)",
           "S1 BB Filled", "S1 NBB Filled", "S1 BB Short", "S1 NBB Short",
           "S1 End All", "S1 End Restricted",
           "S2 Demand (Lift)", "S2 Filled", "S2 Short",
           "End All Pool (post-S2)", "End Restricted (post-S2)", "End Combined"]
    for j, h in enumerate(hdr, start=1):
        ws2.cell(row=4, column=j, value=h).font = Font(bold=True)
    r = 5
    for _, row in bux_walk.iterrows():
        ws2.cell(row=r, column=1, value=MONTH_NUM_TO_NAME[int(row["Month"])])
        ws2.cell(row=r, column=2, value=float(row["InflowGrad"]))
        ws2.cell(row=r, column=3, value=float(row["InflowSNGN"]))
        ws2.cell(row=r, column=4, value=float(row["Stage1_BB_Demand"]))
        ws2.cell(row=r, column=5, value=float(row["Stage1_NBB_Demand"]))
        ws2.cell(row=r, column=6, value=float(row["Stage1_BB_Filled"]))
        ws2.cell(row=r, column=7, value=float(row["Stage1_NBB_Filled"]))
        ws2.cell(row=r, column=8, value=float(row["Stage1_BB_Short"]))
        ws2.cell(row=r, column=9, value=float(row["Stage1_NBB_Short"]))
        ws2.cell(row=r, column=10, value=float(row["Stage1_End_All"]))
        ws2.cell(row=r, column=11, value=float(row["Stage1_End_Restricted"]))
        ws2.cell(row=r, column=12, value=float(row["Stage2_Demand_Total"]))
        ws2.cell(row=r, column=13, value=float(row["Stage2_Filled"]))
        ws2.cell(row=r, column=14, value=float(row["Stage2_Short"]))
        ws2.cell(row=r, column=15, value=float(row["End_All"]))
        ws2.cell(row=r, column=16, value=float(row["End_Restricted"]))
        ws2.cell(row=r, column=17, value=float(row["End_Combined"]))
        r += 1

    # Per-Customer × Month Detail
    r += 2
    ws2.cell(row=r, column=1, value="Per-Customer × Month Detail").font = Font(bold=True)
    r += 1
    hdr2 = ["Customer", "Month", "Item Num", "Item Description",
            "Plan QTY (orig)", "Plan QTY (lifted)",
            "Plan Filled (S1)", "Plan Short (S1)",
            "Lift Filled (S2)", "Lift Short (S2)"]
    for j, h in enumerate(hdr2, start=1):
        ws2.cell(row=r, column=j, value=h).font = Font(bold=True)
    r += 1

    bux_alloc = cust_alloc[cust_alloc["KI"] == bux_ki].copy()
    bux_alloc = bux_alloc.merge(items[["Item Num", "Item Desc"]].drop_duplicates(),
                                 on="Item Num", how="left")
    cust_order_bux = sorted(set(PLAN_CUSTOMERS))
    for c in cust_order_bux:
        sub = bux_alloc[bux_alloc["Customer"] == c].sort_values(["MonthNum", "Item Num"])
        for _, row in sub.iterrows():
            ws2.cell(row=r, column=1, value=row["Customer"])
            ws2.cell(row=r, column=2, value=MONTH_NUM_TO_NAME[int(row["MonthNum"])])
            ws2.cell(row=r, column=3, value=row["Item Num"])
            ws2.cell(row=r, column=4, value=row.get("Item Desc"))
            ws2.cell(row=r, column=5, value=float(row["PlanQty"]))
            ws2.cell(row=r, column=6, value=float(row["LiftedQty"]))
            ws2.cell(row=r, column=7, value=float(row["PlanFilled"]))
            ws2.cell(row=r, column=8, value=float(row["PlanShort"]))
            ws2.cell(row=r, column=9, value=float(row["LiftFilled"]))
            ws2.cell(row=r, column=10, value=float(row["LiftShort"]))
            r += 1

    wb.save(out_path)
    log(f"  wrote {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--rebuild-cache", action="store_true",
                    help="Force rebuild of per-year history parquet caches")
    ap.add_argument("--out", default=None,
                    help="Output xlsx filename (default: NOR CAL Forward Reproduction <MMDDYY>.xlsx)")
    args = ap.parse_args(argv)

    out_name = args.out or f"NOR CAL Forward Reproduction {SNAP_DATE.strftime('%m%d%y')}.xlsx"
    out_path = PROJECT_DIR / out_name

    plan = load_plan()
    v158_region = load_v158_region_dataset()
    dmnd = load_v158_demand_data()
    inv = load_inventory()
    ytd = load_ytd_actuals()
    hist = load_history(rebuild=args.rebuild_cache)
    hd_xref = load_hd_xref()
    lowes_xref = load_lowes_xref()

    items = build_item_universe(plan, inv, dmnd)
    items = attach_ki_grouping(items, v158_region)

    # Merge KI onto plan
    item_ki = items[["Item Num", "KI"]].drop_duplicates()
    plan = plan.merge(item_ki, on="Item Num", how="left")

    pools = build_pools(inv, items)
    lift_groups = build_lift_groups(items, hist, plan, hd_xref, lowes_xref)
    smoothed = compute_smoothed_history(hist, lift_groups)
    plan_lifted = apply_lift(plan, smoothed, lift_groups)
    offplan_demand = build_offplan_stage2_demand(smoothed)

    pool_walk, cust_alloc, offplan_alloc = run_walk(plan_lifted, pools, offplan_demand)
    prices = build_price_cascade(items, hist, ytd, lift_groups)
    ye_miss = compute_ye_miss(plan, plan_lifted, ytd, cust_alloc, items, prices)

    # Sanity report
    log("=== Sanity report ===")
    fwd_orig = plan[plan["MonthNum"].isin(FWD_MONTHS)].groupby("Customer")["PlanQty"].sum()
    fwd_lift = plan_lifted[plan_lifted["MonthNum"].isin(FWD_MONTHS)].groupby("Customer")["LiftedQty"].sum()
    for c in ["HD", "Lowes", "Midwest", "Walmart", "West Coast"]:
        log(f"  {c}: orig={fwd_orig.get(c,0):,.2f}  lifted={fwd_lift.get(c,0):,.2f}  Δ={fwd_lift.get(c,0)-fwd_orig.get(c,0):,.2f}")

    write_output(out_path, plan, plan_lifted, ytd, cust_alloc, pool_walk, ye_miss, items)
    log("DONE")


if __name__ == "__main__":
    main()
