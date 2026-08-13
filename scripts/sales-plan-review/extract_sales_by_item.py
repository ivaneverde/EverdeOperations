#!/usr/bin/env python3
"""
Aggregate Sales by Item workbooks → compact JSON for Claude / portal lookup.

Grain: 445 Year × Tree × Demand Channel × Renamed Rep × Bill To
  - item + channel + rep questions (Justin)
  - customer / account lookups (Meredith) via Bill To Name

Default inputs:
  2024/2025  Shared\\Sales Data\\{year} Sales by Item.xlsx (Year End archive fallback)
  2026       newest *Sales by Item*.xlsx in Sales Plan Review\\WeeklyDrop

Outputs (repo public/):
  sales_by_item_meta.json
  sales_by_item_rows.json.gz   (JSON array of row arrays)

Usage:
  python extract_sales_by_item.py
  python extract_sales_by_item.py --out-dir public
"""
from __future__ import annotations

import argparse
import gzip
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_SHARED = Path(
    r"\\192.168.190.10\Claude Sandbox\JS Files\Shared\Sales Data"
)
DEFAULT_DROP = Path(
    r"\\192.168.190.10\Claude Sandbox\DataDrops\Sales Plan Review\WeeklyDrop"
)
DEFAULT_YEAR_END = Path(
    r"\\VRD-AWSECS\Everde Central Share\Planning & Reporting\Data & Reports\Posted Data\Sales by Item\Year End Sales by Items"
)

# Bump when grain/columns change so share caches are not reused.
CACHE_VERSION = "v2"

COLUMNS = [
    "year",
    "tree",
    "description",
    "common_name",
    "container",
    "demand_channel",
    "rep",
    "bill_to",
    "qty",
    "revenue",
    "lines",
]


def newest_weekly(drop: Path) -> Path | None:
    if not drop.is_dir():
        return None
    hits = [
        p
        for p in drop.glob("*Sales by Item*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not hits:
        return None
    return max(hits, key=lambda p: p.stat().st_mtime)


def hist_file(shared: Path, year: int, year_end: Path | None = None) -> Path | None:
    roots: list[Path] = []
    if shared.is_dir():
        roots.append(shared)
    if year_end is not None and year_end.is_dir():
        roots.append(year_end)
    for root in roots:
        exact = root / f"{year} Sales by Item.xlsx"
        if exact.is_file():
            return exact
        hits = [
            p
            for p in root.glob(f"{year} Sales by Item*.xlsx")
            if p.is_file() and not p.name.startswith("~$")
        ]
        if hits:
            return max(hits, key=lambda p: p.stat().st_mtime)
    return None


def cache_path(cache_dir: Path, src: Path) -> Path:
    st = src.stat()
    safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in src.stem)
    return cache_dir / f"sbi_{CACHE_VERSION}_{safe}_{st.st_size}_{int(st.st_mtime)}.json"


def _num(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return 0.0
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _year(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        y = int(v)
        return y if 2000 <= y <= 2100 else None
    s = str(v).strip()
    if len(s) >= 4 and s[:4].isdigit():
        y = int(s[:4])
        return y if 2000 <= y <= 2100 else None
    return None


def _header_index(header: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    lower = {str(c).strip().lower(): i for i, c in enumerate(header)}
    aliases = {
        "Tree": ["tree"],
        "Description": ["description"],
        "Common Name": ["common name"],
        "Container Code": ["container code"],
        "Demand Channel": ["demand channel"],
        "Rep": ["rep"],
        "Renamed Rep": ["renamed rep"],
        "Bill To Name": ["bill to name", "bill to", "billto"],
        "Qty Inv SUM": ["qty inv sum", "qty inv", "qty"],
        "Revenue Amt Sum": ["revenue amt sum", "revenue"],
        "445 Year": ["445 year"],
        "Gl Yr": ["gl yr", "gl year"],
    }
    for canon, names in aliases.items():
        for n in names:
            if n in lower:
                idx[canon] = lower[n]
                break
    missing = [k for k in ("Tree", "Demand Channel") if k not in idx]
    if missing:
        raise ValueError(f"Sales by Item missing columns: {missing}")
    return idx


def scan_workbook(src: Path) -> dict[str, dict[str, Any]]:
    print(f"  scanning {src.name} ({src.stat().st_size / 1e6:.0f} MB)...", flush=True)
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    agg: dict[str, dict[str, Any]] = {}
    idx: dict[str, int] | None = None
    rows = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip() if c is not None else "" for c in row]
            idx = _header_index(header)
            continue
        assert idx is not None
        rows += 1
        tree = str(row[idx["Tree"]] or "").strip()
        if not tree:
            continue
        year = _year(row[idx["445 Year"]]) if "445 Year" in idx else None
        if year is None and "Gl Yr" in idx:
            year = _year(row[idx["Gl Yr"]])
        if year is None:
            continue
        ch = str(row[idx["Demand Channel"]] or "").strip()
        renamed = ""
        if "Renamed Rep" in idx:
            renamed = str(row[idx["Renamed Rep"]] or "").strip()
        raw_rep = str(row[idx["Rep"]] or "").strip() if "Rep" in idx else ""
        rep = renamed or raw_rep or "(unassigned)"
        bill_to = ""
        if "Bill To Name" in idx:
            bill_to = str(row[idx["Bill To Name"]] or "").strip()
        if not bill_to or bill_to.upper() == "NULL":
            bill_to = "(unknown)"
        key = f"{year}\t{tree}\t{ch}\t{rep}\t{bill_to}"
        rec = agg.get(key)
        if rec is None:
            desc = str(row[idx["Description"]] or "").strip() if "Description" in idx else ""
            common = str(row[idx["Common Name"]] or "").strip() if "Common Name" in idx else ""
            container = (
                str(row[idx["Container Code"]] or "").strip() if "Container Code" in idx else ""
            )
            rec = {
                "year": year,
                "tree": tree,
                "description": desc,
                "common_name": common,
                "container": container,
                "demand_channel": ch,
                "rep": rep,
                "bill_to": bill_to,
                "qty": 0.0,
                "revenue": 0.0,
                "lines": 0,
            }
            agg[key] = rec
        rec["qty"] += _num(row[idx["Qty Inv SUM"]]) if "Qty Inv SUM" in idx else 0.0
        rec["revenue"] += _num(row[idx["Revenue Amt Sum"]]) if "Revenue Amt Sum" in idx else 0.0
        rec["lines"] += 1
        if rows % 200000 == 0:
            print(f"    {rows:,} source rows, {len(agg):,} keys", flush=True)
    wb.close()
    print(f"    done {rows:,} source rows -> {len(agg):,} keys", flush=True)
    return agg


def load_or_scan(src: Path, cache_dir: Path) -> tuple[dict[str, dict[str, Any]], int]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = cache_path(cache_dir, src)
    if cached.is_file():
        print(f"  cache hit {cached.name}", flush=True)
        data = json.loads(cached.read_text(encoding="utf-8"))
        keys = {row["key"]: row["rec"] for row in data["keys"]}
        return keys, int(data.get("sourceRows", 0))
    agg = scan_workbook(src)
    payload = {
        "version": CACHE_VERSION,
        "source": src.name,
        "sourceRows": sum(r["lines"] for r in agg.values()),
        "keys": [{"key": k, "rec": v} for k, v in agg.items()],
    }
    cached.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    print(f"  wrote cache {cached.name}", flush=True)
    return agg, payload["sourceRows"]


def round_num(v: float) -> float | int:
    r = round(float(v), 2)
    if r == int(r):
        return int(r)
    return r


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--shared-sales", default=str(DEFAULT_SHARED))
    ap.add_argument("--weekly-drop", default=str(DEFAULT_DROP))
    ap.add_argument("--year-end", default=str(DEFAULT_YEAR_END))
    ap.add_argument("--out-dir", default="")
    ap.add_argument(
        "--years",
        default="2024,2025,2026",
        help="Comma years. 2026 uses WeeklyDrop; earlier years use Shared / Year End.",
    )
    args = ap.parse_args()

    repo = Path(__file__).resolve().parents[2]
    out_dir = Path(args.out_dir) if args.out_dir else repo / "public"
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = Path(__file__).resolve().parent / "cache"

    shared = Path(args.shared_sales)
    drop = Path(args.weekly_drop)
    year_end = Path(args.year_end) if str(args.year_end).strip() else DEFAULT_YEAR_END
    years = [int(y.strip()) for y in args.years.split(",") if y.strip()]

    sources: list[dict[str, Any]] = []
    merged: dict[str, dict[str, Any]] = {}
    source_rows = 0

    for year in years:
        if year >= 2026:
            src = newest_weekly(drop)
            if src is None:
                print(f"No {year} Sales by Item in WeeklyDrop: {drop}", file=sys.stderr)
                continue
        else:
            src = hist_file(shared, year, year_end)
            if src is None:
                print(
                    f"No {year} Sales by Item in {shared} or {year_end}",
                    file=sys.stderr,
                )
                continue
        print(f"Year {year}: {src}", flush=True)
        agg, n = load_or_scan(src, cache_dir)
        source_rows += n
        sources.append(
            {
                "year": year,
                "file": src.name,
                "path": str(src),
                "bytes": src.stat().st_size,
                "mtime": datetime.fromtimestamp(src.stat().st_mtime, tz=timezone.utc).isoformat(),
                "keys": len(agg),
                "sourceRows": n,
            }
        )
        for k, rec in agg.items():
            prev = merged.get(k)
            if prev is None:
                merged[k] = dict(rec)
            else:
                prev["qty"] += rec["qty"]
                prev["revenue"] += rec["revenue"]
                prev["lines"] += rec["lines"]

    if not merged:
        print("No Sales by Item rows extracted.", file=sys.stderr)
        return 1

    rows = []
    channels: set[str] = set()
    year_set: set[int] = set()
    bill_tos: set[str] = set()
    for rec in merged.values():
        year_set.add(int(rec["year"]))
        if rec["demand_channel"]:
            channels.add(rec["demand_channel"])
        if rec.get("bill_to"):
            bill_tos.add(str(rec["bill_to"]))
        rows.append(
            [
                int(rec["year"]),
                rec["tree"],
                rec["description"],
                rec["common_name"],
                rec["container"],
                rec["demand_channel"],
                rec["rep"],
                rec.get("bill_to") or "(unknown)",
                round_num(rec["qty"]),
                round_num(rec["revenue"]),
                int(rec["lines"]),
            ]
        )
    rows.sort(key=lambda r: (r[0], str(r[7]), str(r[1]), str(r[5]), str(r[6])))

    meta = {
        "asOf": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "grain": "year x tree x demand_channel x renamed_rep x bill_to",
        "columns": COLUMNS,
        "rowCount": len(rows),
        "sourceRowCount": source_rows,
        "years": sorted(year_set),
        "channels": sorted(channels),
        "billToCount": len(bill_tos),
        "sources": sources,
        "note": (
            "bill_to = Bill To Name (customer/account). "
            "Rep = Renamed Rep else Rep. "
            "Fast Growing Trees may appear as Demand Channel and/or Bill To. "
            "West Coast LSC = WEST COAST NORTH + WEST COAST SOUTH. "
            "Use get_sales_by_item focus=query with q= customer and/or item and/or rep and/or year."
        ),
    }

    meta_path = out_dir / "sales_by_item_meta.json"
    rows_path = out_dir / "sales_by_item_rows.json.gz"
    meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    with gzip.open(rows_path, "wt", encoding="utf-8") as f:
        json.dump(rows, f, separators=(",", ":"))

    print(f"Wrote {meta_path} rows={len(rows):,} sourceRows={source_rows:,} billTos={len(bill_tos):,}")
    print(f"Wrote {rows_path} ({rows_path.stat().st_size / 1e6:.1f} MB gzip)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
